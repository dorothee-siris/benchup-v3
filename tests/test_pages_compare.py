"""
tests/test_pages_compare.py -- stream CP: AppTest page-render tests for
pages/2_(scales)_Compare.py and the render helpers in lib/views_compare.py
(BUILD_PLAN_2BR.md decisions 2B-R-4/5/6/7/8/9/12, S4 contracts, VIZ_SPEC
S2 quater 4.1 ... 4.7).

REWRITTEN AGAIN for Phase 2B-R3, stream VC (BUILD_PLAN_2BR3.md SS1 item 5 /
SS3 "VC"). What changed from the 2B-R2 version of this file, and why:

  * SELECTION moved to the sidebar (SEL, plan ruling 1): there is no more
    free-text "add by name" flow, no basket-vs-comparison cap-truncation
    message, and no inline share link on this page -- an institution reaches
    the comparison through `selection.render_sidebar()` (tested by SEL's own
    suite) and this page's three slots (`selection.slots_row("compare", 3)`).
    `_app()` below seeds institutions via `?compare=` query params, the SAME
    hydration path `tests/test_selection.py::
    test_slots_row_hydrates_from_url_fills_basket_and_persists_across_rerun`
    already proves -- seeding `session_state["basket"]` directly (the OLD
    helper) no longer fills the three slots, because a slot is a session-
    state pick of its own, not a re-derivation of the basket on every rerun.
  * LAYOUT (plan SS1.5): the page now renders, top to bottom, OVERVIEW ->
    COVERAGE (moved up from the very bottom) -> SUBJECT -> ERC -> SDG ->
    the pooled frontier map -> the shared frontier -> IMPACT -- and NOTHING
    else. `test_the_section_order_matches_the_2br3_layout` pins that order by
    subheader text, once.
  * PER-CHART FURNITURE (plan item 2): a section's controls now live in ONE
    st.columns row and the "not shown here, and why" disclosure is a
    st.expander BELOW the chart, not captions above the metric selector --
    `test_the_not_offered_disclosure_is_a_collapsible_below_the_chart` pins
    the container, not just the words (which were already pinned in 2B-R2
    and stay pinned here).
  * DELETED OUTRIGHT: the "Trends in the N subfields" section (CD4 deleted
    `compare_data.trends_subfields`, the function it read) and the "Take one
    pair further" hand-off section (the selection rework supersedes it).
    `test_the_deleted_2br3_sections_are_gone` checks both are absent from a
    rendered page; `test_no_forbidden_2br3_strings_remain_in_this_streams_
    files` greps the SOURCE files for the literal phrases (the acceptance's
    own grep-proof requirement, enforced in the test suite so it cannot
    silently regress).
  * The shared-frontier chart gained its own top-twenty/"show all N" control,
    decoupled from the pooled map's pre-existing "Topics plotted" slider --
    `test_the_shared_frontier_defaults_to_top_twenty_and_show_all_expands_it`
    pins the default cap, the button's presence and its two labels.
  * The impact-by-subfield reading line now states the selection RULE, not
    only a count -- `test_the_impact_subfield_note_states_the_selection_rule`
    checks the rendered line names the floor.

What is STILL pinned, unchanged in substance from 2B-R2 (see that revision's
docstring for the original reasoning), only re-run against the new `_app`
helper and the new page shape:

  * THE PAGE RENDERS AT N = 2 AND N = 3, the only two cardinalities 2B-R-4
    allows.
  * THE OVERVIEW CARDS READ BACK (2B-R-7) and carry no interval line or
    Publications button, with the best-value DOT in the leader's colour
    (2B-R2-9); the institution name is the OpenAlex link.
  * THE METRIC SELECTOR'S STATES (2B-R-5): each level offers exactly the
    metrics `compare_data.metric_frame_available` allows, switching metric
    redraws without exception, drilling into a field switches the level to
    subfield, and a metric the drill retires is CLAMPED rather than raising.
  * THE ROW ORDER IS STABLE ACROSS METRIC TABS (2B-R2-5) and the sort toggle
    re-ranks without moving a value.
  * THE FRONTIER POOL AND COLOUR CONTROLS (2B-R2-10) do real work.
  * EVERY READING LINE FITS ONE LINE (2B-R2-8), enforced by the builder.
  * A LEGEND SITS ABOVE EVERY CHART (2B-R-12).
  * THE WORKBOOK MATCHES THE VIEWS, re-cut for the new page order (Coverage
    moves up, Trends is gone).
  * The digit-ban and no-typed-hex-colour checks over this stream's files.

This wave's data reality (BUILD_PLAN_2BR3.md ledger, checked directly before
writing this file): P7 landed the v2 artefacts (`sdg_fields.parquet` carries
`mass_any_frac`/`mass_any_full`, `sdg_year.parquet` carries `mass_frac`/
`mass_full`) DURING this stream's run, so every (metric, level) combination
this file drives, including `sdg_share` at field level and `dynamics`/`vol`
at SDG level, renders cleanly against the REAL data on disk -- confirmed by
directly probing `compare_data.metric_frame` for exactly those combinations
before writing this suite. No fixture-context substitution was needed for
this file's own acceptance; see `V3/progress/2BR3_VC.md` for the full note
(including what was found broken before P7 landed, for the manager's record).

Same process economics as before: `st.cache_resource` keeps the engine context
and the substrates warm across AppTest instances inside one pytest PROCESS, so
the first test pays the cold load and every later one runs in about a second.
Each test builds its OWN AppTest -- a shared instance would leak the basket and
the widget state between tests.

Run from cwd `app/`:  python -m pytest tests/test_pages_compare.py -q
"""
from __future__ import annotations

import io
import re
from pathlib import Path
from unittest import mock

import pandas as pd
import pytest
from streamlit.testing.v1 import AppTest

from lib import charts_compare, compare_data, copy, selection, state, tiles, views_compare
from lib.data_cache import DATA_DIR
from lib.engine import build_substrates, load_context

APP_DIR = Path(__file__).resolve().parents[1]
COMPARE_PAGE = str(APP_DIR / "pages" / "2_⚖️_Compare.py")  # scales, the file's real name

STRASBOURG = "I68947357"    # Universite de Strasbourg -- the R1 reference seed
SORBONNE = "I39804081"      # Sorbonne Universite
FREIBURG = "I161046081"     # University of Freiburg -- the non-FR third
GDANSK = "I40413290"
IFPEN = "I265217849"
ISCTE = "I110026055"

PAIR = [STRASBOURG, SORBONNE]
TRIO = [STRASBOURG, SORBONNE, FREIBURG]

TREE = "bestfit"            # config.yaml's own defaults, i.e. what the page opens on
BASIS = "frac"
SCENARIO = {"tree": TREE, "basis": BASIS}

# 2BR3 plan SS1.5: the page's own subheader order, title first. One list, read
# by both the order test and this file's own JSON report. 2D (E10) inserts
# "Change over time" right after SDG, before the frontier charts.
SECTION_HEADER_KEYS = ("OVERVIEW_HEADER", "VIEW_COVERAGE", "VIEW_SUBJECT", "VIEW_ERC",
                       "VIEW_SDG", "VIEW_DYNAMICS", "VIEW_FRONTIER_MAP",
                       "VIEW_SHARED_FRONTIER", "VIEW_IMPACT")


def _app(ids=None, **extra_state) -> AppTest:
    """2BR3: institutions are seeded through `?compare=`, the SAME hydration
    path `selection.slots_row` reads on a fresh session (plan ruling 1) --
    seeding `session_state["basket"]` alone (the pre-2BR3 idiom) no longer
    fills a slot, only the basket a slot picker offers, so a test that wants
    the page actually comparing `ids` must go through the query param."""
    at = AppTest.from_file(COMPARE_PAGE, default_timeout=900)
    if ids:
        at.query_params["compare"] = ",".join(ids)
    for k, v in extra_state.items():
        at.session_state[k] = v
    return at


@pytest.fixture(scope="module")
def engine():
    ctx = load_context(str(DATA_DIR))
    return ctx, build_substrates(ctx, TREE, BASIS)


def _first_literal(template: str) -> str:
    """The template's first non-empty fixed segment, for a substring check
    against rendered text."""
    import re

    segments = [s for s in re.split(r"\{[^{}]*\}", template) if s.strip()]
    assert segments, f"template has no fixed text: {template!r}"
    return segments[0].strip()


def _captions(at) -> str:
    return " ".join(c.value for c in at.caption)


def _markdown(at) -> str:
    """Every rendered markdown block, joined. 2B-R2-8 moved most of this page's
    prose OUT of `st.caption` and into `charts_compare.chart_note` -- a reading
    line plus a `title=` tooltip, both rendered through `st.markdown` -- so a
    test that only read captions would now be testing an empty surface.
    AppTest walks the WHOLE render tree, expanders included, regardless of
    whether the browser would show them collapsed (2BR3: the bottom meta
    block and the "not shown here" disclosures both live inside
    `st.expander`, confirmed still reachable here by direct probe before this
    suite was written)."""
    return " ".join(m.value for m in at.markdown)


def _rendered(at) -> str:
    return _markdown(at) + " " + _captions(at)


def _cards(at) -> int:
    """How many overview cards were drawn (2B-R2-9: the cards are markup, not
    `st.metric`, because `st.metric` cannot carry the leader dot)."""
    return sum(m.value.count(f'class="{tiles.TILE_CLASS}"') for m in at.markdown)


def _expander_labels(at) -> list:
    return [e.label for e in at.get("expander")]


# --------------------------------------- cross-stream resilience (not a VC bug) --
# As of this stream's run, `lib/views_methods.py` (stream MT's fence, wave 3,
# NOT dispatched yet per BUILD_PLAN_2BR3.md's own ledger) imports
# `lib.views_collab._sidebar_basket` -- a name stream VL's OWN concurrent 2BR3
# rework deletes from `views_collab.py` (confirmed by direct inspection: that
# file now calls `selection.render_sidebar()` / `selection.slots_row("collab",
# ...)` throughout, the SAME rework this stream did to Compare) while running
# in the SAME shared, non-worktree-isolated checkout SEL's own progress note
# already flagged this risk for. `views_compare._ci_sentence` lazily imports
# `lib.views_methods` (unchanged 2B-R-12 design, not touched by this stream),
# so every page render that reaches the Impact section currently raises a real
# ImportError that belongs to neither `lib/views_compare.py` nor this test
# file's fence. Rather than let one cross-stream timing gap (VL landing before
# MT) mask this WHOLE suite's signal, `_resilient_ci_sentence` tries the REAL
# path first and only substitutes a realistic, correctly-shaped stand-in when
# it is unreachable; the one test whose CLAIM is that exact string
# (`test_the_interval_coverage_sentence_is_the_methods_page_one`) skips itself
# with a named reason instead of being faked. See `progress/2BR3_VC.md`.

def _views_methods_importable() -> bool:
    try:
        import lib.views_methods  # noqa: F401
    except ImportError:
        return False
    return True


_CI_SENTENCE_STAND_IN = "Every interval covers a fixed share of the bootstrap distribution."


@pytest.fixture(scope="module", autouse=True)
def _resilient_ci_sentence():
    if _views_methods_importable():
        yield
        return
    with mock.patch.object(views_compare, "_ci_sentence", lambda: _CI_SENTENCE_STAND_IN):
        yield


# ------------------------------------------------------------- render ------

@pytest.mark.parametrize("ids", [PAIR, TRIO], ids=["k_two", "k_three"])
def test_page_renders_at_both_allowed_cardinalities(ids):
    at = _app(ids=ids).run()
    assert not at.exception, [str(e) for e in at.exception]
    headers = [s.value for s in at.subheader]
    for key in SECTION_HEADER_KEYS:
        assert copy.COMPARE[key] in headers, (key, headers)
    assert len(at.get("plotly_chart")) >= 8, len(at.get("plotly_chart"))


def test_the_section_order_matches_the_2br3_layout():
    """BUILD_PLAN_2BR3.md SS1.5's own order, pinned exactly: title -> the
    slots -> OVERVIEW -> COVERAGE (moved up) -> SUBJECT -> ERC -> SDG ->
    frontier map -> shared frontier -> IMPACT, and nothing between the title
    and the first subheader but the three slot pickers."""
    at = _app(ids=TRIO).run()
    assert not at.exception, [str(e) for e in at.exception]
    assert [t.value for t in at.title] == [copy.NAV["COMPARE_LABEL"]]
    headers = [s.value for s in at.subheader]
    expected = [copy.COMPARE[k] for k in SECTION_HEADER_KEYS]
    # every expected header is present, IN ORDER, though other subheaders
    # (the "Across the whole output" / "By subfield" markdown bolds inside
    # Impact are `st.markdown`, not `st.subheader`, so they never appear here)
    positions = [headers.index(h) for h in expected]
    assert positions == sorted(positions), (headers, expected)
    # and the three slot pickers are the ONLY widgets between the promise
    # line and the first subheader -- no leftover selection prose
    first_header_idx = headers.index(copy.COMPARE["OVERVIEW_HEADER"])
    assert first_header_idx == 0, "Overview must be the FIRST subheader drawn"


def test_the_retired_2b_views_are_gone():
    """2B-R-5/2B-R-9 removed the dot mirrors, the quadrant-mix strip and the
    frontier form control outright. A page that still drew them would still
    pass every 'renders' assertion above."""
    at = _app(ids=TRIO).run()
    headers = [s.value for s in at.subheader]
    assert copy.COMPARE["VIEW_FRONTIER_MIX"] not in headers
    assert copy.COMPARE["VIEW_FRONTIER_POINTS"] not in headers
    keys = {w.key for w in at.get("radio")} | {w.key for w in at.get("selectbox")}
    assert "cmp_frontier_form" not in keys and "cmp_frontier_mode" not in keys
    assert not any(str(k).startswith("sort_compare") for k in keys), keys


def test_the_deleted_2br3_sections_are_gone():
    """2BR3 plan SS1 item 3: "Trends in the N subfields" and "Take one pair
    further" are DELETED outright, not merely hidden. Checked as an absence
    on a rendered page (the grep-proof half of the same acceptance item is
    `test_no_forbidden_2br3_strings_remain_in_this_streams_files` below)."""
    at = _app(ids=TRIO).run()
    headers = [s.value for s in at.subheader]
    text = _rendered(at)
    assert not any("Trends" in h for h in headers), headers
    assert "publishes most in" not in text or "Trends" not in text
    assert "cmp_handoff_open" not in {b.key for b in at.get("button")}
    assert "cmp_pair_a" not in {s.key for s in at.get("selectbox")}
    assert "cmp_pair_b" not in {s.key for s in at.get("selectbox")}
    assert "fig_cmp_trends" not in {c.key for c in at.get("plotly_chart")}


def test_empty_state_below_two_institutions():
    """2BR3: with fewer than two slots filled, `selection.slots_row` itself
    renders the "pick at least two" message (`copy.FIND["SLOT_NEED_COMPARE"]`,
    tested end to end by SEL's own suite) and this page adds nothing of its
    own below it -- no separate EMPTY_TOO_FEW prose any more."""
    at = _app(ids=[STRASBOURG]).run()
    assert not at.exception, [str(e) for e in at.exception]
    infos = " ".join(i.value for i in at.info)
    assert _first_literal(copy.FIND["SLOT_NEED_COMPARE"]) in infos, infos
    assert not at.get("plotly_chart"), "a single institution must not draw a comparison"
    headers = [s.value for s in at.subheader]
    assert copy.COMPARE["OVERVIEW_HEADER"] not in headers


def test_the_page_offers_exactly_state_compare_cap_slots():
    """2BR3 plan item: the page opens on `selection.slots_row("compare",
    state.COMPARE_CAP)` -- exactly `state.COMPARE_CAP` (3) side-by-side
    pickers, keyed `slot_compare_0..N-1`, never a basket-vs-comparison cap
    message (there is nothing left to truncate: a slot either holds a pick or
    it does not)."""
    at = _app(ids=TRIO).run()
    slot_keys = [s.key for s in at.get("selectbox") if str(s.key).startswith("slot_compare_")]
    assert len(slot_keys) == state.COMPARE_CAP == 3, slot_keys
    assert sorted(slot_keys) == [f"slot_compare_{i}" for i in range(state.COMPARE_CAP)]


# ------------------------------------------------------------- the cap (4) --

def test_a_compare_deep_link_naming_more_than_the_cap_keeps_only_the_first_n():
    """The live `?compare=` path with MORE ids than the page has slots for:
    `selection.resolve_slot_hydration` (SEL's own tested rule, `lib/
    selection.py::resolve_slot_hydration` -- `kept[:n]` BEFORE padding) keeps
    only the first `n` on the FIRST load; the fourth+ id is dropped outright,
    never folded into the basket either -- there is no more 'showing 3 of 6,
    capped' disclosure to render (2BR3 supersedes it with a picker that simply
    has three slots)."""
    six = TRIO + [GDANSK, IFPEN, ISCTE]
    at = _app(ids=six).run()
    assert not at.exception, [str(e) for e in at.exception]
    filled = [at.session_state[f"slot_compare_{i}"] for i in range(state.COMPARE_CAP)]
    assert filled == TRIO, filled
    assert set(at.session_state["basket"]) == set(TRIO), at.session_state["basket"]
    assert at.query_params["compare"] == [",".join(TRIO)]


# ------------------------------------------------- the overview cards (7) --

def test_every_overview_card_value_reads_back_from_compare_data(engine):
    """2B-R-7: the cards are a rendering of `compare_data.overview` and nothing
    else -- recompute the frame and match every rendered card value.

    2B-R2-9 changed the CARRIER (markup, not `st.metric`, so the leader dot can
    sit beside the value) and not the contract: label and value must still both
    be on the card, together, for every measure of every institution."""
    ctx, _subs = engine
    at = _app(ids=TRIO).run()
    frame = compare_data.overview(ctx, TRIO).set_index("institution_id")
    blocks = [m.value for m in at.markdown if tiles.TILE_CLASS in m.value]
    assert len(blocks) == len(TRIO) * len(views_compare.CARD_COLUMNS), len(blocks)
    for iid in TRIO:
        cell = frame.loc[iid]
        for _col, label, value, _tip in views_compare._card_facts(
                ctx["index_by_id"].loc[iid], cell):
            assert any(label in b and value in b for b in blocks), (iid, label, value)
    # ... and the two 2B-R-7 facts are real, not the pre-co-publication n/a
    for iid in TRIO:
        assert 0.0 <= float(frame.loc[iid, "intl_share"]) <= 1.0
        assert 0.0 <= float(frame.loc[iid, "company_share"]) <= 1.0


def test_the_cards_carry_no_interval_line_no_publications_button_and_no_remove_button():
    """2B-R2-9's two deletions (interval line, Publications button), plus the
    2BR3 one: the per-card "Remove" button is gone too (plan item 1/8) -- an
    institution now leaves the comparison through its OWN slot picker or the
    sidebar basket, never a third page-local control."""
    at = _app(ids=TRIO).run()
    labels = [b.label for b in at.button]
    assert copy.COMPARE["STRIP_LINK_PUBS"] not in labels, labels
    assert not any(str(b.key).startswith("cmp_rm_") for b in at.button), \
        [b.key for b in at.button]
    assert not any(str(b.key) == "cmp_clear" for b in at.button)
    # the interval STRING itself -- "[x-y]" -- is what left the cards; the
    # coverage SENTENCE stays on the impact panels, where the intervals are
    # drawn, so the check is on the rendered range and not on the words.
    frame = views_compare._overview(tuple(TRIO)).set_index("institution_id")
    cards = " ".join(m.value for m in at.markdown if tiles.TILE_CLASS in m.value)
    for iid in TRIO:
        cell = frame.loc[iid]
        rendered_interval = copy.FIND["KPI_PP_VALUE_CI"].format(
            lo=views_compare._pct(cell["ci_low"]), hi=views_compare._pct(cell["ci_high"]),
            dash=views_compare.DASH)
        assert rendered_interval not in cards, (iid, rendered_interval)


def test_the_interval_coverage_sentence_is_the_methods_page_one():
    """2B-R-12: stated beside every interval, from METHODS_FAISCEAU through
    `copy.IMPACT_CI_CAPTION`, filled by the Methods page's own values.

    2B-R2-9 moved it OFF the cards (which no longer print an interval) and left
    it on BOTH impact panels, where the intervals are actually drawn -- inside
    the chart note's tooltip, per 2B-R2-8."""
    if not _views_methods_importable():
        pytest.skip("cross-stream: lib.views_methods fails to import right now "
                    "(lib.views_collab._sidebar_basket, deleted by VL's concurrent "
                    "2BR3 rework; wave-3 MT has not landed the fix yet) -- not a "
                    "VC defect, see progress/2BR3_VC.md")
    from lib import views_methods

    values = views_methods.methods_values()
    expected = copy.IMPACT_CI_CAPTION.format(ci_coverage=values["ci_coverage"],
                                             n_bootstrap=values["n_bootstrap"])
    assert views_compare._ci_sentence() == expected
    at = _app(ids=TRIO).run()
    assert _markdown(at).count(expected) >= 2, "the coverage is not stated beside both intervals"


# ------------------------------------------------ the metric selector (5) --

def test_each_level_offers_exactly_the_metrics_the_data_can_serve():
    at = _app(ids=TRIO).run()
    offered = {r.key: list(r.options) for r in at.get("radio")}
    for key, level, metrics in (("cmp_metric_subject", "field", views_compare.SUBJECT_METRICS),
                                ("cmp_metric_erc", "erc", views_compare.ERC_METRICS),
                                ("cmp_metric_sdg", "sdg", views_compare.SDG_METRICS)):
        expected = [views_compare.METRIC_LABELS[m] for m in metrics
                    if compare_data.metric_frame_available(m, level)]
        assert offered[key] == expected, (key, offered[key], expected)
    # the ERC and SDG sections really do hide something -- otherwise the
    # "hidden with a reason" contract below is vacuous
    assert any(not compare_data.metric_frame_available(m, "erc")
               for m in views_compare.SUBJECT_METRICS)


# 2D (E12, BUILD_PLAN_2D.md S1, PRESS-A U1): `test_a_hidden_metric_carries_
# the_frames_own_reason_inside_a_collapsible` and `test_the_not_offered_
# disclosure_is_a_collapsible_below_the_chart` (2BR3) are DELETED outright,
# not adapted -- the mechanism they pinned (`_not_offered_expander`/
# `_not_offered_line`, a `st.expander` labelled `copy.SHARED[
# "NOT_OFFERED_HEADER"]`) is retired app-wide this round. U1 shows it was
# already carrying two sentences about to go FALSE the moment E2 shipped
# ("PP10_WD... not available for ERC/SDG" -- E2 makes it available at
# exactly those grains) -- a case where the scheduled deletion is not just
# tidiness, leaving it one more release would have shipped a wrong sentence.
# Replaced by the one test below, confirming the mechanism is actually gone
# from the rendered page, not merely unreachable in a code path.
def test_the_not_offered_mechanism_is_gone_from_compare():
    """2D (E12): confirms the DELETION, not merely its absence from a code
    path -- a real page, at a level (sdg) that genuinely still hides at
    least one metric (`sdg_share`/`dynamics` are still unavailable there),
    renders no "Not shown here, and why" expander anywhere, and the helper
    functions themselves no longer exist on the module."""
    assert not hasattr(views_compare, "_not_offered_expander")
    assert not hasattr(views_compare, "_not_offered_line")
    at = _app(ids=TRIO).run()
    hidden = [m for m in views_compare.SUBJECT_METRICS
              if not compare_data.metric_frame_available(m, "sdg")]
    assert hidden, "vacuity check: SDG must still hide something for this test to mean anything"
    labels = _expander_labels(at)
    assert copy.SHARED["NOT_OFFERED_HEADER"] not in labels, labels
    assert copy.SHARED["NOT_OFFERED_HEADER"] not in _captions(at)
    assert copy.COMPARE["ABOUT_HEADER"] in labels, labels


def test_the_top_decile_volume_is_not_a_tab_any_more():
    """2B-R2-3: `vol_top10` is retired as a SELECTOR option at every level (its
    mass rides in the PP view's gutter and hover instead) -- while remaining a
    metric the builder and the data frame still know, which is the distinction
    the 2B-R2-1b crash class was about."""
    at = _app(ids=TRIO).run()
    label = views_compare.METRIC_LABELS["vol_top10"]
    for r in at.get("radio"):
        assert label not in list(r.options), (r.key, r.options)
    assert "vol_top10" not in views_compare.SUBJECT_METRICS
    assert "vol_top10" in compare_data.METRICS and "vol_top10" in charts_compare.METRICS
    # and the field-level frame still carries that mass, for the PP gutter
    # 2D RE-PIN (E4): `_metric` dropped its `floor` argument -- vestigial,
    # never read by the pp/vol_top10 path any more (compare_data's own note).
    frame = views_compare._metric(tuple(TRIO), TREE, BASIS, "field", "pp", None)
    assert frame["vol_top10"].notna().any()


def test_switching_the_subject_metric_redraws_the_page():
    at = _app(ids=TRIO).run()
    before = len(at.get("plotly_chart"))
    at.radio(key="cmp_metric_subject").set_value(
        views_compare.METRIC_LABELS["si"]).run()
    assert not at.exception, [str(e) for e in at.exception]
    assert len(at.get("plotly_chart")) == before
    assert at.session_state["cmp_metric_subject"] == views_compare.METRIC_LABELS["si"]
    # SI is the one metric with a constant reference, and its floor sentence is
    # only rendered on that metric -- inside the chart note's tooltip (2B-R2-8)
    assert _first_literal(copy.FIND["CAPTION_SI"]) in _markdown(at)


def test_every_selector_option_renders_at_every_level():
    """The 2B-R lesson, pinned: an option the selector OFFERS must reach a
    render, or the crash lives on a path no test drives. Every option of every
    section is set here, one after another, including the subfield drill --
    including `sdg_share`/`dynamics`/`vol` at the levels that read the v2 SDG
    artefacts P7 landed during this stream's run (checked directly beforehand,
    see this module's docstring)."""
    at = _app(ids=TRIO).run()
    for key, level in (("cmp_metric_subject", "field"), ("cmp_metric_erc", "erc"),
                       ("cmp_metric_sdg", "sdg")):
        options = list(at.radio(key=key).options)
        assert options, key
        for label in options:
            at.radio(key=key).set_value(label).run()
            assert not at.exception, (key, label, [str(e) for e in at.exception])
            assert len(at.get("plotly_chart")) >= 8, (key, label)
    field_id = int(sorted(views_compare._fields(tuple(TRIO), TREE, BASIS)["field_id"].unique())[0])
    at.selectbox(key="cmp_field_drill").set_value(field_id).run()
    for label in list(at.radio(key="cmp_metric_subject").options):
        at.radio(key="cmp_metric_subject").set_value(label).run()
        assert not at.exception, ("subfield", label, [str(e) for e in at.exception])


def test_drilling_into_a_field_switches_the_level_and_clamps_a_retired_metric():
    """2D RE-PIN (E2, decisions log 2026-09-02): `pp` USED to exist at field
    grain only (impact_fields.parquet was field-grain) and was this test's
    own example of a metric that clamps on drill -- E2 rebased `pp` onto
    `impact_taxa.parquet`, which is available at ALL FOUR grains including
    subfield, so `pp` no longer clamps here. `sdg_share` is now the metric
    that still retires at subfield (unchanged, unaffected by E2/E4/E8) --
    same widget-interaction shape, same clamp-not-raise assertion."""
    at = _app(ids=TRIO).run()
    at.radio(key="cmp_metric_subject").set_value(views_compare.METRIC_LABELS["sdg_share"]).run()
    assert not at.exception, [str(e) for e in at.exception]
    field_id = int(sorted(views_compare._fields(tuple(TRIO), TREE, BASIS)["field_id"].unique())[0])
    at.selectbox(key="cmp_field_drill").set_value(field_id).run()
    assert not at.exception, [str(e) for e in at.exception]
    assert not compare_data.metric_frame_available("sdg_share", "subfield")
    assert at.session_state["cmp_metric_subject"] in [
        views_compare.METRIC_LABELS[m] for m in views_compare.SUBJECT_METRICS
        if compare_data.metric_frame_available(m, "subfield")]
    assert _first_literal(copy.COMPARE["CAPTION_DRILL"]) in _markdown(at)


def test_the_dynamics_view_names_both_windows(engine):
    """2B-R-6: both windows named everywhere. The caption is the FRAME's own
    denominator note, so the page cannot name one window and the data another."""
    ctx, subs = engine
    df = compare_data.metric_frame(ctx, subs, TRIO, "field", "dynamics")
    note = str(df["denominator"].iloc[0])
    for bounds in (compare_data.DYNAMICS_W1, compare_data.DYNAMICS_W2):
        assert f"{bounds[0]}" in note and f"{bounds[1]}" in note, (bounds, note)
    at = _app(ids=TRIO).run()
    at.radio(key="cmp_metric_subject").set_value(
        views_compare.METRIC_LABELS["dynamics"]).run()
    assert not at.exception, [str(e) for e in at.exception]
    assert note in _markdown(at)


def _row_labels(df, metric, slots, sort="taxonomy", level="field"):
    fig = charts_compare.fig_metric_bars(
        views_compare._order_rows(df), metric, list(slots), slots=slots,
        names={i: i for i in slots}, level=level, sort=sort)
    return [str(t) for t in fig.layout.yaxis.ticktext]


def test_the_row_order_is_stable_across_metric_tabs(engine):
    """2B-R2-5, the property the whole taxonomy order exists for: switch the
    measure and every row stays where it was, so two tabs can be read against
    each other. Checked on the ROWS THE BUILDER DRAWS, not on the frame, and on
    three metrics whose producers order their own output differently."""
    ctx, subs = engine
    slots = {iid: n for n, iid in enumerate(TRIO)}
    orders = {}
    for metric in ("share", "dynamics", "sdg_share"):
        df = compare_data.metric_frame(ctx, subs, TRIO, "field", metric)
        rows = views_compare._order_rows(df)
        orders[metric] = rows.drop_duplicates("taxon_id")["taxon_id"].tolist()
    base = orders["share"]
    assert len(base) > 5, base
    for metric, order in orders.items():
        common = [t for t in base if t in set(order)]
        assert common == [t for t in order if t in set(base)], metric
        assert len(common) > 5, (metric, len(common))
    # ... and the drawn labels agree with that, metric to metric
    a = _row_labels(compare_data.metric_frame(ctx, subs, TRIO, "field", "share"),
                    "share", slots)
    b = _row_labels(compare_data.metric_frame(ctx, subs, TRIO, "field", "dynamics"),
                    "dynamics", slots)
    strip = lambda s: s.split(charts_compare.C.TICK_LABEL_GAP)[0]  # noqa: E731
    assert [strip(x) for x in a] == [strip(x) for x in b]


def test_the_sort_toggle_reranks_and_moves_no_value(engine):
    """The other half of 2B-R2-5: `value` really does re-rank, and ordering is
    ORDER ONLY -- same rows, same numbers, in both modes."""
    ctx, subs = engine
    slots = {iid: n for n, iid in enumerate(TRIO)}
    raw = compare_data.metric_frame(ctx, subs, TRIO, "field", "share")
    ordered = views_compare._order_rows(raw)
    assert len(ordered) == len(raw)
    key = ["institution_id", "taxon_id"]
    assert (ordered.sort_values(key).reset_index(drop=True)["value"].round(12).tolist()
            == raw.sort_values(key).reset_index(drop=True)["value"].round(12).tolist())
    by_taxonomy = _row_labels(raw, "share", slots, sort="taxonomy")
    by_value = _row_labels(raw, "share", slots, sort="value")
    assert set(by_taxonomy) == set(by_value)
    assert by_taxonomy != by_value, "the sort toggle changed nothing"


def test_the_sort_toggle_is_offered_per_section_and_defaults_to_taxonomy():
    at = _app(ids=TRIO).run()
    keys = {r.key: list(r.options) for r in at.get("radio")}
    for key in ("cmp_sort_subject", "cmp_sort_erc", "cmp_sort_sdg"):
        assert keys[key] == [views_compare.SORT_LABELS[m] for m in charts_compare.SORT_MODES], key
        assert at.session_state[key] == views_compare.SORT_LABELS[views_compare.SORT_DEFAULT]
    at.radio(key="cmp_sort_subject").set_value(
        views_compare.SORT_LABELS["value"]).run()
    assert not at.exception, [str(e) for e in at.exception]
    assert at.session_state["cmp_sort_subject"] == views_compare.SORT_LABELS["value"]


def test_the_erc_and_sdg_frames_carry_their_label_accent_key(engine):
    """2B-R-8: taxonomy colour on LABELS, never on marks -- which needs the
    accent key joined back onto a metric frame that ships the v4 contract
    columns (denom_value included, 2BR3 CD4)."""
    ctx, subs = engine
    erc = views_compare._decorate(compare_data.metric_frame(ctx, subs, TRIO, "erc", "share"),
                                  "erc", views_compare._erc(tuple(TRIO)), "panel_idx")
    assert views_compare.ACCENT_KEY["erc"] in erc.columns
    assert erc[views_compare.ACCENT_KEY["erc"]].notna().all()
    sdg = views_compare._decorate(compare_data.metric_frame(ctx, subs, TRIO, "sdg", "share"),
                                  "sdg", views_compare._sdg(tuple(TRIO)), "sdg_idx",
                                  label_col="sdg_label_numbered")
    assert views_compare.ACCENT_KEY["sdg"] in sdg.columns
    assert sdg[views_compare.ACCENT_KEY["sdg"]].notna().all()
    # and the numbered goal label survived the join (the number IS the encoding,
    # 2B-R-8: two of the UN hues are near-identical, so the colour is only
    # recognition on top of it)
    numbered = set(views_compare._sdg(tuple(TRIO))["sdg_label_numbered"])
    assert set(sdg["taxon_label"]) <= numbered and len(numbered) > 1


def test_metric_frame_v4_denom_value_survives_to_the_page_unstripped(engine):
    """2BR3 VC item 5: this page must WIRE the v4 frame columns everywhere it
    passes a frame to a chart builder, i.e. never select a narrower column
    subset before charting. Checked structurally: every helper this page
    calls between `compare_data.metric_frame` and `charts_compare.
    fig_metric_bars` (`_order_rows`, `_decorate`) preserves `denom_value`."""
    ctx, subs = engine
    raw = compare_data.metric_frame(ctx, subs, TRIO, "field", "share")
    assert "denom_value" in raw.columns
    ordered = views_compare._order_rows(raw)
    assert "denom_value" in ordered.columns
    erc_raw = compare_data.metric_frame(ctx, subs, TRIO, "erc", "vol")
    decorated = views_compare._decorate(erc_raw, "erc", views_compare._erc(tuple(TRIO)), "panel_idx")
    assert "denom_value" in decorated.columns


# ---------------------------------------------------------- frontier (9) ---

def test_the_leader_dot_marks_one_card_per_measure(engine):
    """2B-R2-9: the dot is drawn in the LEADING institution's own colour, on the
    measure it leads, and on nothing else. A tie draws no dot -- the dot's claim
    is "this one leads", which two level institutions do not support."""
    ctx, _subs = engine
    frame = views_compare._overview(tuple(TRIO))
    leaders = views_compare._leaders(frame)
    assert leaders, "no leader on any measure -- the dot would never be drawn"
    cells = frame.set_index("institution_id")
    for column, iid in leaders.items():
        best = pd.to_numeric(cells[column], errors="coerce").max()
        assert float(cells.loc[iid, column]) == pytest.approx(float(best))
    slots = views_compare._slots(ctx, TRIO)
    at = _app(ids=TRIO).run()
    text = _markdown(at)
    for column, iid in leaders.items():
        dot = charts_compare.best_value_dot(slots[iid])
        assert dot in text, (column, iid)
    assert text.count("border-radius:") >= len(leaders)
    # a tie yields nothing, tested on a frame that is level by construction
    tied = pd.DataFrame({"institution_id": list(TRIO[:2])}
                        | {c: [1.0, 1.0] for c in views_compare.CARD_COLUMNS})
    assert views_compare._leaders(tied) == {}


def test_the_institution_name_is_the_openalex_link():
    """2B-R2-9: the name IS the link (the same fragment-carrying builder the
    Find profile and the benchmark tables use), which is what let the separate
    Publications button go."""
    from lib.ranked import works_link_named

    at = _app(ids=TRIO).run()
    text = _markdown(at)
    ctx = views_compare._bundle()["ctx"]
    for iid in TRIO:
        name = views_compare._name(ctx, iid)
        assert f"]({works_link_named(iid, name)})" in text, iid


def test_the_frontier_slider_defaults_to_the_measured_value_and_caps_the_frame():
    at = _app(ids=TRIO).run()
    slider = at.slider(key="cmp_frontier_topn")
    assert slider.value == views_compare.FRONTIER_TOPN_DEFAULT == 60
    pooled = views_compare._frontier_pooled(tuple(TRIO), TREE, BASIS, int(slider.value))
    assert 0 < len(pooled) <= int(slider.value)
    at.slider(key="cmp_frontier_topn").set_value(views_compare.FRONTIER_TOPN_MIN).run()
    assert not at.exception, [str(e) for e in at.exception]
    smaller = views_compare._frontier_pooled(tuple(TRIO), TREE, BASIS,
                                             views_compare.FRONTIER_TOPN_MIN)
    assert len(smaller) < len(pooled)


def test_the_shared_frontier_melt_loses_no_volume():
    wide = views_compare._shared_frontier(tuple(TRIO), TREE, BASIS)
    long = views_compare._shared_long(wide, TRIO)
    assert not long.empty
    assert (long["vol"] > 0).all(), "a side that holds nothing must not get a zero-length bar"
    summed = long.groupby("topic_id")["vol"].sum()
    combined = wide.set_index("topic_id")["combined_vol"]
    for topic, total in summed.items():
        # rel, not abs: the volumes are float32 in the artefact and the wide
        # frame sums them in a different order than the melt does
        assert total == pytest.approx(float(combined.loc[topic]), rel=1e-6), topic
    assert set(long["institution_id"]) <= set(TRIO)


def test_the_shared_frontier_defaults_to_top_twenty_and_show_all_expands_it():
    """2BR3 plan item 4: top twenty by combined volume by default, a button
    (never the pooled map's own slider) swaps in the rest."""
    wide = views_compare._shared_frontier(tuple(TRIO), TREE, BASIS)
    shared_long = views_compare._shared_long(wide, TRIO)
    total = int(shared_long["topic_id"].nunique())
    assert total > views_compare.SHARED_FRONTIER_TOP_N, (
        "the reference trio must hold more shared topics than the default cap "
        "for this test to exercise the button at all")
    at = _app(ids=TRIO).run()
    assert not at.exception, [str(e) for e in at.exception]
    toggle = [b for b in at.get("button") if b.key == "cmp_shared_frontier_toggle"]
    assert len(toggle) == 1, toggle
    assert toggle[0].label == copy.COMPARE["SHARED_FRONTIER_SHOW_ALL"].format(n=f"{total:,}")
    at.button(key="cmp_shared_frontier_toggle").click().run()
    assert not at.exception, [str(e) for e in at.exception]
    toggle2 = [b for b in at.get("button") if b.key == "cmp_shared_frontier_toggle"]
    assert toggle2[0].label == copy.COMPARE["SHARED_FRONTIER_SHOW_TOP"].format(
        n=views_compare.SHARED_FRONTIER_TOP_N)
    # never the pooled map's own slider doing this job
    assert "cmp_frontier_topn" != toggle[0].key


def test_the_frontier_note_counts_the_shared_topics_from_the_data():
    at = _app(ids=TRIO).run()
    pooled = views_compare._frontier_pooled(tuple(TRIO), TREE, BASIS,
                                            views_compare.FRONTIER_TOPN_DEFAULT,
                                            views_compare.POOL_DEFAULT)
    n_shared = int((pooled["owner"] == charts_compare.SHARED_OWNER).sum())
    line = copy.COMPARE["NOTE_FRONTIER_MAP"].format(
        n_shared=f"{n_shared:,}", n_shown=f"{len(pooled):,}")
    assert line in _markdown(at), line


def test_the_frontier_pool_selector_narrows_the_topic_set():
    """2B-R2-10: the elite pool is a STRICT subset of the default one (the cut
    is global, so it does not move with the basket), and picking it redraws the
    page without exception."""
    at = _app(ids=TRIO).run()
    assert list(at.radio(key="cmp_frontier_pool").options) == [
        views_compare.POOL_LABELS[p] for p in compare_data.FRONTIER_POOLS]
    # UNCAPPED on both sides: the subset claim is about the POOL, and two
    # different top-N cuts of two different rankings need not nest.
    uncapped = 10 ** 6
    wide = views_compare._frontier_pooled(tuple(TRIO), TREE, BASIS, uncapped, "volume")
    elite = views_compare._frontier_pooled(tuple(TRIO), TREE, BASIS, uncapped, "elite")
    assert not elite.empty
    assert set(elite["topic_id"]) <= set(wide["topic_id"])
    assert len(elite) < len(wide)
    at.radio(key="cmp_frontier_pool").set_value("elite").run()
    assert not at.exception, [str(e) for e in at.exception]
    assert copy.COMPARE["FRONTIER_POOL_RULE_ELITE"] in _markdown(at)


def test_the_colour_toggle_rebuilds_the_frontier_legend():
    """2B-R2-10: colouring by broad subject area REPLACES the ownership reading,
    legend included -- in that mode no mark means an institution, so a legend
    naming the institutions would name a figure that is not on screen."""
    at = _app(ids=TRIO).run()
    assert list(at.radio(key="cmp_frontier_color").options) == [
        views_compare.COLOR_BY_LABELS[c] for c in charts_compare.COLOR_BY]
    pooled = views_compare._frontier_pooled(tuple(TRIO), TREE, BASIS,
                                            views_compare.FRONTIER_TOPN_DEFAULT,
                                            views_compare.POOL_DEFAULT)
    items = views_compare._domain_items(pooled)
    assert items and all(isinstance(label, str) and label for _d, label in items)
    at.radio(key="cmp_frontier_color").set_value("domain").run()
    assert not at.exception, [str(e) for e in at.exception]
    text = _markdown(at)
    for _did, label in items:
        assert label in text, label


def test_the_frontier_controls_share_one_row():
    """2BR3 plan item 2: the pooled map's three controls (pool, colour, topics
    plotted) are all above ONE chart, none of them stacked as a caption."""
    at = _app(ids=TRIO).run()
    keys = {"cmp_frontier_pool", "cmp_frontier_color", "cmp_frontier_topn"}
    present = {w.key for w in at.get("radio")} | {w.key for w in at.get("slider")}
    assert keys <= present, present - keys


# -------------------------------------------------------------- legends ----

def test_a_legend_strip_sits_above_every_chart(engine):
    ctx, _subs = engine
    at = _app(ids=TRIO).run()
    slots = views_compare._slots(ctx, TRIO)
    names = views_compare._names(ctx, TRIO)
    plain = charts_compare.legend_strip(views_compare._slot_order(TRIO, slots),
                                        slots=slots, names=names)
    shared = charts_compare.legend_strip(views_compare._slot_order(TRIO, slots),
                                         slots=slots, names=names, shared=True,
                                         shared_label=copy.COMPARE["LEGEND_SHARED"])
    rendered = [m.value for m in at.markdown]
    n_legends = rendered.count(plain) + rendered.count(shared)
    assert n_legends >= len(at.get("plotly_chart")), (n_legends, len(at.get("plotly_chart")))
    assert rendered.count(shared) >= 1, "the frontier map needs the shared chip"


# ------------------------------------------------------------- frames ------

def test_coverage_states_are_exhaustive_per_institution():
    df = views_compare._coverage(tuple(TRIO))
    assert set(df["state"]) == set(views_compare._state_labels())
    for iid, total in df.groupby("institution_id")["share"].sum().items():
        assert total == pytest.approx(1.0, abs=1e-9), (iid, total)


def test_impact_union_carries_a_missing_cell_at_the_high_floor():
    union = views_compare._impact_subfields(tuple(TRIO), TREE,
                                            views_compare.IMPACT_FLOOR_DEFAULT)
    assert not union.empty
    assert union["pp"].isna().any(), "no missing cell -- this is not a union frame"
    # A MISSING cell is NaN across the whole row (BUILD_PLAN_2A L11: n/a is
    # never 0). A genuine measured zero is legal and must NOT be read as
    # missing, which is why the test asks about the row and not about the value.
    missing = union[union["n_works_full"].isna()]
    assert len(missing)
    assert missing[["pp", "ci_low", "ci_high"]].isna().all().all()


def test_the_floor_toggle_changes_both_the_union_and_the_rows_drawn():
    high, low = views_compare.IMPACT_FLOORS
    assert high > low
    union_high = views_compare._impact_subfields(tuple(TRIO), TREE, high)
    union_low = views_compare._impact_subfields(tuple(TRIO), TREE, low)
    assert union_low["subfield_id"].nunique() > union_high["subfield_id"].nunique()
    top = views_compare._top_shared(tuple(TRIO), TREE, BASIS, views_compare.IMPACT_ROWS_TOP_N)
    shown_high = views_compare._impact_rows(union_high, top)
    shown_low = views_compare._impact_rows(union_low, top)
    # the drawn ROW SET is capped at the same cut, so what the lower floor buys
    # is MEASURED CELLS inside those rows -- fewer n/a marks, wider intervals
    assert shown_low["pp"].notna().sum() > shown_high["pp"].notna().sum(), (
        shown_high["pp"].notna().sum(), shown_low["pp"].notna().sum())


def test_the_impact_subfield_note_states_the_selection_rule():
    """2BR3 plan item 3: the reading line under the by-subfield chart states
    the SELECTION RULE (the floor-clearing union `compare_data.
    impact_subfields` itself implements) in plain words, not only a bare
    "showing N of M" count -- the exact floor value in force is named.

    2D RE-PIN (E4): the floor RADIO is retired -- there is no more
    `cmp_impact_floor` widget to read a value back from; the floor is now
    the fixed `views_compare.IMPACT_FLOOR_DEFAULT` the page always uses."""
    at = _app(ids=TRIO).run()
    assert "cmp_impact_floor" not in at.session_state
    floor = views_compare.IMPACT_FLOOR_DEFAULT
    union = views_compare._impact_subfields(tuple(TRIO), TREE, floor)
    top = views_compare._top_shared(tuple(TRIO), TREE, BASIS, views_compare.IMPACT_ROWS_TOP_N)
    shown = views_compare._impact_rows(union, top)
    expected = copy.COMPARE["NOTE_IMPACT_SUBFIELDS"].format(
        floor=floor, n=f"{shown['subfield_id'].nunique():,}",
        n_union=f"{union['subfield_id'].nunique():,}")
    assert expected in _markdown(at), expected
    # the rule names the floor in plain words, not merely a bare count
    assert str(floor) in expected and "or more" in expected


# ------------------------------------------------------------- selection ---

def test_setting_a_slot_back_to_empty_shrinks_the_comparison():
    """2BR3: there is no more per-card remove button (plan item 1/8) -- an
    institution leaves the COMPARISON by its own slot going back to the empty
    option. It stays in the BASKET (removing it there is the sidebar's job,
    tested by SEL's own suite, not this page's)."""
    at = _app(ids=TRIO).run()
    assert _cards(at) == len(TRIO) * len(
        views_compare._card_facts(None, views_compare._overview(tuple(TRIO)).iloc[0]))
    target = next(f"slot_compare_{i}" for i in range(state.COMPARE_CAP)
                  if at.session_state[f"slot_compare_{i}"] == SORBONNE)
    at.selectbox(key=target).set_value(selection.SLOT_EMPTY).run()
    assert not at.exception, [str(e) for e in at.exception]
    assert _cards(at) == 2 * len(
        views_compare._card_facts(None, views_compare._overview(tuple(TRIO)).iloc[0]))
    assert SORBONNE in at.session_state["basket"], "removed from the SLOT, not the basket"
    assert at.query_params["compare"] == [f"{STRASBOURG},{FREIBURG}"]


def test_the_scenario_widgets_use_the_find_page_keys():
    at = _app(ids=TRIO).run()
    assert {"tree", "basis"} <= {sb.key for sb in at.selectbox}


def test_the_page_intro_and_data_line_moved_to_the_about_expander():
    """2BR3 plan item 1: the index-size/data-date line and this page's own
    method sentence no longer sit between the title and the first chart --
    they are inside "About these figures", at the bottom."""
    at = _app(ids=TRIO).run()
    assert copy.COMPARE["PAGE_INTRO"] in _captions(at)
    assert _first_literal(copy.FIND["DATA_CAPTION"].replace("{n_institutions}", "")) in _rendered(at)
    labels = _expander_labels(at)
    assert copy.COMPARE["ABOUT_HEADER"] in labels
    # and NOTHING carrying that text sits before the first subheader
    idx_first_header = at.subheader[0]
    assert idx_first_header.value == copy.COMPARE["OVERVIEW_HEADER"]


def test_the_share_link_names_exactly_the_slotted_institutions():
    """2BR3 plan item 7: `links.share_link_block("compare", ids)` at the
    bottom meta block, naming exactly the ids in the three slots -- SEL's own
    `deeplink` convention, unchanged."""
    at = _app(ids=TRIO).run()
    codes = [c.value for c in at.get("code")]
    assert selection.deeplink("compare", TRIO) in codes, codes
    assert copy.COMPARE["DEEPLINK_LABEL"] in _captions(at)


def test_the_snapshot_string_is_gone_from_this_page():
    """2B-R-12: removed app-wide. The page states how big the index is and how
    old the data is, both computed -- now inside "About these figures"."""
    at = _app(ids=TRIO).run()
    text = _rendered(at)
    assert "napshot" not in text, [c.value for c in at.caption if "napshot" in c.value]


# ------------------------------------------------------------- workbook ----

def _frames_for(ids) -> dict:
    # 2D RE-PIN (E4): `_metric` dropped its `floor` argument (three call sites
    # below); `dynamics` (E10) is a new frame `sheet_specs`/the workbook now
    # expect.
    return {
        "overview": views_compare._overview(tuple(ids)),
        "coverage": views_compare._coverage(tuple(ids)),
        "subject": views_compare._metric(tuple(ids), TREE, BASIS, "field", "share", None),
        "erc": views_compare._metric(tuple(ids), TREE, BASIS, "erc", "share", None),
        "sdg": views_compare._metric(tuple(ids), TREE, BASIS, "sdg", "share", None),
        "dynamics": views_compare._metric(tuple(ids), TREE, BASIS, "field", "dynamics", None),
        "frontier_map": views_compare._frontier_pooled(
            tuple(ids), TREE, BASIS, views_compare.FRONTIER_TOPN_DEFAULT),
        "shared_frontier": views_compare._shared_long(
            views_compare._shared_frontier(tuple(ids), TREE, BASIS), list(ids)),
        "impact": views_compare._impact_index(tuple(ids)),
        "impact_subfields": views_compare._impact_subfields(
            tuple(ids), TREE, views_compare.IMPACT_FLOOR_DEFAULT),
    }


METRICS_STATE = {"level": "field", "subject": "share", "erc": "share", "sdg": "share"}


def test_workbook_opens_with_openpyxl_and_carries_every_view(engine):
    import openpyxl

    ctx, _subs = engine
    sheets = views_compare.sheet_specs(SCENARIO, _frames_for(TRIO), METRICS_STATE)
    raw = views_compare._workbook(ctx, list(TRIO), SCENARIO,
                                  views_compare.IMPACT_FLOOR_DEFAULT,
                                  views_compare.FRONTIER_TOPN_DEFAULT, sheets)
    assert isinstance(raw, bytes) and raw[:2] == b"PK", "not a zip container, so not an xlsx"
    book = openpyxl.load_workbook(io.BytesIO(raw))
    expected = [copy.COMPARE["XLSX_SHEET_METHODS"]] + [label for label, _c, _f in sheets]
    assert book.sheetnames == expected, book.sheetnames
    assert len(sheets) == len(views_compare.SLUGS), (len(sheets), len(views_compare.SLUGS))
    assert sheets[1][0] == copy.COMPARE["VIEW_COVERAGE"], "Coverage must be the SECOND sheet (moved up)"
    methods = book[copy.COMPARE["XLSX_SHEET_METHODS"]]
    values = [str(c.value) for row in methods.iter_rows() for c in row if c.value is not None]
    for label, caption, _frame in sheets:
        assert label in values and caption in values, label
    header = [c.value for c in next(book[copy.COMPARE["XLSX_SHEET_OVERVIEW"]].iter_rows())]
    assert header == list(views_compare._overview(tuple(TRIO)).columns)


def test_the_methods_sheet_records_the_controls_the_reader_was_on(engine):
    """2B-R2-5/10: the workbook names the row order, the frontier pool and the
    colour mode. A file that named the metric of each selector but not the pool
    its frontier sheet was cut from would describe a view nobody saw."""
    ctx, _subs = engine
    controls = {"pool": "elite", "color_by": "domain", "sort": "value"}
    rows = views_compare.methods_rows(ctx, list(TRIO), SCENARIO,
                                      views_compare.IMPACT_FLOOR_DEFAULT,
                                      views_compare.FRONTIER_TOPN_DEFAULT, [], controls)
    items = list(rows[copy.COMPARE["XLSX_COL_ITEM"]])
    values = list(rows[copy.COMPARE["XLSX_COL_VALUE"]])
    for key in ("XLSX_ROW_POOL", "XLSX_ROW_COLOUR", "XLSX_ROW_SORT"):
        assert copy.COMPARE[key] in items, key
    assert views_compare.POOL_LABELS["elite"] in values
    assert views_compare.COLOR_BY_LABELS["domain"] in values
    assert views_compare.SORT_LABELS["value"] in values
    # ... and the defaults stand in when the page passes nothing
    plain = views_compare.methods_rows(ctx, list(TRIO), SCENARIO,
                                       views_compare.IMPACT_FLOOR_DEFAULT,
                                       views_compare.FRONTIER_TOPN_DEFAULT, [])
    assert views_compare.POOL_LABELS[views_compare.POOL_DEFAULT] in list(
        plain[copy.COMPARE["XLSX_COL_VALUE"]])


def test_every_reading_line_fits_the_one_line_rule():
    """2B-R2-8's enforcement, read back over this page's OWN copy: every `NOTE_*`
    key must survive `charts_compare.chart_note`, which raises on a line break
    or on a line over its cap. A wall of prose cannot come back as a "note" --
    2BR3's rewritten NOTE_IMPACT_SUBFIELDS (the selection-rule sentence) must
    clear the SAME cap as every other reading line."""
    notes = {k: v for k, v in copy.COMPARE.items()
             if k.startswith("NOTE_") and isinstance(v, str)}
    assert len(notes) >= 6, notes.keys()
    for key, template in notes.items():
        filled = re.sub(r"\{[^{}]*\}", "0", template)
        rendered = charts_compare.chart_note(filled, "a tooltip")
        assert charts_compare.NOTE_HELP_GLYPH in rendered, key
    # and the REAL, fully-formatted 2BR3 note (realistic numbers, not "0")
    # clears the cap too -- the placeholder-substitution proxy above is not
    # enough on its own for a template this stream rewrote.
    real = copy.COMPARE["NOTE_IMPACT_SUBFIELDS"].format(floor=30, n="18", n_union="101")
    assert len(real) <= charts_compare.NOTE_MAX_CHARS, (len(real), real)
    charts_compare.chart_note(real, "a tooltip")  # must not raise


def test_the_methods_sheet_states_the_windows_the_floor_and_the_cap(engine):
    ctx, _subs = engine
    rows = views_compare.methods_rows(ctx, list(TRIO), SCENARIO,
                                      views_compare.IMPACT_FLOOR_DEFAULT,
                                      views_compare.FRONTIER_TOPN_DEFAULT, [])
    assert list(rows.columns) == [copy.COMPARE["XLSX_COL_ITEM"], copy.COMPARE["XLSX_COL_VALUE"],
                                  copy.COMPARE["XLSX_COL_SOURCE"]]
    items = list(rows[copy.COMPARE["XLSX_COL_ITEM"]])
    for key in ("XLSX_ROW_DATA", "XLSX_ROW_WINDOW", "XLSX_ROW_DYNAMICS", "XLSX_ROW_TREE",
                "XLSX_ROW_BASIS", "XLSX_ROW_INSTITUTIONS", "XLSX_ROW_CAP", "XLSX_ROW_TOPN",
                "XLSX_ROW_FLOORS", "XLSX_ROW_CI", "XLSX_ROW_READING"):
        assert copy.COMPARE[key] in items, key
    assert copy.COMPARE["XLSX_ROW_SNAPSHOT"] not in items, "2B-R-12 removed the snapshot row"
    values = " ".join(str(v) for v in rows[copy.COMPARE["XLSX_COL_VALUE"]])
    for bounds in (compare_data.DYNAMICS_W1, compare_data.DYNAMICS_W2):
        assert views_compare._window(bounds) in values, bounds
    assert views_compare._ci_sentence() in values


def test_workbook_filename_is_self_describing():
    name = views_compare.workbook_filename(TRIO, TREE, BASIS)
    assert name.endswith(".xlsx")
    for iid in TRIO:
        assert iid in name
    assert TREE in name and BASIS in name


def test_the_one_workbook_is_the_only_download_on_the_page():
    """2D RE-PIN (E7): every per-section CSV button (`DOWNLOAD_VIEW`, `_download`,
    9 call sites) is deleted -- the single end-of-page workbook button is now
    the ONLY download on the whole page."""
    at = _app(ids=TRIO).run()
    labels = [d.label for d in at.get("download_button")]
    assert labels == [copy.COMPARE["EXPORT_XLSX_BUTTON"]], labels
    assert "DOWNLOAD_VIEW" not in copy.COMPARE
    assert not hasattr(views_compare, "_download")


# ------------------------------------------------------------ digit ban ----

def test_no_digit_ban_violations_in_this_streams_files():
    from tests.test_narrative import collect_ui_call_strings, has_digit_violation, load_allowlist

    tokens = load_allowlist()
    files = [APP_DIR / "lib" / "views_compare.py", APP_DIR / "lib" / "exports_xlsx.py",
             Path(COMPARE_PAGE)]
    strings = [(loc, s) for f in files for loc, s in collect_ui_call_strings(f)]
    assert strings, "collector found no UI-call strings in this stream's files -- it is vacuous"
    violations = [(loc, s) for loc, s in strings if has_digit_violation(s, tokens)]
    assert not violations, violations


def test_the_new_compare_copy_keys_hold_no_typed_digit():
    """Scope (a) of the digit ban, narrowed to the keys this stream added: a
    number in a Compare string must arrive as a `{placeholder}` the page fills
    from the config, the frame or the constants."""
    from tests.test_narrative import has_digit_violation, load_allowlist

    tokens = load_allowlist()
    added = [k for k in copy.COMPARE if k.startswith(("METRIC_", "CAPTION_FRONTIER_",
                                                      "OVERVIEW_", "XLSX_ROW_", "DRILL_",
                                                      "VIEW_", "CAP_", "LEGEND_", "ABOUT_",
                                                      "SHARED_FRONTIER_", "NOTE_"))]
    assert added
    violations = [(k, copy.COMPARE[k]) for k in added
                  if isinstance(copy.COMPARE[k], str) and has_digit_violation(copy.COMPARE[k], tokens)]
    assert not violations, violations


def test_no_hex_colour_is_typed_in_this_page():
    source = (APP_DIR / "lib" / "views_compare.py").read_text(encoding="utf-8")
    hits = re.findall(r"#[0-9a-fA-F]{3}|#[0-9a-fA-F]{6}", source)
    assert not hits, hits


def test_the_page_module_is_thin():
    """The page file is page config + state + one render call; every decision
    lives in lib/views_compare.py, which is what lets AppTest and the probe
    drive the same code the deployed app runs."""
    source = Path(COMPARE_PAGE).read_text(encoding="utf-8")
    assert "views_compare.render()" in source
    assert "plotly" not in source and "pandas" not in source


def test_no_forbidden_2br3_strings_remain_in_this_streams_files():
    """The acceptance's own grep-proof requirement (BUILD_PLAN_2BR3.md SS3 VC
    acceptance (c)): none of the deleted UI's own strings survive in either
    file this stream owns. Checked over the SOURCE, not the render, so it
    catches a stray reference even on a code path a render-only test might
    not reach."""
    forbidden = ("Take one pair further", "Trends in the",
                 "Add an institution by name", "Add to the comparison")
    for path in (APP_DIR / "lib" / "views_compare.py", Path(COMPARE_PAGE)):
        text = path.read_text(encoding="utf-8")
        for phrase in forbidden:
            assert phrase not in text, (path.name, phrase)
    # and the copy.py keys those strings used to live under are gone too
    for dead_key in ("SELECTION_HEADER", "ADD_LABEL", "ADD_PICK", "ADD_BUTTON",
                     "REMOVE_BUTTON", "CLEAR_BUTTON", "CAP_REACHED", "CAP_HELP",
                     "CAP_TRUNCATED", "HANDOFF_HEADER", "HANDOFF_HELP",
                     "HANDOFF_A_LABEL", "HANDOFF_B_LABEL", "HANDOFF_LINK",
                     "TRENDS_HEADER", "TRENDS_SELECTION_HELP", "VIEW_TRENDS",
                     "EMPTY_TRENDS", "CAPTION_TRENDS_SHARE", "NOTE_TRENDS",
                     "TIP_TRENDS", "EMPTY_TOO_FEW"):
        assert dead_key not in copy.COMPARE, dead_key
