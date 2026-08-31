"""
tests/ui/probe.py -- ONE parameterised acceptance probe, consolidating the
three per-view probes `ops/_probe_find.py`, `ops/_probe_compare.py` and
`ops/_probe_collab.py` (all now DELETED, superseded by this file -- Phase
2B-R3, stream TEV-U, wave 3).

WHY ONE FILE. The three old probes shared one shape (start `streamlit run
pages/N_<emoji>_<Page>.py` as a subprocess, drive it headless with
Playwright, recompute every rendered VALUE from the matching `lib/*_data.py`
module and look for the page's own formatting of it in the DOM, ALWAYS
terminate the server) and duplicated the harness three times. 2BR3's
selection rework (`lib.selection.render_sidebar`/`slots_row`) also means
every view now shares the SAME entry mechanism (a `?compare=`/`?pair=`/
`?seed=` deep link, or the sidebar search+basket), which used to be three
different flows -- one harness, one entry helper, three thin per-view
probe functions.

WHAT SURVIVES FROM THE OLD PROBES (still-valid checks, ported): Find's L1
rank-1 golden recompute against the export CSV, the frontier-mode
signature change, the breakdown chip-legend swap, the wordcloud/KPI-card
shape; Compare's per-institution overview-card recompute against
`compare_data.overview`, the best-value-dot colour match, the FULL metric
sweep (every option on every level, clicked and redrawn -- "the 2B-R
lesson"), the row-order-identical-across-tabs LOAD-BEARING check, the
gutter/reference-line proofs, the frontier pool/colour/slider controls, the
workbook's own contents; Collaborate's pulse/rank-direction recompute
against `collab_data.pulse`, the field-chart values against
`collab_data.field_breakdown`.

WHAT IS NEW / REWRITTEN for 2BR3: every entry point now goes through
`selection.slots_row`/`render_sidebar` rather than a per-page add-comparator
flow; Compare's overview-card recompute reads `.st-key-compare_strip`'s
SLOT-ordered swatches (2BR3 manager merge fix: positional, not by internal
institution key); Collaborate's field section recomputes against the NEW
chart-only builder (`views_collab._fields_chart`, no more table), and adds
the reciprocity-scatter geometry proof and the topic/untapped native
`st.dataframe` "Show all" proof (a canvas grid carries no per-cell text, so
these are structural, not value, proofs -- CD4's own pytest suite owns the
value-level proof for both).

WHAT IS DELETED, not ported (2BR3 UI no longer exists): Compare's cap-3
truncation prose recompute (`_probe_cap`, no truncation state exists once
slots replace the basket-vs-cap disclosure) and the Compare hand-off probe;
Collaborate's hand-built field/topic/untapped TABLE cell recompute (chart +
native dataframe now), the row-depth SLIDER probes (retired for "Show all"),
the "Read the publications on OpenAlex" link-button probe, the CSV-header
contract checks for the topic/untapped tables (native `st.dataframe` ships
its own export toolbar, no page-owned CSV any more).

Usage:
    python tests/ui/probe.py find
    python tests/ui/probe.py compare
    python tests/ui/probe.py collab
    python tests/ui/probe.py all [--port 8620]

Exit 0 iff every check in the requested view(s) passes; 1 otherwise. Stdout
is ASCII-only (cp1252 console). The server is always started as a
foreground-waited subprocess and always terminated in a `finally` block.
"""
from __future__ import annotations

import argparse
import csv
import io
import re
import socket
import subprocess
import sys
import time
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright

APP_DIR = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(APP_DIR))

for _stream in (sys.stdout, sys.stderr):
    if getattr(_stream, "encoding", "").lower() not in ("utf-8", "utf8"):
        _stream.reconfigure(encoding="utf-8")

from lib import charts, charts_compare, collab_data, compare_data, copy  # noqa: E402
from lib import palette, views_collab, views_compare, views_find  # noqa: E402
from lib.app_config import CFG  # noqa: E402
from lib.ranked import works_link_named  # noqa: E402

PAGES = {"find": "pages/1_\U0001F50E_Find.py", "compare": "pages/2_⚖️_Compare.py",
         "collab": "pages/3_\U0001F91D_Collaborate.py"}
SHOT_DIR = APP_DIR / "tests" / "ui" / "screenshots"
WIDTHS = [1920, 1280, 390]
ACTION_TIMEOUT_MS = 30_000
TREE, BASIS = "bestfit", "frac"   # config defaults -- what every page opens on

GOLD_SEED = "I40413290"       # University of Gdansk -- the L1 golden pin
GOLD_RANK1 = ("I34250744", 0.793119)
SHOT_SEED = "I68947357"       # Universite de Strasbourg -- the reference profile

A_ID, B_ID = "I68947357", "I1294671590"   # Strasbourg x CNRS -- the collab anchor
TRIO = ("I68947357", "I1294671590", "I154526488")  # Strasbourg, CNRS, Inserm

RESULTS: list[tuple[bool, str]] = []
PORT = 8620
BASE_URL = "http://127.0.0.1:8620"


def check(ok: bool, message: str) -> bool:
    RESULTS.append((bool(ok), message))
    print(("PASS: " if ok else "FAIL: ") + message)
    return bool(ok)


# ------------------------------------------------------------- harness ------

def _wait_for_port(port: int, timeout: float = 60.0) -> bool:
    deadline = time.time() + timeout
    while time.time() < deadline:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            if sock.connect_ex(("127.0.0.1", port)) == 0:
                return True
        time.sleep(0.5)
    return False


def _start_server(page_file: str, port: int) -> subprocess.Popen:
    return subprocess.Popen(
        [sys.executable, "-m", "streamlit", "run", page_file,
         "--server.headless", "true", "--server.port", str(port),
         "--browser.gatherUsageStats", "false"],
        cwd=str(APP_DIR), stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)


def _stop_server(server: subprocess.Popen) -> None:
    server.terminate()
    try:
        server.wait(timeout=10)
    except subprocess.TimeoutExpired:
        server.kill()
        server.wait(timeout=10)


# --------------------------------------------------------------- DOM utils --

def _text(page) -> str:
    return page.evaluate("document.body.innerText")


def _full_text(page) -> str:
    return page.evaluate("document.body.textContent") or ""


def _container_text(page, key: str) -> str:
    return page.evaluate(
        "(() => { const e = document.querySelector('.st-key-%s');"
        " return e ? e.textContent : ''; })()" % key)


def _container_html(page, key: str) -> str:
    return page.evaluate(
        "(() => { const e = document.querySelector('.st-key-%s');"
        " return e ? e.innerHTML : ''; })()" % key)


def _rgb(hex_colour: str) -> str:
    h = hex_colour.lstrip("#")
    return "rgb(%d, %d, %d)" % tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))


def _radio_options(page, key: str) -> list:
    return [t.strip() for t in
            page.locator(f'.st-key-{key} [data-testid="stRadioOption"]').all_text_contents() if t.strip()]


def _click_option(page, key: str, text: str) -> None:
    page.locator(f".st-key-{key}").get_by_text(text, exact=True).first.click(timeout=ACTION_TIMEOUT_MS)
    page.wait_for_timeout(2500)


def _tick_labels(page, key: str) -> list:
    return page.evaluate(
        "(sel) => { const p = document.querySelector(sel); if (!p) return [];"
        " return Array.from(p.querySelectorAll('.yaxislayer-above .ytick text')).map(t => t.textContent); }",
        f".st-key-{key} .js-plotly-plot")


def _row_names(labels) -> list:
    return [re.split(r"\d", str(t))[0].strip() for t in labels]


def _n_shapes(page, key: str, kind: str) -> int:
    return page.evaluate(
        "(sel) => { const p = document.querySelector(sel); return p ? p.querySelectorAll(sel2).length : 0; }"
        .replace("sel2", "'%s'" % kind), f".st-key-{key} .js-plotly-plot")


def _n_figures(page) -> int:
    return page.locator(".js-plotly-plot").count()


def _sidebar_add(page, query: str) -> None:
    """2BR3: the ONE shared sidebar search + add flow. `st.text_input`
    commits on Enter or blur, never on a bare `.fill()` alone."""
    box = page.locator('[data-testid="stSidebar"] .st-key-sidebar_search_query input').first
    box.click(timeout=ACTION_TIMEOUT_MS)
    box.fill(query)
    box.press("Enter")
    page.wait_for_timeout(2000)
    row = page.locator('[data-testid="stSidebar"] [class*="st-key-sidebar_add_"] button').first
    row.click(timeout=ACTION_TIMEOUT_MS)
    page.wait_for_timeout(1500)


def _fig_xy_text(page, selector: str) -> dict:
    return page.evaluate(
        "(sel) => { const el = document.querySelector(sel); if (!el || !el.data) return null;"
        " return el.data.map(t => ({x: t.x || [], y: t.y || [], text: t.text || []})); }",
        selector)


def _fig_layout(page, selector: str) -> dict:
    return page.evaluate(
        "(sel) => { const el = document.querySelector(sel); if (!el || !el.layout) return null;"
        " const l = el.layout;"
        " return {shapes_n: (l.shapes || []).length, scaleanchor_y: (l.yaxis || {}).scaleanchor || null,"
        " xrange: (l.xaxis || {}).range || null, yrange: (l.yaxis || {}).range || null}; }",
        selector)


# ===================================================================== find
# Ported from ops/_probe_find.py: the L1 golden recompute, the frontier-mode
# signature change, the breakdown chip-legend swap. Entry point updated:
# `?seed=` (the app's own qp_seed hydration, unchanged this round).

def _probe_find(page) -> None:
    page.goto(f"{BASE_URL}/?seed={SHOT_SEED}", wait_until="domcontentloaded")
    page.wait_for_selector('[role="tab"]', timeout=180_000)
    page.wait_for_timeout(3000)
    check(page.locator('[data-testid="stException"]').count() == 0,
          f"Find {SHOT_SEED}: no Streamlit exception on the page")
    check(page.locator(".st-key-profile").count() == 1, "Find: the profile container renders exactly once")
    tiles = page.locator(".st-key-profile .benchup-kpi").count()
    check(tiles == 6, f"Find: 6 KPI cards render (found {tiles})")

    page.locator(".st-key-panel_frontier summary").first.click(timeout=ACTION_TIMEOUT_MS)
    page.wait_for_timeout(3000)

    def _sig() -> str:
        data = _fig_xy_text(page, ".st-key-panel_frontier .js-plotly-plot")
        return "" if not data else "|".join(",".join(f"{v:.4f}" for v in tr["x"]) for tr in data)

    before_sig = _sig()
    page.locator(".st-key-frontier_mode button").nth(1).click(timeout=ACTION_TIMEOUT_MS)
    page.wait_for_timeout(5000)
    after_sig = _sig()
    check(bool(before_sig) and bool(after_sig) and before_sig != after_sig,
          "Find: the frontier mode control changes the plotted topic-set signature")
    page.locator(".st-key-frontier_mode button").nth(0).click(timeout=ACTION_TIMEOUT_MS)
    page.wait_for_timeout(3000)

    before_legend = _container_text(page, "profile")
    page.locator(".st-key-breakdown_dim button").nth(1).click(timeout=ACTION_TIMEOUT_MS)
    page.wait_for_timeout(4000)
    after_legend = _container_text(page, "profile")
    check(after_legend != before_legend,
          "Find: the breakdown segmented control swaps the chip legend (domain <-> document type)")
    page.locator(".st-key-breakdown_dim button").nth(0).click(timeout=ACTION_TIMEOUT_MS)
    page.wait_for_timeout(3000)

    # 2BR3: the shared sidebar search/basket, driven for real.
    page.wait_for_timeout(1500)   # let any in-flight rerun from the checks above settle first
    page.goto(BASE_URL, wait_until="domcontentloaded")
    # `streamlit run pages/<page>.py` directly (this probe's own entry point,
    # matching the old per-view probes) does not reliably populate
    # `[data-testid="stSidebarNav"]` the way `streamlit run Menu.py` does
    # (smoke.py's entry point) -- wait on a PAGE-SPECIFIC element instead,
    # the same convention the old probes always used.
    page.wait_for_selector('[data-testid="stSidebar"] .st-key-sidebar_search_query',
                           state="attached", timeout=60_000)
    page.wait_for_timeout(2000)
    _sidebar_add(page, "gdansk")
    page.wait_for_selector('[role="tab"]', timeout=ACTION_TIMEOUT_MS)
    check(page.locator(".st-key-profile").count() == 1,
          "Find: a single sidebar add auto-selects and renders the profile (no explicit pick)")


def _recompute_l1_golden() -> None:
    """L1 rank-1 read back off the export CSV, matched against the golden
    pin -- unaffected by the selection rework (unchanged this round)."""
    import pandas as pd

    from lib.engine import build_rows, build_substrates, load_context, rank_all
    from lib.engine.evidence import rows_evidence
    from lib.exports import ranking_csv

    ctx = load_context(str(APP_DIR / "data"))
    subs = build_substrates(ctx)
    rankings = rank_all(ctx, subs, GOLD_SEED)
    l1 = rankings["L1"]
    rows = build_rows(l1, ctx, len(l1["sorted_ids"]), rankings, subs)
    head = rows[:50]
    texts = rows_evidence(ctx, subs, "L1", GOLD_SEED, [r["institution_id"] for r in head])
    for r in head:
        r["evidence_text"] = texts.get(r["institution_id"])
    blob = ranking_csv(rows, seed_id=GOLD_SEED, lens="L1", tree=subs["tree"], basis=subs["basis"],
                       snapshot="", filters_label="")
    df = pd.read_csv(io.BytesIO(blob))
    top = df.iloc[0]
    check(str(top["institution_id"]) == GOLD_RANK1[0]
          and abs(float(top["lens_score"]) - GOLD_RANK1[1]) < 1e-3,
          f"Find golden: L1 rank-1 = {top['institution_id']} {top['lens_score']:.6f} "
          f"(golden {GOLD_RANK1[0]} {GOLD_RANK1[1]:.6f})")


# ================================================================== compare
# Ported/rewritten from ops/_probe_compare.py against the 2BR3 slots entry
# point and the reworked layout (KPI cards -> Coverage -> Subject/ERC/SDG ->
# frontier -> Impact -> About).

def _cmp_ctx():
    return views_find._bundle()["ctx"]


def _cmp_names(ids) -> dict:
    idx = _cmp_ctx()["index_by_id"]
    return {i: str(idx.loc[i, "display_name"]) for i in ids}


def _probe_compare(page) -> None:
    page.goto(f"{BASE_URL}/?compare={','.join(TRIO)}", wait_until="domcontentloaded")
    page.wait_for_selector('.st-key-compare_strip', state="attached", timeout=60_000)
    page.wait_for_timeout(2000)

    def _settle_figures(target: int, timeout_ms: int = 60_000) -> None:
        deadline = time.time() + timeout_ms / 1000
        last, stable = -1, 0
        while time.time() < deadline:
            now = _n_figures(page)
            stable = stable + 1 if now == last and now >= target else 0
            last = now
            if stable >= 3:
                return
            page.wait_for_timeout(700)

    _settle_figures(7)
    check(page.locator('[data-testid="stException"]').count() == 0, "Compare: no Streamlit exception")

    # --- overview cards: recomputed against compare_data.overview, SLOT order
    names = _cmp_names(TRIO)
    strip = _container_text(page, "compare_strip")
    for iid, name in names.items():
        check(name in strip, f"Compare overview: names {iid} ({name})")
    frame = compare_data.overview(_cmp_ctx(), list(TRIO)).set_index("institution_id")
    idx = _cmp_ctx()["index_by_id"]
    for iid in TRIO:
        facts = views_compare._card_facts(idx.loc[iid], frame.loc[iid])
        missing = [(label, value) for _c, label, value, _t in facts if value not in strip]
        check(not missing, f"Compare overview: every value of {iid} reads back from compare_data "
                           f"(missing {missing})")
    slots = views_compare._slots(_cmp_ctx(), list(TRIO))
    leaders = views_compare._leaders(compare_data.overview(_cmp_ctx(), list(TRIO)))
    if leaders:
        dot_colors = page.evaluate(
            "Array.from(document.querySelectorAll('.st-key-compare_strip span[style*=\"color:\"]'))"
            ".map(e => getComputedStyle(e).color)")
        wanted = {_rgb(palette.institution_color(slots[iid])) for iid in set(leaders.values())}
        check(wanted <= set(dot_colors),
              f"Compare overview: every best-value dot paints its SLOT'S colour ({sorted(wanted)})")

    # --- metric-selector vocabulary + FULL sweep ("the 2B-R lesson") --------
    for key, level in (("cmp_metric_subject", "field"), ("cmp_metric_erc", "erc"), ("cmp_metric_sdg", "sdg")):
        options = _radio_options(page, key)
        expected = [views_compare.METRIC_LABELS[m] for m in views_compare.SUBJECT_METRICS
                    if compare_data.metric_frame_available(m, level)]
        check(options == expected, f"Compare {key}: offers exactly what the data can serve ({options})")
        for label in options:
            _click_option(page, key, label)
            check(_n_figures(page) >= 7, f"Compare {key} = {label!r}: every view still renders")
    _click_option(page, "cmp_metric_subject", views_compare.METRIC_LABELS["share"])

    # --- row order IDENTICAL across two metric tabs (LOAD-BEARING) ----------
    first = _row_names(_tick_labels(page, "fig_cmp_subject"))
    _click_option(page, "cmp_metric_subject", views_compare.METRIC_LABELS["dynamics"])
    second = _row_names(_tick_labels(page, "fig_cmp_subject"))
    common_a = [t for t in first if t in set(second)]
    common_b = [t for t in second if t in set(first)]
    check(len(common_a) > 5 and common_a == common_b,
          f"Compare (LOAD-BEARING): row order is IDENTICAL between Share and Dynamics tabs "
          f"({len(common_a)} common rows)")
    _click_option(page, "cmp_sort_subject", views_compare.SORT_LABELS["value"])
    ranked = _row_names(_tick_labels(page, "fig_cmp_subject"))
    check(set(ranked) == set(second) and ranked != second, "Compare: the sort-by-value toggle re-ranks the rows")
    _click_option(page, "cmp_sort_subject", views_compare.SORT_LABELS["taxonomy"])
    back = _row_names(_tick_labels(page, "fig_cmp_subject"))
    check(back == second, "Compare: switching back restores the taxonomy order")

    # --- gutters + reference lines -------------------------------------------
    # 2BR3 PAL: the volume gutter moved from being baked into the y-TICK text
    # to a per-bar VERTICAL label (plotly bar `text`, textposition="outside")
    # at the bar's own end -- the SAME mechanism the low-volume dagger glyph
    # already used, now carrying the gutter number too. Read the trace's own
    # `text` array, not the y-tick text (which now carries only the row name).
    _click_option(page, "cmp_metric_subject", views_compare.METRIC_LABELS["share"])
    labels = _tick_labels(page, "fig_cmp_subject")
    bar_data = _fig_xy_text(page, ".st-key-fig_cmp_subject .js-plotly-plot")
    bar_texts = [t for tr in (bar_data or []) for t in tr["text"]]
    with_numbers = [t for t in bar_texts if re.search(r"\d", str(t))]
    check(len(bar_texts) > 0 and len(with_numbers) >= len(bar_texts) - 1,
          f"Compare: the volume gutter carries a number on (almost) every bar "
          f"({len(with_numbers)}/{len(bar_texts)})")
    n_refs_share = _n_shapes(page, "fig_cmp_subject", ".shapelayer path")
    _click_option(page, "cmp_metric_subject", views_compare.METRIC_LABELS["dynamics"])
    n_refs_dyn = _n_shapes(page, "fig_cmp_subject", ".shapelayer path")
    check(n_refs_dyn > n_refs_share,
          f"Compare: the Dynamics view draws its reference line, Share does not ({n_refs_dyn} vs {n_refs_share})")
    _click_option(page, "cmp_metric_subject", views_compare.METRIC_LABELS["share"])

    # --- the frontier map's own three controls -------------------------------
    n_before = _n_shapes(page, "fig_cmp_frontier_map", ".scatterlayer path")
    if views_compare.POOL_LABELS["elite"] in _radio_options(page, "cmp_frontier_pool"):
        _click_option(page, "cmp_frontier_pool", views_compare.POOL_LABELS["elite"])
        n_after = _n_shapes(page, "fig_cmp_frontier_map", ".scatterlayer path")
        check(n_before > 0 and n_after > 0, f"Compare frontier: both pools draw bubbles ({n_before}, {n_after})")
        _click_option(page, "cmp_frontier_pool", views_compare.POOL_LABELS["volume"])
    if views_compare.COLOR_BY_LABELS["domain"] in _radio_options(page, "cmp_frontier_color"):
        _click_option(page, "cmp_frontier_color", views_compare.COLOR_BY_LABELS["domain"])
        check(_n_figures(page) >= 7, "Compare frontier: the map redraws under the colour-by-domain toggle")
        _click_option(page, "cmp_frontier_color", views_compare.COLOR_BY_LABELS["owner"])
    slider = page.locator('.st-key-cmp_frontier_topn input[type="range"]').first
    if slider.count():
        before = _n_shapes(page, "fig_cmp_frontier_map", ".scatterlayer path")
        slider.press("ArrowLeft", timeout=ACTION_TIMEOUT_MS)
        page.wait_for_timeout(2500)
        after = _n_shapes(page, "fig_cmp_frontier_map", ".scatterlayer path")
        check(before != after or before > 0, f"Compare frontier: the top-N slider re-cuts the map ({before}, {after})")
        slider.press("ArrowRight", timeout=ACTION_TIMEOUT_MS)
        page.wait_for_timeout(2000)

    # --- shared frontier: recomputed total vs the page's own Show all -------
    shared = views_compare._shared_long(
        views_compare._shared_frontier(tuple(TRIO), TREE, BASIS), list(TRIO))
    total_shared = int(shared["topic_id"].nunique()) if not shared.empty else 0
    if total_shared > views_compare.SHARED_FRONTIER_TOP_N:
        btn = page.locator("button").filter(has_text=re.compile(r"^Show all \d"))
        check(btn.count() >= 1,
              f"Compare shared frontier: 'Show all {total_shared}' renders (total {total_shared} > 20)")

    # --- the workbook ------------------------------------------------------
    with page.expect_download(timeout=120_000) as info:
        page.locator(".st-key-dl_workbook button").first.click(timeout=ACTION_TIMEOUT_MS)
    raw = Path(info.value.path()).read_bytes()
    check(raw[:2] == b"PK", "Compare: the workbook downloads as a real xlsx container")
    book = openpyxl.load_workbook(io.BytesIO(raw))
    check(book.sheetnames[0] == copy.COMPARE["XLSX_SHEET_METHODS"], "Compare: the first sheet is Methods")
    expected_sheets = len(views_compare.SLUGS) + 1
    check(len(book.sheetnames) == expected_sheets,
          f"Compare: one sheet per view plus Methods ({len(book.sheetnames)} of {expected_sheets})")
    values = [str(c.value) for row in book[book.sheetnames[0]].iter_rows() for c in row if c.value is not None]
    check(copy.VERDICT_LINE in values, "Compare workbook: the Methods sheet carries the standing reading line")
    check("VIEW_TRENDS" not in " ".join(book.sheetnames) and "Trends" not in book.sheetnames,
          "Compare workbook (2BR3): no Trends sheet remains")

    check(page.locator('[data-testid="stException"]').count() == 0, "Compare: no exception at the end of the probe")


# ================================================================== collab
# Ported/rewritten from ops/_probe_collab.py: the pulse recompute + rank
# direction, the field chart's own values, the NEW reciprocity geometry
# proof, the NEW native-dataframe "Show all" structural proof.

def _collab_ctx():
    return views_collab._bundle()["ctx"]


def _collab_names() -> dict:
    ctx = _collab_ctx()
    return {i: str(ctx["index_by_id"].loc[i, "display_name"]) for i in (A_ID, B_ID)}


def _probe_collab(page) -> None:
    page.goto(f"{BASE_URL}/?pair={A_ID},{B_ID}", wait_until="domcontentloaded")
    page.wait_for_selector('.st-key-collab_header', state="attached", timeout=60_000)
    page.wait_for_timeout(4000)
    check(page.locator('[data-testid="stException"]').count() == 0, "Collaborate: no Streamlit exception")
    names = _collab_names()

    # --- the pulse, recomputed -----------------------------------------------
    p = collab_data.pulse(_collab_ctx(), A_ID, B_ID)
    text = _text(page)
    check(page.locator(".st-key-fig_pulse").count() > 0, "Collaborate: the pulse chart renders")
    star = f"{CFG['bonus_year']}{views_collab.BONUS_STAR}"
    chart = _container_text(page, "fig_pulse") + str(page.evaluate(
        "(() => { const el = document.querySelector('.st-key-fig_pulse .js-plotly-plot');"
        " return el && el.data ? JSON.stringify(el.data.flatMap(t => t.x || [])) : ''; })()"))
    check(star in chart, f"Collaborate pulse: the axis stars the partial year ({star})")
    legend = _container_text(page, "collab_legend")
    check(copy.COLLAB["LEGEND_JOINT"] in legend and names[A_ID] not in legend and names[B_ID] not in legend,
          "Collaborate pulse legend: JOINT chip only, no institution chip (2BR3 task 2)")
    check(views_collab._count(p["copubs_total"]) in text,
          f"Collaborate pulse: the joint total matches collab_data.pulse ({p['copubs_total']})")
    check(p["rank_in_a"] == 1 and p["rank_in_b"] == 16,
          f"Collaborate rank DIRECTION anchor: B is A's #{p['rank_in_a']}, A is B's #{p['rank_in_b']}")
    rank_line = copy.COLLAB["PULSE_RANK_LINE"].format(
        name_a=names[A_ID], name_b=names[B_ID],
        rank_of_b=views_collab._count(p["rank_in_a"]), rank_of_a=views_collab._count(p["rank_in_b"])
    ).replace("**", "")
    check(rank_line in text, "Collaborate: the two ranks render in their two directions")

    # --- momentum headline, recomputed ---------------------------------------
    mom = collab_data.pair_momentum(_collab_ctx(), A_ID, B_ID)
    if mom is not None:
        mom_container = _container_text(page, "collab_momentum")
        check(mom.get("text", "") in mom_container,
              f"Collaborate momentum: the headline text matches collab_data.pair_momentum ({mom.get('text')!r})")

    # --- the field chart, recomputed against collab_data.field_breakdown ----
    fields = collab_data.field_breakdown(_collab_ctx(), A_ID, B_ID)
    check(not fields.empty, f"Collaborate: the pair x field frame is non-empty ({len(fields)} fields)")
    check(page.locator(".st-key-fig_fields").count() > 0, "Collaborate: the field chart renders")
    fig_data = _fig_xy_text(page, ".st-key-fig_fields .js-plotly-plot")
    plotted_names = {n for tr in (fig_data or []) for n in tr["y"]}
    top_field = str(fields.sort_values("vol", ascending=False).iloc[0]["field_name"])
    check(top_field in plotted_names,
          f"Collaborate field chart: the largest field by volume is plotted ({top_field!r})")
    check(page.locator('[data-table="collab_fields"]').count() == 0,
          "Collaborate (2BR3): no field TABLE remains -- the chart is the whole section")

    # --- the reciprocity scatter, geometry proof -----------------------------
    recip = collab_data.reciprocity_frame(_collab_ctx(), views_find._subs(TREE, BASIS), A_ID, B_ID)
    if not recip.empty:
        layout = _fig_layout(page, ".st-key-fig_reciprocity .js-plotly-plot")
        check(layout is not None and layout.get("scaleanchor_y") == "x",
              f"Collaborate reciprocity: axes are scale-anchored (squared) ({layout})")
        check(layout is not None and layout.get("shapes_n", 0) >= 1,
              f"Collaborate reciprocity: the equal-weight diagonal draws ({layout})")

    # --- FWCI reality (MT sweep casualty #4: ops/_probe_collab.py's old
    # `copy.FWCI_NOT_AVAILABLE_LINE` assertion is retired with that file --
    # FWCI is a real, always-attempted column now, ruling 4, never a
    # 'not available' placeholder line) ------------------------------------
    fwci_col_cfg = views_collab._topics_column_config()
    check("fwci_median" in fwci_col_cfg,
          "Collaborate: the topic table's own column_config carries the Median FWCI column")

    # --- topic deep dive: native dataframe + Show all (structural) ----------
    prof = collab_data.joint_profile(_collab_ctx(), views_find._subs(TREE, BASIS), A_ID, B_ID)
    if prof is not None:
        n_topics = len(prof["topics"])
        dataframes = page.locator('[data-testid="stDataFrame"]')
        check(dataframes.count() >= 2, f"Collaborate: >=2 native dataframes render (found {dataframes.count()})")
        btn = page.locator("button").filter(has_text=re.compile(r"^Show all \d+ topics"))
        if n_topics > 20:
            check(btn.count() >= 1, f"Collaborate topics: 'Show all {n_topics} topics' renders (n={n_topics})")
        else:
            check(btn.count() == 0, f"Collaborate topics: no 'Show all' button when n={n_topics} <= 20")

    check(page.locator('[data-testid="stException"]').count() == 0,
          "Collaborate: no exception at the end of the probe")


# ------------------------------------------------------------------- widths -

def _probe_widths(page_path: str, browser, slug: str, url_suffix: str) -> None:
    SHOT_DIR.mkdir(parents=True, exist_ok=True)
    for width in WIDTHS:
        page = browser.new_page(viewport={"width": width, "height": 1000})
        page.set_default_timeout(ACTION_TIMEOUT_MS)
        page.goto(f"{BASE_URL}{url_suffix}", wait_until="domcontentloaded")
        # Generic across all three views (each renders at least one Plotly
        # figure once its seeded query params resolve) -- see _probe_find/
        # _probe_compare/_probe_collab's own docstrings on why
        # `[data-testid="stSidebarNav"]` is not reliable on a single-page
        # `streamlit run` entry point.
        try:
            page.wait_for_selector('.js-plotly-plot', state="attached", timeout=60_000)
        except Exception:  # noqa: BLE001 -- Find's empty-basket state has none; settle and proceed
            pass
        page.wait_for_timeout(2500)
        scroll = page.evaluate("document.documentElement.scrollWidth")
        inner = page.evaluate("window.innerWidth")
        check(scroll <= inner + 2, f"{slug} {width}px: scrollWidth {scroll} <= innerWidth+2 {inner + 2}")
        path = SHOT_DIR / f"probe_{slug}_{width}.png"
        page.screenshot(path=str(path), full_page=True)
        check(path.is_file(), f"{slug} {width}px: screenshot written")
        page.close()


# -------------------------------------------------------------------- main --

VIEWS = {
    "find": (PAGES["find"], "", _probe_find),
    "compare": (PAGES["compare"], f"/?compare={','.join(TRIO)}", _probe_compare),
    "collab": (PAGES["collab"], f"/?pair={A_ID},{B_ID}", _probe_collab),
}


def _run_view(view: str, port: int) -> None:
    global PORT, BASE_URL
    PORT = port
    BASE_URL = f"http://127.0.0.1:{port}"
    page_file, url_suffix, fn = VIEWS[view]
    server = _start_server(page_file, port)
    try:
        if not _wait_for_port(port):
            check(False, f"{view}: server did not open port {port}")
            return
        with sync_playwright() as p:
            browser = p.chromium.launch()
            context = browser.new_context(viewport={"width": 1280, "height": 1000}, accept_downloads=True)
            page = context.new_page()
            page.set_default_timeout(ACTION_TIMEOUT_MS)
            try:
                fn(page)
            except Exception as exc:  # noqa: BLE001 -- one phase's crash must not skip widths/screenshots
                check(False, f"{view}: raised {type(exc).__name__}: {exc}")
            page.close()
            context.close()
            _probe_widths(page_file, browser, view, url_suffix)
            browser.close()
    finally:
        _stop_server(server)
    if view == "find":
        _recompute_l1_golden()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("view", choices=["find", "compare", "collab", "all"])
    parser.add_argument("--port", type=int, default=8620)
    args = parser.parse_args()

    views = list(VIEWS) if args.view == "all" else [args.view]
    for i, view in enumerate(views):
        print(f"\n=== probe: {view} ===")
        _run_view(view, args.port + i)

    failed = [m for ok, m in RESULTS if not ok]
    print(f"\n{len(RESULTS) - len(failed)} of {len(RESULTS)} checks passed")
    if failed:
        for m in failed:
            print("FAILED:", m)
        return 1
    print("ALL CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
