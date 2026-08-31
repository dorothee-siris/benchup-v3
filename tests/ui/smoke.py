"""
tests/ui/smoke.py -- Playwright smoke test against the LIVE Streamlit server
(BUILD_PLAN_2A.md Stream H; extended for R1/R2, Phase 2B, and now re-cut for
Phase 2B-R -- BUILD_PLAN_2BR.md Stream H -- against the search-on-validate
Find page, the L0..L9 bare-code benchmark tabs, the cap-3 Compare page and the
four-section Collaborate v2 page).

Cross-page persistence is still the load-bearing claim: the basket (a plain,
non-widget session_state list) and every keyed widget (persist_state="session")
must survive real Menu<->Find<->Compare<->Collaborate<->Methods navigation with
their widget KEYS unchanged.

2B-R changes this file must track (BUILD_PLAN_2BR.md S0/S1, S3 row H):
  * A12 -- Find is search-ON-VALIDATE: typing a query and pressing Enter opens
    a results selectbox but renders NOTHING else (no profile, no tabs) until a
    match is actually picked.
  * 2B-R-2 -- the profile carries FOUR KPI cards (not eight tiles), each with
    ONE subline (not two); the bonus year is starred on the yearly axis
    ("2025*") instead of a banner/caption; the data caption reads "<n>
    institutions . data from <date>", the old verbose snapshot stamp is gone.
  * 2B-R-7 -- the identity column carries two REAL co-publication facts
    (international / with a company) now that P2/P4 have landed the artefacts
    -- no longer "n/a" placeholders.
  * 2B-R-13 -- Top topics is cut at 30 (was 20) with no sort control; the SI
    unit grid is RETIRED for an outer-end text label on the SI marker itself
    (fields/subfields/ERC panels); the frontier panel's single top-N slider
    drives BOTH modes.
  * A11/2B-R-11 -- the tab strip carries ONLY the bare display code (L0..L9),
    the full name moved inside the tab body; with both optional lenses on, the
    twelve-tab strip must fit at 1280px with no silent Streamlit-tab scroll;
    institution names are the OpenAlex-works link everywhere (fragment trick,
    A10) -- proven with a real click + captured popup, since a canvas grid
    carries no DOM <a> to query.
  * 2B-R-4/5/6/7/8/9 -- Compare is cap-3 (basket stays 6), reorder buttons and
    the frontier facets/overlay toggle are RETIRED, replaced by ONE metric
    selector per section (share/vol_top10/pp/sdg_share/dynamics/si, plus a
    "vol" option 2B-R-8 asks for at ERC/SDG level) and two frontier charts
    (a pooled map + a diverging "who holds the shared frontier" list).
  * 2B-R-10 -- Collaborate is four sections (pulse / joint corpus / untapped /
    links) over a pair; a real below-floor pair renders the honest notice
    instead of an empty table.
  * 2B-R-11 -- Methods gains a lens concordance table ("Reading the lens
    codes") alongside the existing digit-ban / placeholder guarantees.

Navigation uses the app's OWN sidebar nav link (`[data-testid="stSidebarNav"] a`)
for every page hop -- NEVER `page.goto()` for a persistence check. `page.goto()`
tears down and recreates the browser's WebSocket session, which silently resets
exactly the state a persistence test exists to catch (Lorraine Phase 2
tests/ui/smoke.py; Portfolio Mapping INSPECTION_PLAYBOOK.md "Known pitfalls").
`goto` IS used for the very first page load, and for the below-floor
Collaborate pair check below, which is deliberately a fresh, standalone
session (it asserts no persistence claim).

All selectors are locale-independent: `.st-key-<key>` classes from the keyed
widgets/containers the app already emits, `[role=...]`, `[data-testid=...]` --
text is read only via `textContent` (never `innerText`, empty for an inactive
tab panel or a collapsed expander body even though its DOM nodes exist) and
only to ASSERT content, never to locate an element. `st.dataframe` renders a
canvas grid with no real text nodes for cell values, so row-level facts are
read from captions/keyed containers/a real downloaded file, or (for the
institution-link proof) from a captured popup after a real click.

Usage:
    python tests/ui/smoke.py --port 8611

Exit 0 iff every check passes, 1 otherwise. Prints one PASS/FAIL line per
check. Stdout is ASCII-only (cp1252 console).
"""
from __future__ import annotations

import argparse
import csv
import io
import re
import socket
import subprocess
import sys
import time
from pathlib import Path

import openpyxl
from playwright.sync_api import TimeoutError as PWTimeoutError
from playwright.sync_api import sync_playwright

# Windows consoles default to cp1252, on which a bare print() of "Gdańsk"
# raises UnicodeEncodeError and aborts the journey mid-flight (inspection I-2).
for _stream in (sys.stdout, sys.stderr):
    if getattr(_stream, "encoding", "").lower() not in ("utf-8", "utf8"):
        _stream.reconfigure(encoding="utf-8")

DEFAULT_APP_DIR = Path(__file__).resolve().parents[2]  # tests/ui/smoke.py -> app/
WIDTHS = [1920, 1280, 390]
GDANSK_QUERY = "gdansk"
# 2B-R-11a renumbers the DISPLAY codes but the SHOWN-LENS COUNT is unchanged:
# 8 defaults (L0..L7) + Overview + Aspirational = 10; + L7(optional, ->L9) = 11;
# + C1(optional, ->L8) too = 12 (A11's own tab-overflow acceptance number).
GDANSK_TAB_COUNT = 10
L7_ON_TAB_COUNT = 11
BOTH_OPTIONAL_TAB_COUNT = 12          # A11: C1 + L7 both on
ACTION_TIMEOUT_MS = 30_000     # time-box every wait so a hang FAILS, never blocks

SEP = "·"  # middle dot, matches lib/copy.py's own separator

SUBFIELDS_TOP_N = 30
# 2B-R-13 (FB handoff): topics panel cut raised 20 -> 30, sort control removed
# (same reasoning as subfields under R2/L34 -- "top N" is itself a
# volume-ordered concept).
TOPICS_TOP_N = 30

# 2B-R2-6: the profile's cards column is a SIX-card 2 x 3 grid now (was the
# 2B-R FOUR-tile grid) -- `lib/tiles.py`'s TILE_CLASS/SUBLINE_CLASS hooks. ONE
# small line per card: the index-baseline line for five of them, and for
# Publications the fractional-counting NOTE instead (own `.benchup-kpi-value2`
# hook) -- both share the SAME `.benchup-kpi-sub` class, which is why the sub
# count below is checked against N_CARDS, not N_CARDS - 1.
N_CARDS = 6
CARD_LABELS = ["Publications", "SDG-tagged share", "Frontier top-quartile share",
              "PP(top10%)", "International co-publications", "Industrial co-publications"]
# 2B-R2-1a: the Ifremer profile is BOTH an umbrella and type-corrected -- the
# assert that crashed the app at gate 2B-R. Rendered as its own dedicated
# check (fresh session, `?seed=`), never folded into the generic Gdansk walk.
CRASH_SEED = "I154202486"
IDENTITY_TYPE_CORRECTED_RE = re.compile(r"[A-Za-z_]+\*\s*\(was:\s*[A-Za-z_]+\)")
NO_LONGER_ON_FIND = "What counts as a publication"

# 2B-R2-13: a strategy officer must never meet a plan code, a build artefact,
# a pipeline/table name or a stream name anywhere this file renders. Kept as a
# short, hardcoded list here (never re-imported from the app's own forbidden-
# vocabulary test) so a rename of that test's own list cannot silently widen
# what this file is willing to accept.
FORBIDDEN_VOCAB = ("2B-R", "BUILD_PLAN", "artefact", "pipeline", "parquet")
FORBIDDEN_CODES_RE = re.compile(r"\b(MU3|CP3|LP3|VS3|FA3|CD3|WT2?|P6|G2|H2|I2)\b")

# 2B-R2-3: the Compare selector's own vocabulary and order. "Publications in
# the world top decile" (the old vol_top10 TAB) must never appear among the
# options; "Volume" is offered only where a level actually defines one.
SUBJECT_METRIC_LABELS = ["Share", "Specialisation", "PP(top10%)", "SDG-tagged share",
                         "Change in mean annual volume"]
VOL_TOP10_LABEL = "Publications in the world top decile"
VOLUME_LABEL = "Volume"
SORT_TAXONOMY_LABEL = "By subject area"
SORT_VALUE_LABEL = "Largest first"
POOL_VOLUME_LABEL = "Most published by this set"
POOL_ELITE_LABEL = "The most emerging topics only"
COLOR_OWNER_LABEL = "Who holds the topic"
COLOR_DOMAIN_LABEL = "Broad subject area"
LEGEND_SHARED_TEXT = "held by more than one"
LOW_VOLUME_GLYPH = "\N{DAGGER}"
NOT_OFFERED_HEADER = "Not shown here, and why"

PANEL_LABELS = [
    ("fields", "Fields"),
    ("subfields", f"Top {SUBFIELDS_TOP_N} subfields"),
    ("topics", "Top topics"),
    ("frontier", "Frontier positioning"),
    ("sdg", "SDG profile"),
    ("erc", "ERC profile"),
]

TREE_LABEL_BESTFIT = "Repaired taxonomy (best fit, default)"
TREE_LABEL_ORIGINAL = "OpenAlex taxonomy as published"
BASIS_LABEL_FRAC = "Fractional counting"
STRIP_TREE_ORIGINAL = f"taxonomy: {TREE_LABEL_ORIGINAL}"

LENS_GUIDE_HEADER = "How to read the lenses"
# A11 (2B-R-11a): the TAB itself now carries only the bare code; the full name
# ("L0 . Field overlap") moved inside the tab BODY, as `_lens_intro`'s own
# opening line -- checked separately below, never conflated with the tab text.
LENS0_TAB_CODE = "L0"
LENS0_FULL_NAME = f"L0 {SEP} Field overlap"
LENS_LEGEND_SUBSTR = "see the lens guide above"
# 2B-R-11a renumbers the internal "L2f" lens to display code "L4" -- the tab
# now carries ONLY that bare code, and `UNDEFINED_LENS_TEMPLATE` formats with
# `copy.LENS_DISPLAY_NAMES["L2f"]`, so the literal substring "L2f" no longer
# appears anywhere in the rendered page (copy.py: "L2f": "L4 . Shared
# specialisations").
L2F_TAB_CODE = "L4"
L2F_DISPLAY_NAME = f"L4 {SEP} Shared specialisations"

FRONTIER_MODE_TOP_IDX, FRONTIER_MODE_EMERGING_IDX = 0, 1
BREAKDOWN_DOMAIN_IDX, BREAKDOWN_DOCTYPE_IDX = 0, 1

# 2B-R-2: no bonus-year banner -- the bonus year is starred ON the yearly
# breakdown's own x-axis instead (config.yaml bonus_year: 2025).
BONUS_YEAR_AXIS_LABEL = "2025*"

DATA_CAPTION_RE = re.compile(
    r"[\d,]+\s+institutions\s+" + re.escape(SEP) + r"\s+data from\s+[A-Za-z]+\s+\d{1,2},\s+\d{4}")

RESULTS: list[tuple[bool, str]] = []
PORT = 8611
BASE_URL = "http://127.0.0.1:8611"

# ---------------------------------------------------------------------------
# Phase 2B / 2B-R narrative journey: Menu -> Find -> Compare -> Collaborate ->
# Methods. Distinct from the R2 "Basket" section (Sorbonne, Bologna, left in
# place -- untouched): this journey CLEARS the basket and rebuilds it from the
# Gdansk seed's own L1 (subfield overlap) ranking, so the compared set is a
# real top-overlap peer group rather than three arbitrary names.
#
# Every label compared for exact text below is a HARDCODED literal (never
# re-imported from lib/copy.py): importing the very string under test would
# make a rename compare against itself and pass vacuously -- the point of the
# non-vacuity proofs at the bottom of this file / README.md.
NAV_CARD_LABELS = ["Find peers", "Compare", "Collaborate", "How it is built"]
NAV_COMPARE, NAV_COLLAB, NAV_METHODS = "Compare", "Collaborate", "Methods"

COMPARE_MIN_FIGURES = 8       # subject/erc/sdg/frontier-map/shared-frontier/impact x2/trends/coverage
COMPARE_CAP = 3               # state.COMPARE_CAP (2B-R-4)
BASKET_CAP = 6                # state.BASKET_CAP (unchanged)
MIN_LEGEND_STRIPS = 4         # 2B-R-12: a legend strip above every chart section
CAP_TRUNCATED_SUBSTR = "The basket holds more institutions than a comparison can show at once"

XLSX_METHODS_SHEET = "Methods"
XLSX_SHEET_COUNT = 11          # Methods + 10 view sheets (sheet_specs, 2B-R re-cut)

# 2B-R2-11a Collaborate section headers, hardcoded literals (non-vacuity).
# NOTE: "What the two publish on together" (old order) is GONE -- the joint
# corpus is now read off the field-breakdown chart+table FIRST, then the
# top-shared-topics table; "The joint corpus, field by field" is the new
# section this stream's field-breakdown chart lives under.
COLLAB_SECTION_HEADERS = [
    "The relationship, year by year",
    "The joint corpus, field by field",
    "The topics the two publish on together",
    "Where the two overlap without publishing together",
    "Read the publications on OpenAlex",
]
COLLAB_TABLES = ("collab_fields", "collab_topics", "collab_untapped", "collab_siblings")
COLLAB_TABLES_BELOW_FLOOR = ("collab_untapped", "collab_siblings")
# 2B-R2-12: the pair tables ship at floor 5 now (was 3).
PAIR_FLOOR = 5
BELOW_FLOOR_NOTICE_RE = re.compile(
    r"This pair holds \d+ publications, under the " + str(PAIR_FLOOR)
    + r" a breakdown needs to stay readable")
GAP_TABLE_HEADER_SUBSTR = "does not publish in"
DOWNLOAD_GAPS_TEXT = "Download this gap list (CSV)"
JOINT_TOPICS_CSV_HEADER = ("topic_id,topic_name,subfield_id,subfield_name,field_id,field_name,"
                           "domain_id,domain_name,vol_w1,vol_w2,vol_2025,vol_total,n_covered,"
                           "n_top10,sdg_tagged_n,arrow,url")
FIELD_BREAKDOWN_CSV_HEADER = ("field_id,field_name,domain_id,domain_name,vol_w1,vol_w2,vol_2025,"
                              "vol_total,n_covered,n_top10,mean_citations,arrow,url")
UNTAPPED_CSV_HEADER = ("topic_id,topic_name,subfield_id,subfield_name,vol_a,vol_b,"
                       "joint_observed,joint_expected,gap,url")
TOPICS_TOP_N_CAP = 100
TOPICS_ROWS_DEFAULT = 20
TOPICS_ROWS_STEP = 10
TAXON_FILTER_KEYS = ("primary_topic.id:", "primary_topic.subfield.id:", "primary_topic.field.id:")
# LP's own proven real sub-floor pair: 2 joint works, under the floor of 5.
BELOW_FLOOR_A_ID = "I68947357"     # Universite de Strasbourg
BELOW_FLOOR_B_ID = "I109144446"    # Bavarian Academy of Sciences and Humanities
COLLAB_PAIR_A = "I68947357"        # Universite de Strasbourg
COLLAB_PAIR_B = "I1294671590"      # CNRS -- Strasbourg's own first partner

METHODS_MIN_SECTIONS = 14      # MU shipped 20; 10 was the pre-2B-R floor
LENS_CODES_TITLE = "Reading the lens codes"
PLACEHOLDER_RE = re.compile(r"\{[a-zA-Z_][a-zA-Z0-9_]*\}")


def check(ok: bool, message: str) -> bool:
    RESULTS.append((bool(ok), message))
    print(("PASS: " if ok else "FAIL: ") + message)
    return bool(ok)


def fail_section(name: str, exc: Exception) -> None:
    check(False, f"{name}: raised {type(exc).__name__}: {exc}")


# ------------------------------------------------------------- server -------

def _wait_for_port(port: int, timeout: float = 60.0) -> bool:
    deadline = time.time() + timeout
    while time.time() < deadline:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            if sock.connect_ex(("127.0.0.1", port)) == 0:
                return True
        time.sleep(0.5)
    return False


def _start_server(app_dir: Path, port: int) -> subprocess.Popen:
    # DEVNULL, not PIPE: every rerun logs a `use_container_width` deprecation
    # per st.dataframe call, which fills an unread pipe buffer and blocks the
    # server mid-probe (app/ops/_probe_find.py's own note).
    return subprocess.Popen(
        [sys.executable, "-m", "streamlit", "run", "Menu.py",
         "--server.headless", "true", "--server.port", str(port),
         "--browser.gatherUsageStats", "false"],
        cwd=str(app_dir), stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)


def _stop_server(server: subprocess.Popen) -> None:
    server.terminate()
    try:
        server.wait(timeout=10)
    except subprocess.TimeoutExpired:
        server.kill()
        server.wait(timeout=10)


# --------------------------------------------------------- DOM helpers ------

def _settle(page, ms: int = 2500) -> None:
    page.wait_for_timeout(ms)


def _wait_for(page, predicate, timeout_ms: int = 15_000, interval_ms: int = 300) -> bool:
    """Poll `predicate()` instead of a blind sleep -- needed after a scenario
    switch (tree/basis), which pays a real, measured cold `build_substrates`
    cost the FIRST time that (tree, basis) pair is hit in this server
    process."""
    deadline = time.time() + timeout_ms / 1000
    while time.time() < deadline:
        if predicate():
            return True
        page.wait_for_timeout(interval_ms)
    return False


def _all_text(page, selector: str) -> str:
    """textContent (not innerText) joined across every match -- reads content
    inside an inactive Streamlit tab panel too (st.tabs runs every tab body
    every rerun; only the active panel has non-empty innerText), and inside a
    collapsed `st.expander` body (same story: the body executes and mounts,
    only the visual display folds)."""
    return page.evaluate(
        "(sel) => Array.from(document.querySelectorAll(sel)).map(e => e.textContent).join('|')",
        selector)


def _full_page_text(page) -> str:
    return page.evaluate("document.body.textContent") or ""


def _no_exception(page, label: str) -> bool:
    return check(page.locator('[data-testid="stException"]').count() == 0,
                 f"{label}: no Streamlit exception on the page")


def _open_select(page, key: str) -> None:
    """Open a keyed selectbox: click it, wait for its (portal-rendered) option
    list. Streamlit 1.61's selectbox is a react-aria ComboBox -- the FIRST
    click on a fresh widget instance opens the listbox reliably, but a SECOND,
    already-focused round on the SAME widget can leave `aria-expanded="false"`
    after an identical click; `ArrowDown` (react-aria's own keyboard-accessible
    open) is the fallback."""
    loc = page.locator(f".st-key-{key} [data-baseweb='select']")
    if loc.count() == 0:
        loc = page.locator(f".st-key-{key}")
    loc.first.click(timeout=ACTION_TIMEOUT_MS)
    try:
        page.wait_for_selector('[role="option"]', timeout=3000)
    except PWTimeoutError:
        page.keyboard.press("ArrowDown")
        page.wait_for_selector('[role="option"]', timeout=ACTION_TIMEOUT_MS)


def _pick_option(page, text: str | None = None) -> None:
    opts = page.locator('[role="option"]')
    target = opts.filter(has_text=text).first if text else opts.first
    target.click(timeout=ACTION_TIMEOUT_MS)


def _selectbox_value(page, key: str) -> str:
    """A keyed selectbox's CURRENT selection -- the react-aria ComboBox
    input's own `value` property, not the container's text."""
    return page.locator(f".st-key-{key} input").first.input_value()


def _ensure_sidebar_open(page) -> None:
    """At a narrow viewport Streamlit collapses the sidebar (and its nav
    links) behind a hamburger control; open it first so the nav is
    interactable. A no-op when the sidebar is already expanded."""
    ctrl = page.locator('[data-testid="stSidebarCollapsedControl"] button, '
                         '[data-testid="stSidebarCollapsedControl"]')
    if ctrl.count() and ctrl.first.is_visible():
        ctrl.first.click(timeout=ACTION_TIMEOUT_MS)
        page.wait_for_timeout(500)


def _ensure_expander_open(page, key: str, probe_selector: str) -> None:
    """Open a keyed `st.expander` if its content is not currently visible.
    Every panel's/expander's body EXECUTES every rerun regardless of the
    expander's visual state, but that visual open/closed state resets to the
    coded `expanded=` default on the very next rerun -- so this is called
    before every interaction inside one, never assumed to still be open."""
    probe = page.locator(probe_selector).first
    if probe.count() == 0 or not probe.is_visible():
        page.locator(f".st-key-{key} summary").first.click(timeout=ACTION_TIMEOUT_MS)
        page.wait_for_timeout(700)


def _ensure_expander_open_by_text(page, text: str, probe_selector: str) -> None:
    """Same idiom as `_ensure_expander_open`, for the rare expander that
    carries no `key=` (`SHARED_EXPANDER` inside Collaborate's untapped
    section) -- located by its own summary TEXT instead of a `.st-key-`
    class. Used only to NAVIGATE (click to open), never to assert content, so
    this does not reopen the non-vacuity question `PANEL_LABELS` exists to
    answer."""
    probe = page.locator(probe_selector).first
    if probe.count() == 0 or not probe.is_visible():
        page.locator("summary").filter(has_text=text).first.click(timeout=ACTION_TIMEOUT_MS)
        page.wait_for_timeout(700)


def _click_nav(page, label: str) -> None:
    """Real in-app sidebar nav-link click -- the ONLY way this file changes
    page for a persistence check."""
    _ensure_sidebar_open(page)
    link = page.locator('[data-testid="stSidebarNav"] a').filter(has_text=label).first
    link.wait_for(state="visible", timeout=ACTION_TIMEOUT_MS)
    try:
        link.click(timeout=ACTION_TIMEOUT_MS)
    except Exception:
        link.evaluate("el => el.click()")
    _settle(page, 3000)


def _basket_count(page) -> int:
    return page.locator('[class*="st-key-rm_"]').count()


def _seed_heading(page) -> str:
    return page.locator(".st-key-profile h3").first.text_content() or ""


def _strip_text(page) -> str:
    return _all_text(page, ".st-key-strip")


def _chip_legend(page) -> str:
    return _all_text(page, '.st-key-profile div[style*="flex-wrap"]')


def _plotly_point_count(page, selector: str) -> int:
    """Total marker count across a live Plotly figure's own traces -- reads
    what is actually PLOTTED, never a caption."""
    return page.evaluate(
        "(sel) => { const el = document.querySelector(sel); if (!el || !el.data) return -1;"
        " return el.data.reduce((a, t) => a + ((t.x && t.x.length) || 0), 0); }",
        selector)


def _frontier_points(page) -> int:
    return _plotly_point_count(page, ".st-key-panel_frontier .js-plotly-plot")


def _si_label_info(page, fig_key: str) -> dict:
    """2B-R-13: the SI unit grid is retired for an outer-end text label on
    each row's own SI marker. Reads the live figure's own trace/layout data:
    the count of non-empty text labels on any `mode` that includes "text",
    and `showgrid` on the SI axis (the LAST `xaxis*` key, same idiom the old
    grid-tick reader used)."""
    return page.evaluate(
        """(sel) => {
            const el = document.querySelector(sel);
            if (!el || !el.data) return {n_labels: -1, showgrid: null};
            let n = 0;
            for (const t of el.data) {
                if (t.mode && t.mode.indexOf('text') >= 0 && Array.isArray(t.text)) {
                    n += t.text.filter(x => x).length;
                }
            }
            const keys = Object.keys(el.layout || {}).filter(k => /^xaxis/.test(k)).sort();
            const showgrid = keys.length ? (el.layout[keys[keys.length - 1]].showgrid) : null;
            return {n_labels: n, showgrid: showgrid};
        }""",
        f".st-key-{fig_key} .js-plotly-plot")


def _yearly_axis_labels(page) -> list:
    return page.evaluate(
        "(() => { const el = document.querySelector('.st-key-fig_breakdown_yearly .js-plotly-plot');"
        " if (!el || !el.data) return [];"
        " return el.data.flatMap(t => t.x || []); })()")


def _frontier_slider_locator(page):
    # Measured (debug probe, 2026-08-31): Streamlit's slider thumb carries no
    # `role="slider"` on this pinned build -- it is a visually-hidden real
    # `<input type="range">` (react-aria's own accessible-hide pattern), which
    # DOES respond to Arrow keys once focused. `.press()` on a locator focuses
    # the element first, so no separate click is needed (and a click would
    # miss: the input has zero visual size).
    return page.locator('[class*="st-key-frontier_topn_"] input[type="range"]').first


def _frontier_slider_step(page, steps: int) -> None:
    slider = _frontier_slider_locator(page)
    key = "ArrowRight" if steps > 0 else "ArrowLeft"
    for _ in range(abs(steps)):
        slider.press(key, timeout=ACTION_TIMEOUT_MS)


def _frontier_signature(page) -> str:
    """2B-R-13: BOTH frontier modes share ONE top-N slider, so a mode switch
    can leave the plotted POINT COUNT unchanged (both cut at the same top_n)
    even though the underlying topic SET differs -- `_frontier_points` alone
    is no longer sufficient proof the mode control does anything. Reads the
    live figure's own x/y arrays instead."""
    return page.evaluate(
        "(() => { const el = document.querySelector('.st-key-panel_frontier .js-plotly-plot');"
        " if (!el || !el.data) return '';"
        " return JSON.stringify(el.data.map(t => [t.x, t.y])); })()")


def _fig_xy_text(page, selector: str) -> dict:
    """Every trace's x/y/text arrays off a LIVE Plotly figure -- used to prove a
    chart actually carries data (2B-R2-11's new field-breakdown chart), not
    just that its container exists."""
    return page.evaluate(
        "(sel) => { const el = document.querySelector(sel); if (!el || !el.data) return null;"
        " return el.data.map(t => ({x: t.x || [], y: t.y || [], text: t.text || []})); }",
        selector)


def _table_rows(page, name: str) -> int:
    return page.locator(f'[data-table="{name}"] tbody tr[data-row]').count()


def _table_cells(page, name: str, selector: str, attr: str) -> list:
    return page.evaluate(
        "([n, sel, a]) => Array.from(document.querySelectorAll("
        "'[data-table=\"' + n + '\"] ' + sel)).map(e => e.getAttribute(a))",
        [name, selector, attr])


def _hrefs(page) -> list:
    return page.evaluate(
        "Array.from(document.querySelectorAll('a[href]')).map(a => a.getAttribute('href'))")


def _no_forbidden_vocab(page, label: str) -> None:
    """2B-R2-13: no plan code, build artefact, pipeline/table name or stream
    code anywhere the page renders -- text, tooltips (`title=`) and captions
    alike."""
    text = (_full_page_text(page) + "|"
            + page.evaluate("Array.from(document.querySelectorAll('[title]'))"
                            ".map(e => e.getAttribute('title')).join('|')"))
    low = text.lower()
    hits = [w for w in FORBIDDEN_VOCAB if w.lower() in low]
    check(not hits, f"{label}: no forbidden-vocabulary term renders ({hits})")
    code_hit = FORBIDDEN_CODES_RE.search(text)
    check(code_hit is None,
          f"{label}: no stream code renders on the page (found {code_hit.group(1) if code_hit else None!r})")


def _outer_label_bbox_check(page, fig_selector: str, label: str) -> None:
    """2B-R2-7: the SI marker's own outer-end value label must stay fully
    inside the figure's plot area at the widest measured cell (Ifremer's SI
    0.17-21.35 at 1280px). Reads every non-empty Plotly TEXT node's real
    bounding box against the figure's own `.main-svg` box -- a clip/overflow
    is invisible to a DOM-presence check, only a geometry one catches it."""
    fig = page.locator(fig_selector).first
    if fig.count() == 0 or not fig.is_visible():
        check(False, f"{label}: figure is visible for the outer-label bbox check")
        return
    plot_box = fig.locator(".main-svg").first.bounding_box()
    if plot_box is None:
        check(False, f"{label}: could not read the plot's own .main-svg bounding box")
        return
    texts = fig.locator(".scatterlayer text")
    n = texts.count()
    left, right = plot_box["x"], plot_box["x"] + plot_box["width"]
    offenders = []
    for i in range(n):
        t = texts.nth(i)
        content = (t.text_content() or "").strip()
        if not content:
            continue
        box = t.bounding_box()
        if box is None:
            continue
        if box["x"] < left - 1 or box["x"] + box["width"] > right + 1:
            offenders.append(f"{content!r} at x={box['x']:.1f}..{box['x'] + box['width']:.1f} "
                             f"(plot {left:.1f}..{right:.1f})")
    check(n > 0 and not offenders,
          f"{label}: every outer-end SI value label stays inside the plot ({n} checked)"
          + (f" -- offenders: {offenders}" if offenders else ""))


def _search_and_pick(page, query: str, pick_key: str = "seed_pick",
                      query_key: str = "seed_query", option_text: str | None = None) -> None:
    box = page.locator(f".st-key-{query_key} input").first
    box.click(timeout=ACTION_TIMEOUT_MS)
    box.fill(query)
    box.press("Enter")
    _settle(page, 2500)
    _open_select(page, pick_key)
    _pick_option(page, option_text)
    _settle(page, 3000)


# ------------------------------------------------------ undefined-L2f seed --

def _find_undefined_l2f_seed(app_dir: Path, tree: str = "original") -> tuple[str, str] | None:
    """Smallest-first scan: the smaller an institution, the more likely L2f's
    own floor-of-papers-per-cell rule leaves it undefined."""
    sys.path.insert(0, str(app_dir))
    from lib.data_cache import index
    from lib.engine import build_substrates, load_context, rank_all

    idx = index().sort_values("total_full_2020_2024")
    ctx = load_context(str(app_dir / "data"))
    subs = build_substrates(ctx, tree)
    for row in idx.itertuples(index=False):
        ranking = rank_all(ctx, subs, row.institution_id)
        l2f = ranking.get("L2f")
        if l2f and l2f.get("undefined"):
            return row.institution_id, row.display_name
    return None


# ------------------------------------------------------------- sections -----

def check_menu(page) -> None:
    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.wait_for_selector('[data-testid="stSidebarNav"]', state="attached", timeout=ACTION_TIMEOUT_MS)
    _settle(page, 2000)
    check(page.get_by_role("heading").count() >= 1, "Menu: heading present")
    nav = page.locator(".st-key-nav_cards")
    check(nav.count() >= 1, "Menu: .st-key-nav_cards container present")
    cards = nav.locator("[class*='st-key-nav_card_']")
    try:
        cards.first.wait_for(state="visible", timeout=ACTION_TIMEOUT_MS)
    except Exception:  # noqa: BLE001 -- the count check reports the failure
        pass
    check(cards.count() >= 3, f"Menu: >=3 nav cards (found {cards.count()})")
    find_link = nav.locator("a").filter(has_text="Find")
    check(find_link.count() >= 1, "Menu: Find card is live (st.page_link anchor present)")
    check(cards.count() == len(NAV_CARD_LABELS),
          f"Menu: exactly {len(NAV_CARD_LABELS)} nav cards render (found {cards.count()})")
    live_links = nav.locator("a")
    check(live_links.count() == len(NAV_CARD_LABELS),
          f"Menu: all {len(NAV_CARD_LABELS)} cards are live st.page_link anchors, none greyed "
          f"(found {live_links.count()} anchors, expected {len(NAV_CARD_LABELS)})")
    for label in NAV_CARD_LABELS:
        check(live_links.filter(has_text=label).count() >= 1,
              f"Menu: a live card links to {label!r}")
    _no_exception(page, "Menu")


def check_find_search(page) -> None:
    """2B-R-12/A12: search-on-validate. Typing a query and pressing Enter
    opens the results selectbox but renders NOTHING else below it -- no
    profile container, no benchmark tabs -- until an actual pick is made."""
    _click_nav(page, "Find")
    box = page.locator(".st-key-seed_query input")
    box.first.wait_for(state="visible", timeout=ACTION_TIMEOUT_MS)
    check(box.count() >= 1, "Find: seed search input present (.st-key-seed_query)")

    # 2B-R-12: the data caption, checked before anything is searched -- it is
    # part of the header, not the profile.
    caption_text = _all_text(page, '[data-testid="stCaptionContainer"]')
    check(bool(DATA_CAPTION_RE.search(caption_text)),
          f"Find header: the data caption reads '<n> institutions {SEP} data from <date>' "
          f"(2B-R-12) (captions: {caption_text[:200]!r})")
    check("(generated" not in caption_text,
          "Find header: the old verbose snapshot stamp ('(generated ...)') is gone (2B-R-12)")

    box.first.click(timeout=ACTION_TIMEOUT_MS)
    box.first.fill(GDANSK_QUERY)
    box.first.press("Enter")
    _settle(page, 2500)

    # A12 (NEW): nothing below the search box renders before a pick.
    check(page.locator(".st-key-profile").count() == 0,
          "Find (A12): no profile container renders after typing, before a pick")
    check(page.locator('[role="tab"]').count() == 0,
          "Find (A12): no benchmark tabs render after typing, before a pick")

    check(page.locator(".st-key-seed_pick").count() >= 1,
          "Find: results selectbox appeared after typing 'gdansk'")
    _open_select(page, "seed_pick")
    _pick_option(page)
    page.wait_for_selector('[role="tab"]', timeout=ACTION_TIMEOUT_MS)
    _settle(page, 3000)
    heading = _seed_heading(page)
    check("Gda" in heading, f"Find: seed profile heading contains 'Gda' (got {heading!r})")
    check(page.locator(".st-key-profile").count() == 1,
          "Find (A12): the profile renders exactly once after the pick")
    tabs = page.locator('[role="tab"]').count()
    check(tabs == GDANSK_TAB_COUNT, f"Find: default tab count is {GDANSK_TAB_COUNT} (got {tabs})")
    _no_exception(page, "Find (Gdansk seed)")


def _add_comparator(page, name: str) -> None:
    _search_and_pick(page, name, pick_key="basket_pick", query_key="basket_query")
    page.locator(".st-key-basket_add button").first.click(timeout=ACTION_TIMEOUT_MS)
    _settle(page, 2500)


def check_basket(page) -> None:
    _add_comparator(page, "Sorbonne")
    n1 = _basket_count(page)
    check(n1 == 1, f"Basket: 1 item after adding Sorbonne (got {n1})")
    _add_comparator(page, "Bologna")
    n2 = _basket_count(page)
    check(n2 == 2, f"Basket: 2 items after adding Bologna (got {n2})")
    _no_exception(page, "Basket add flow")


# --------------------------------------------------- controls / sidebar -----

def check_controls_placement(page) -> None:
    """L16: the sidebar holds ONLY the scenario selects (tree, basis) and the
    basket; depth/C1/L7/post-filters render in the MAIN area's controls row."""
    sidebar = page.locator('[data-testid="stSidebar"]')
    check(sidebar.locator(".st-key-tree").count() >= 1, "Sidebar: .st-key-tree (scenario) present")
    check(sidebar.locator(".st-key-basis").count() >= 1, "Sidebar: .st-key-basis (scenario) present")
    check(sidebar.locator(".st-key-depth").count() == 0, "Sidebar: no .st-key-depth")
    check(sidebar.locator(".st-key-f_types").count() == 0, "Sidebar: no .st-key-f_types")
    check(sidebar.locator(".st-key-c1_on").count() == 0, "Sidebar: no .st-key-c1_on")

    check(page.locator(".st-key-depth").count() >= 1,
          "Controls row: .st-key-depth renders in the main area")
    check(page.locator(".st-key-c1_on").count() >= 1,
          "Controls row: .st-key-c1_on renders in the main area")
    check(page.locator(".st-key-l7_on").count() >= 1,
          "Controls row: .st-key-l7_on renders in the main area")

    tree_val = _selectbox_value(page, "tree")
    check(tree_val == TREE_LABEL_BESTFIT,
          f"Sidebar: taxonomy selectbox shows the default DISPLAY label (got {tree_val!r})")
    check("bestfit" not in tree_val, "Sidebar: the internal taxonomy value never appears")
    basis_val = _selectbox_value(page, "basis")
    check(basis_val == BASIS_LABEL_FRAC,
          f"Sidebar: counting-basis selectbox shows the default DISPLAY label (got {basis_val!r})")
    check("frac" not in basis_val.replace(BASIS_LABEL_FRAC, ""),
          "Sidebar: the internal counting-basis value never appears")

    _ensure_expander_open(page, "postfilters", ".st-key-f_types input")
    check(page.locator(".st-key-f_types").count() >= 1,
          "Post-filters expander: reveals .st-key-f_types")
    check(page.locator(".st-key-f_countries").count() >= 1,
          "Post-filters expander: reveals .st-key-f_countries")

    cinp = page.locator(".st-key-f_countries input").first
    cinp.click(timeout=ACTION_TIMEOUT_MS)
    cinp.fill("Fra")
    page.wait_for_selector('[role="option"]', timeout=ACTION_TIMEOUT_MS)
    opt = page.locator('[role="option"]').filter(has_text="France")
    check(opt.count() >= 1, "Country filter: typing 'Fra' surfaces an option containing 'France'")
    page.keyboard.press("Escape")
    _settle(page, 500)
    cinp.fill("")
    _no_exception(page, "Controls placement / sidebar labels / post-filters")


# ----------------------------------------------------------- R2 profile -----

def check_profile_and_panels(page) -> dict:
    """2B-R-2/13: the profile container, its FOUR KPI tiles (each with one
    index-baseline subline), its wordcloud, its six chart-panel expanders
    (exact labels), the top-subfields cut with no sort control, the SDG
    panel's numbered labels, the frontier panel's two modes, and the
    breakdown pair's segmented control.

    Returns `{"frontier_points": int, "breakdown_legend": str}`, the OFF-
    DEFAULT values this function deliberately leaves the page in (frontier
    mode swapped to "emerging", breakdown swapped to "Document type")."""
    check(page.locator(".st-key-profile").count() == 1,
          "Profile: .st-key-profile container renders exactly once")
    check(page.locator('.st-key-profile [data-testid="stImage"] img').count() >= 1,
          "Profile: subfield wordcloud renders as an <img>")

    # ---- 2B-R2-6: the SIX cards, title-first, one small line each --------
    tiles = page.locator(".st-key-profile .benchup-kpi")
    check(tiles.count() == N_CARDS, f"Profile: {N_CARDS} cards render (found {tiles.count()})")
    full_text = _full_page_text(page)
    for label in CARD_LABELS:
        check(label in full_text, f"Profile: card {label!r} renders")
    sublines = page.locator(".st-key-profile .benchup-kpi-sub")
    n_sub = sublines.count()
    check(n_sub == N_CARDS, f"Profile: every card carries exactly one small line (found {n_sub})")
    sub_texts = [sublines.nth(i).text_content() or "" for i in range(n_sub)]
    n_baseline = sum(1 for t in sub_texts if "index median" in t)
    check(n_baseline == N_CARDS - 1,
          f"Profile: 5 of {N_CARDS} cards' lines read 'index median ...' (found {n_baseline})")
    n_notes = page.locator(".st-key-profile .benchup-kpi-value2").count()
    check(n_notes == 1,
          f"Profile: exactly the Publications card carries the fractional-counting note "
          f"(found {n_notes})")
    check("in fractional counting" in "|".join(
        sublines.nth(i).text_content() or "" for i in range(n_sub)),
        "Profile: the fractional-counting note names its own basis")
    check("Key figures" in full_text, "Profile: 'Key figures' header renders")
    check("ERC-classified share" not in full_text,
          "Profile: the retired coverage-line phrase 'ERC-classified share' is nowhere on the page")

    # ---- title-first card anatomy (2B-R2-6): name is the FIRST child of the
    # card, in a smaller/lighter weight than the bold value under it ---------
    first_card = tiles.first
    kids = first_card.locator("> div")
    n_kids = kids.count()
    check(n_kids >= 3, f"Profile: a card carries a label/value/subline stack ({n_kids} children)")
    if n_kids >= 2:
        label_size = kids.nth(0).evaluate("e => parseFloat(getComputedStyle(e).fontSize)")
        value_size = kids.nth(1).evaluate("e => parseFloat(getComputedStyle(e).fontSize)")
        label_weight = kids.nth(0).evaluate("e => parseInt(getComputedStyle(e).fontWeight)")
        value_weight = kids.nth(1).evaluate("e => parseInt(getComputedStyle(e).fontWeight)")
        check(value_size > label_size,
              f"Profile (2B-R2-6): the card's SECOND element (the value, {value_size}px) is bigger "
              f"than its FIRST (the name, {label_size}px) -- title-first, value bold under it")
        check(value_weight >= label_weight,
              f"Profile: the value is at least as bold as the name (value {value_weight} vs "
              f"name {label_weight})")

    # ---- 2B-R2-6: identity -- name-as-link, no separate publication link ---
    id_box = page.locator(".st-key-profile")
    name_link = id_box.locator("h3 a").first
    check(name_link.count() >= 1, "Identity: the institution NAME renders as a link")
    name_href = name_link.get_attribute("href") or ""
    check("openalex.org/works" in name_href,
          f"Identity: the institution name links to its own OpenAlex works ({name_href!r})")
    id_text = id_box.inner_text()
    check(NO_LONGER_ON_FIND not in id_text,
          f"Identity (2B-R2-6): the separate {NO_LONGER_ON_FIND!r} link is gone")
    check("ROR" in id_text, "Identity: the ROR link label still renders")

    # ---- the six panels, exact labels --------------------------------------
    for name, label in PANEL_LABELS:
        summary = page.locator(f".st-key-panel_{name} summary").first
        check(summary.count() >= 1, f"Panel '{name}': expander present (.st-key-panel_{name})")
        raw = (summary.text_content() or "").strip()
        clean = raw.replace("keyboard_arrow_right", "").replace("keyboard_arrow_down", "").strip()
        check(clean == label, f"Panel '{name}': header label is exactly {label!r} (got {raw!r})")

    # ---- top subfields: no sort control, cut at SUBFIELDS_TOP_N -----------
    _ensure_expander_open(page, "panel_subfields", ".st-key-fig_subfields")
    _settle(page, 1500)
    fig = page.locator(".st-key-fig_subfields .js-plotly-plot").first
    check(fig.count() >= 1 and fig.is_visible(),
          "Panel Top subfields: opening it reveals a visible Plotly figure")
    check(page.locator(".st-key-panel_subfields .st-key-sort_subfields").count() == 0,
          "Panel Top subfields: carries NO sort control (2B-R-13)")
    sf_ticks = page.locator(".st-key-fig_subfields .ytick")
    n_sf = sf_ticks.count()
    check(0 < n_sf <= SUBFIELDS_TOP_N,
          f"Panel Top subfields: {n_sf} y-tick group(s), within (0, {SUBFIELDS_TOP_N}]")

    # ---- 2B-R-13 (NEW): Top topics -- cut at 30, no sort control ----------
    _ensure_expander_open(page, "panel_topics", ".st-key-fig_topics")
    _settle(page, 1500)
    check(page.locator(".st-key-panel_topics .st-key-sort_topics").count() == 0,
          "Panel Top topics: carries NO sort control (2B-R-13)")
    tp_ticks = page.locator(".st-key-fig_topics .ytick")
    n_tp = tp_ticks.count()
    check(0 < n_tp <= TOPICS_TOP_N,
          f"Panel Top topics: {n_tp} y-tick group(s), within (0, {TOPICS_TOP_N}] (2B-R-13 raised 20->30)")

    # ---- SDG numbered labels --------------------------------------
    _ensure_expander_open(page, "panel_sdg", ".st-key-fig_sdg")
    _settle(page, 1500)
    sdg_ticks = page.locator(".st-key-fig_sdg .ytick")
    n_sdg = sdg_ticks.count()
    sdg_texts = [sdg_ticks.nth(i).text_content() or "" for i in range(n_sdg)]
    non_sdg = [t for t in sdg_texts if not t.strip().startswith("SDG")]
    check(n_sdg > 0 and not non_sdg,
          f"Panel SDG profile: all {n_sdg} y-tick labels start with 'SDG' (offenders: {non_sdg})")

    # ---- frontier panel, two modes --------------------------------
    _ensure_expander_open(page, "panel_frontier", ".st-key-frontier_mode button")
    _settle(page, 1500)
    top_points = _frontier_points(page)
    top_sig = _frontier_signature(page)
    check(top_points > 0, f"Panel Frontier positioning: default mode plots points ({top_points})")
    page.locator(".st-key-frontier_mode button").nth(FRONTIER_MODE_EMERGING_IDX).click(
        timeout=ACTION_TIMEOUT_MS)
    _settle(page, 4000)
    emerging_points = _frontier_points(page)
    emerging_sig = _frontier_signature(page)
    # 2B-R-13: both modes share ONE top-N slider, so the plotted COUNT can tie
    # (both cut at the same top_n) even though the topic SET differs -- the
    # figure's own data is the honest signal, the count is reported alongside
    # it for information only.
    check(emerging_points > 0 and emerging_sig != top_sig,
          f"Panel Frontier positioning: the mode control changes the plotted topic set "
          f"(counts {top_points} -> {emerging_points}; data signature changed: "
          f"{emerging_sig != top_sig})")
    # Deliberately LEFT on "emerging": the persistence checks assert this
    # exact value survives later Menu<->Find hops.

    # ---- the breakdown pair's shared segmented control ------------
    before_legend = _chip_legend(page)
    check(bool(before_legend.strip()), "Breakdown: chip legend renders")
    page.locator(".st-key-breakdown_dim button").nth(BREAKDOWN_DOCTYPE_IDX).click(
        timeout=ACTION_TIMEOUT_MS)
    _settle(page, 4000)
    after_legend = _chip_legend(page)
    check(after_legend != before_legend and bool(after_legend.strip()),
          "Breakdown: segmented control swaps the chip legend (domain <-> document type)")
    check(page.locator(".st-key-fig_breakdown_global .js-plotly-plot").first.is_visible()
          and page.locator(".st-key-fig_breakdown_yearly .js-plotly-plot").first.is_visible(),
          "Breakdown: both plotly figures still render after the swap")
    # Deliberately LEFT on "Document type".

    _no_exception(page, "Profile / panels")
    return {"frontier_points": emerging_points, "frontier_signature": emerging_sig,
            "breakdown_legend": after_legend}


def check_bonus_year_axis(page) -> None:
    """2B-R-2: no bonus-year banner/caption -- the bonus year is starred ON
    the yearly breakdown's own x-axis ("2025*"), read off the live figure's
    own data, never a caption."""
    try:
        labels = _yearly_axis_labels(page)
        check(BONUS_YEAR_AXIS_LABEL in labels,
              f"Breakdown yearly axis: the bonus year is starred "
              f"({BONUS_YEAR_AXIS_LABEL!r} in {labels})")
    except Exception as exc:
        fail_section("Bonus year axis", exc)


def check_si_value_labels(page) -> None:
    """2B-R-13: the SI unit grid is retired for an outer-end value label on
    each row's own SI marker -- checked on the three panels that carry an SI
    column (fields, subfields, ERC): at least one non-empty text label on the
    SI trace, and `showgrid` false on the SI axis (no per-integer
    gridlines)."""
    for name, fig_key, probe in (
            ("fields", "fig_fields", ".st-key-fig_fields"),
            ("subfields", "fig_subfields", ".st-key-fig_subfields"),
            ("erc", "fig_erc", ".st-key-sort_erc [data-testid='stRadioOption']")):
        try:
            _ensure_expander_open(page, f"panel_{name}", probe)
            _settle(page, 1200)
            info = _si_label_info(page, fig_key)
            check(info.get("n_labels", -1) > 0,
                  f"Panel '{name}': the SI marker carries an outer-end value label "
                  f"({info.get('n_labels')} label(s))")
            check(info.get("showgrid") is False,
                  f"Panel '{name}': the retired per-integer SI unit grid stays off "
                  f"(showgrid={info.get('showgrid')!r})")
        except Exception as exc:
            fail_section(f"SI value labels ({name})", exc)


def check_frontier_slider_modes(page, expect: dict) -> None:
    """2B-R-13: the frontier panel's ONE top-N slider actually re-cuts the
    plotted set in BOTH modes. Runs right after check_profile_and_panels
    (which left the panel on "emerging" mode) and restores EXACTLY that state
    (mode=emerging, slider=default) before returning, since the persistence
    checks downstream compare against `expect["frontier_points"]`."""
    try:
        _ensure_expander_open(page, "panel_frontier", ".st-key-frontier_mode button")
        before = _frontier_points(page)
        _frontier_slider_step(page, -1)
        _settle(page, 2500)
        after = _frontier_points(page)
        check(before > 0 and after > 0 and after != before,
              f"Frontier slider (Emerging mode): moving it changes the plotted point count "
              f"({before} -> {after})")
        _frontier_slider_step(page, 1)
        _settle(page, 2500)
        restored = _frontier_points(page)
        check(restored == before,
              f"Frontier slider (Emerging mode): moving it back restores the point count "
              f"({restored} vs {before})")

        page.locator(".st-key-frontier_mode button").nth(FRONTIER_MODE_TOP_IDX).click(
            timeout=ACTION_TIMEOUT_MS)
        _settle(page, 3000)
        top_before = _frontier_points(page)
        _frontier_slider_step(page, -1)
        _settle(page, 2500)
        top_after = _frontier_points(page)
        check(top_before > 0 and top_after > 0 and top_after != top_before,
              f"Frontier slider (Top mode): moving it changes the plotted point count "
              f"({top_before} -> {top_after})")
        _frontier_slider_step(page, 1)
        _settle(page, 2500)

        # Back to "emerging", matching the state check_profile_and_panels left.
        page.locator(".st-key-frontier_mode button").nth(FRONTIER_MODE_EMERGING_IDX).click(
            timeout=ACTION_TIMEOUT_MS)
        _settle(page, 3000)
        final = _frontier_points(page)
        final_sig = _frontier_signature(page)
        check(final == expect.get("frontier_points") and final_sig == expect.get("frontier_signature"),
              f"Frontier slider: panel restored to the emerging-mode default state "
              f"(points expected {expect.get('frontier_points')}, got {final}; "
              f"signature match: {final_sig == expect.get('frontier_signature')})")
    except Exception as exc:
        fail_section("Frontier slider (both modes)", exc)


def check_tab_overflow_a11(page) -> None:
    """A11: with BOTH optional lenses on, the tab strip fits at 1280px with no
    silent Streamlit-tab scroll, and every tab carries only its bare display
    code (L0..L9). Toggles are switched on, measured, then switched back off
    so downstream sections see the coded defaults again."""
    try:
        page.locator(".st-key-c1_on label").first.click(timeout=ACTION_TIMEOUT_MS)
        page.locator(".st-key-l7_on label").first.click(timeout=ACTION_TIMEOUT_MS)
        _wait_for(page, lambda: page.locator('[role="tab"]').count() == BOTH_OPTIONAL_TAB_COUNT)
        _settle(page, 800)
        tabs_n = page.locator('[role="tab"]').count()
        check(tabs_n == BOTH_OPTIONAL_TAB_COUNT,
              f"A11: with both optional lenses on, tab count is {BOTH_OPTIONAL_TAB_COUNT} "
              f"(got {tabs_n})")
        # Measured (debug probe, 2026-08-31): `[data-testid="stTabs"]` itself
        # (the outer wrapper) carries a few px of its OWN padding/border and
        # is never the scrollable element -- `[role="tablist"]` (BaseWeb's
        # `[data-baseweb="tab-list"]` is absent on this pinned build) is the
        # element that actually scrolls, and is what FC's own 820==820
        # measurement (progress/2BR_FC.md) reads.
        info = page.evaluate(
            """(() => {
                const el = document.querySelector('[role="tablist"]')
                        || document.querySelector('[data-testid="stTabs"]');
                if (!el) return null;
                return {scroll: el.scrollWidth, client: el.clientWidth};
            })()""")
        check(info is not None and info["scroll"] <= info["client"] + 2,
              f"A11: the tab strip fits with no silent scroll at 1280px "
              f"(scrollWidth {info and info.get('scroll')} <= clientWidth {info and info.get('client')})")
        tab_text = _all_text(page, '[role="tab"]')
        for code in ("L0", "L1", "L2", "L3", "L4", "L5", "L6", "L7", "L8", "L9"):
            check(code in tab_text, f"A11: tab strip carries the bare display code {code!r}")
    except Exception as exc:
        fail_section("A11 tab overflow", exc)
    finally:
        try:
            page.locator(".st-key-c1_on label").first.click(timeout=ACTION_TIMEOUT_MS)
            page.locator(".st-key-l7_on label").first.click(timeout=ACTION_TIMEOUT_MS)
            _wait_for(page, lambda: page.locator('[role="tab"]').count() == GDANSK_TAB_COUNT)
            _settle(page, 800)
        except Exception:  # noqa: BLE001 -- best-effort restore
            pass


def check_benchmark_lens_guide(page) -> None:
    """The "How to read the lenses" expander at the head of the Benchmark
    section, and A11's bare-code tab + full-name-inside-body split."""
    _ensure_expander_open(page, "lens_guide", ".st-key-lens_guide strong")
    summary = page.locator(".st-key-lens_guide summary").first
    raw = (summary.text_content() or "").strip()
    clean = raw.replace("keyboard_arrow_right", "").replace("keyboard_arrow_down", "").strip()
    check(clean == LENS_GUIDE_HEADER,
          f"Lens guide: header label is exactly {LENS_GUIDE_HEADER!r} (got {raw!r})")
    n_lines = page.locator(".st-key-lens_guide strong").count()
    check(n_lines >= 8, f"Lens guide: at least 8 lens lines render (found {n_lines})")

    tabs = page.locator('[role="tab"]')
    first_lens_tab_text = (tabs.nth(1).text_content() or "").strip()
    check(first_lens_tab_text == LENS0_TAB_CODE,
          f"A11: the first default-lens TAB carries only the bare code "
          f"(expected {LENS0_TAB_CODE!r}, got {first_lens_tab_text!r})")
    tabs.nth(1).click(timeout=ACTION_TIMEOUT_MS)
    _settle(page, 1500)
    body_text = _all_text(page, '[role="tabpanel"]')
    check(LENS0_FULL_NAME in body_text,
          f"A11: the full lens name opens the tab BODY (looking for {LENS0_FULL_NAME!r})")
    tabs.nth(0).click(timeout=ACTION_TIMEOUT_MS)
    _settle(page, 1000)

    caption = _all_text(page, '[data-testid="stCaptionContainer"]')
    check(LENS_LEGEND_SUBSTR in caption,
          f"Overview: the legend caption points at the lens guide (looking for {LENS_LEGEND_SUBSTR!r})")
    _no_exception(page, "Benchmark lens guide")


# ----------------------------------------------------------- R1 tables -----

def _download_csv_header(page, click_selector: str) -> str:
    with page.expect_download(timeout=ACTION_TIMEOUT_MS) as dl_info:
        page.locator(click_selector).click(timeout=ACTION_TIMEOUT_MS)
    download = dl_info.value
    path = download.path()
    with open(path, "r", encoding="utf-8") as fh:
        return fh.readline()


def check_tables_and_export(page) -> None:
    """A lens's ranked table renders; its CSV export carries
    `total_frac_2020_2024`, `country` and `evidence` but never a `badge`
    column. The Aspirational tab renders its own table, with no "Interval"
    column (2B-R-11)."""
    tabs = page.locator('[role="tab"]')
    tabs.nth(1).click(timeout=ACTION_TIMEOUT_MS)  # first default lens tab (L0)
    _settle(page, 2000)
    check(page.locator('.st-key-tbl_L0 [data-testid="stDataFrame"]').count() >= 1,
          "Lens table: L0's ranked table renders (.st-key-tbl_L0)")

    header = _download_csv_header(page, ".st-key-dl_L0 button")
    check("total_frac_2020_2024" in header, f"CSV export: header carries total_frac_2020_2024 ({header!r})")
    check("country" in header, f"CSV export: header carries country ({header!r})")
    check("evidence" in header, f"CSV export: header carries evidence ({header!r})")
    check("badge" not in header, f"CSV export: header carries NO badge column ({header!r})")

    tabs.last.click(timeout=ACTION_TIMEOUT_MS)  # Aspirational
    _settle(page, 2500)
    check(page.locator('.st-key-tbl_aspirational [data-testid="stDataFrame"]').count() >= 1,
          "Aspirational tab: its own table renders (.st-key-tbl_aspirational)")
    header_text = _all_text(page, '.st-key-tbl_aspirational [data-testid="stColumnHeader"], '
                                  '.st-key-tbl_aspirational [role="columnheader"]')
    check("Interval" not in header_text,
          f"Aspirational tab (2B-R-11): no 'Interval' column (headers seen: {header_text[:200]!r})")
    tabs.nth(0).click(timeout=ACTION_TIMEOUT_MS)  # back to Overview
    _settle(page, 1500)
    _no_exception(page, "Tables / export")


def check_institution_link_popup(page) -> None:
    """A10: the institution NAME is the clickable OpenAlex-works link (URL
    fragment trick) -- proven with a REAL click and a captured popup, since a
    canvas grid carries no DOM <a> to query. Tried at a few plausible
    (column, row) pixel offsets since the exact grid geometry is not
    otherwise exposed."""
    try:
        tabs = page.locator('[role="tab"]')
        tabs.last.click(timeout=ACTION_TIMEOUT_MS)  # Aspirational
        _settle(page, 2500)
        grid = page.locator('.st-key-tbl_aspirational [data-testid="stDataFrame"] canvas').first
        box = grid.bounding_box()
        check(box is not None, "Institution link: the Aspirational grid has a real bounding box")
        opened = False
        url = None
        if box is not None:
            for frac_x in (0.14, 0.20, 0.28):
                x = box["x"] + box["width"] * frac_x
                y = box["y"] + 49   # ~header height + half a data row
                try:
                    with page.context.expect_page(timeout=4000) as pop_info:
                        page.mouse.click(x, y)
                    popup = pop_info.value
                    popup.wait_for_load_state("domcontentloaded", timeout=ACTION_TIMEOUT_MS)
                    url = popup.url
                    popup.close()
                    opened = True
                    break
                except PWTimeoutError:
                    continue
        check(opened, "Institution link (A10): a real click on the Institution cell opens a popup")
        if opened:
            check("openalex.org/works" in (url or ""),
                  f"Institution link (A10): the popup opens an OpenAlex-works URL (got {url!r})")
        tabs.nth(0).click(timeout=ACTION_TIMEOUT_MS)  # back to Overview
        _settle(page, 1200)
    except Exception as exc:
        fail_section("Institution link popup", exc)


# ------------------------------------------------------------- settings ----

def check_settings(page) -> None:
    """The settings a reader would touch on a first visit -- depth to max,
    L7 on, a type filter picked, tree switched to its non-default DISPLAY
    label. `frontier_mode` and `breakdown_dim` are already off-default and
    are left untouched here."""
    before = _all_text(page, '[data-testid="stCaptionContainer"]')

    _ensure_expander_open(page, "postfilters", ".st-key-f_types input")
    tinp = page.locator(".st-key-f_types input").first
    tinp.click(timeout=ACTION_TIMEOUT_MS)
    tinp.fill("education")
    page.wait_for_selector('[role="option"]', timeout=ACTION_TIMEOUT_MS)
    page.locator('[role="option"]').filter(has_text="education").first.click(timeout=ACTION_TIMEOUT_MS)
    _settle(page, 3500)

    page.locator('.st-key-depth [data-testid="stRadioOption"]').last.click(timeout=ACTION_TIMEOUT_MS)
    _settle(page, 3000)
    after = _all_text(page, '[data-testid="stCaptionContainer"]')
    check(before != after, "Settings: depth caption changed after switching depth to its max")

    _open_select(page, "tree")
    _pick_option(page, TREE_LABEL_ORIGINAL)
    _wait_for(page, lambda: page.locator('[role="tab"]').count() >= GDANSK_TAB_COUNT)
    _settle(page, 1000)

    page.locator(".st-key-l7_on label").first.click(timeout=ACTION_TIMEOUT_MS)
    _wait_for(page, lambda: page.locator('[role="tab"]').count() == L7_ON_TAB_COUNT)
    _settle(page, 800)
    tabs = page.locator('[role="tab"]').count()
    check(tabs == L7_ON_TAB_COUNT, f"Settings: L7 tab appeared, tab count is {L7_ON_TAB_COUNT} (got {tabs})")

    check(page.locator(".st-key-strip").count() >= 1, "Settings: off-default strip is visible")
    strip = _strip_text(page)
    check(STRIP_TREE_ORIGINAL in strip,
          f"Settings: strip shows the taxonomy's DISPLAY label (looking for {STRIP_TREE_ORIGINAL!r} "
          f"in {strip!r})")
    check("original" not in strip.replace(TREE_LABEL_ORIGINAL, ""),
          "Settings: the internal taxonomy value never appears in the strip")
    check("depth = 50" in strip, f"Settings: strip mentions depth = 50 (strip: {strip!r})")
    check("type: " in strip and "education" in strip,
          f"Settings: strip mentions the type filter (strip: {strip!r})")
    _no_exception(page, "Settings")


def _capture_persisted_state(page) -> dict:
    _ensure_expander_open(page, "panel_frontier", ".st-key-frontier_mode button")
    return {"basket": _basket_count(page), "tabs": page.locator('[role="tab"]').count(),
            "heading": _seed_heading(page), "strip": _strip_text(page),
            # 2B-R-13: BOTH modes share one top-N slider, so the raw point
            # COUNT can tie between modes (see `_frontier_signature`'s own
            # docstring) -- the topic-set SIGNATURE is the only proof that
            # actually distinguishes "still on emerging" from "reset to top".
            "frontier_points": _frontier_points(page),
            "frontier_signature": _frontier_signature(page),
            "breakdown_legend": _chip_legend(page)}


def _assert_persisted(state: dict, tag: str, expect: dict) -> None:
    check(state["basket"] == 2, f"{tag}: basket still lists 2 items (got {state['basket']})")
    check(state["tabs"] == L7_ON_TAB_COUNT,
          f"{tag}: L7 tab still present, tab count {L7_ON_TAB_COUNT} (got {state['tabs']})")
    check("Gda" in state["heading"], f"{tag}: seed still selected, heading 'Gda...' (got {state['heading']!r})")
    check(STRIP_TREE_ORIGINAL in state["strip"], f"{tag}: taxonomy's display label still in the strip")
    check("depth = 50" in state["strip"], f"{tag}: depth still at max in the strip")
    check("type: " in state["strip"] and "education" in state["strip"],
          f"{tag}: type filter (education) still active in the strip")
    fs_expected = expect.get("frontier_signature")
    check(bool(fs_expected) and state["frontier_signature"] == fs_expected,
          f"{tag}: frontier_mode still shows its off-default (emerging) topic set "
          f"(signature match: {state.get('frontier_signature') == fs_expected}; "
          f"points {state['frontier_points']} vs baseline {expect.get('frontier_points')})")
    bl_expected = expect.get("breakdown_legend")
    check(bool(bl_expected) and state["breakdown_legend"] == bl_expected,
          f"{tag}: breakdown_dim still shows the swapped (document-type) chip legend")


def check_persistence(page, expect: dict) -> None:
    """The load-bearing claim: basket + every keyed widget survive real
    Menu<->Find hops (4 hops total), with a second-visit re-mount check at
    the 2-hop midpoint."""
    _assert_persisted(_capture_persisted_state(page), "Persistence: baseline captured before any hop", expect)

    _click_nav(page, "Menu")
    _no_exception(page, "Menu (hop 1 of 4)")
    _click_nav(page, "Find")
    _no_exception(page, "Find (hop 2 of 4, second-visit re-mount)")
    _assert_persisted(_capture_persisted_state(page), "Persistence: 2nd Find visit (re-mount check)", expect)

    _click_nav(page, "Menu")
    _no_exception(page, "Menu (hop 3 of 4)")
    _click_nav(page, "Find")
    _no_exception(page, "Find (hop 4 of 4, final)")
    _assert_persisted(_capture_persisted_state(page), "Persistence: 3rd Find visit (after 4 hops)", expect)


def check_type_filter_clear(page) -> None:
    strip = _strip_text(page)
    check("type: " in strip and "education" in strip,
          f"Type filter: still active going into the clear check (strip: {strip!r})")
    _ensure_expander_open(page, "postfilters", ".st-key-f_types input")

    tag_close = page.locator(".st-key-f_types [data-baseweb='tag'] [role='button'], "
                              ".st-key-f_types [data-baseweb='tag'] svg")
    if tag_close.count():
        tag_close.first.click(timeout=ACTION_TIMEOUT_MS)
    else:
        page.locator(".st-key-f_types input").first.click(timeout=ACTION_TIMEOUT_MS)
        page.keyboard.press("Backspace")
    _settle(page, 3500)
    strip2 = _strip_text(page)
    check("education" not in strip2, f"Type filter: strip no longer names it after clearing (strip: {strip2!r})")
    _no_exception(page, "Type filter cleared")


def check_undefined_lens(page, seed_id: str, seed_name: str) -> None:
    """2B-R-11a renumbered L2f's display code to L4 -- the tab now carries
    only that bare code and the undefined message reads the full DISPLAY name
    (`L4 . Shared specialisations`), never the literal internal id."""
    _search_and_pick(page, seed_name)
    heading = _seed_heading(page)
    check(len(heading) > 0, f"Undefined lens: seed '{seed_name}' ({seed_id}) loaded, heading present")
    tab = page.locator('[role="tab"]').filter(has_text=L2F_TAB_CODE).first
    check(tab.count() >= 1, f"Undefined lens: {L2F_TAB_CODE!r} (L2f) tab present")
    tab.click(timeout=ACTION_TIMEOUT_MS)
    _settle(page, 1500)
    text = _all_text(page, '[role="tabpanel"]')
    check(L2F_DISPLAY_NAME in text and "cannot be computed for this seed" in text,
          f"Undefined lens: the L2f undefined message present for {seed_name} "
          f"(looking for {L2F_DISPLAY_NAME!r} in {text[:300]!r})")
    _no_exception(page, "Undefined L2f seed")


def check_subfields_panel_no_overlap(page, width: int) -> None:
    """Bounding-box proof that opening the TOP-SUBFIELDS panel at this width
    never lets a y-axis tick label collide with anything."""
    fig = page.locator(".st-key-fig_subfields .js-plotly-plot").first
    fig.wait_for(state="visible", timeout=ACTION_TIMEOUT_MS)
    plot_box = fig.locator(".main-svg").first.bounding_box()
    if plot_box is None:
        check(False, f"Top-subfields panel {width}px: could not read the plot's own .main-svg bounding box")
        return
    ticks = fig.locator(".ytick")
    n = ticks.count()
    plot_left = plot_box["x"]
    plot_right = plot_box["x"] + plot_box["width"]
    offenders = []
    for i in range(n):
        box = ticks.nth(i).bounding_box()
        if box is None:
            continue
        text = ticks.nth(i).text_content() or ""
        if box["x"] < plot_left - 1:
            offenders.append(f"{text!r} clipped at left (x={box['x']:.1f} < plot left {plot_left:.1f})")
        elif box["x"] + box["width"] > plot_right + 1:
            offenders.append(f"{text!r} overflows right "
                             f"(right={box['x'] + box['width']:.1f} > plot right {plot_right:.1f})")
    check(n > 0 and not offenders,
          f"Top-subfields panel {width}px: {n} y-tick group(s) all stay inside the plot's own svg"
          + (f" -- offenders: {offenders}" if offenders else ""))


def check_screenshots(browser, shot_dir: Path) -> None:
    """At each width, the seed is loaded AND the Top-subfields panel is opened
    before the scrollWidth assertion."""
    shot_dir.mkdir(parents=True, exist_ok=True)
    for width in WIDTHS:
        page = browser.new_page(viewport={"width": width, "height": 900})
        page.set_default_timeout(ACTION_TIMEOUT_MS)
        try:
            page.goto(BASE_URL, wait_until="domcontentloaded")
            page.wait_for_selector('[data-testid="stSidebarNav"]', state="attached",
                                    timeout=ACTION_TIMEOUT_MS)
            _settle(page, 1500)
            scroll = page.evaluate("document.documentElement.scrollWidth")
            inner = page.evaluate("window.innerWidth")
            check(scroll <= inner + 2, f"Menu {width}px: scrollWidth {scroll} <= innerWidth+2 {inner + 2}")
            p = shot_dir / f"smoke_menu_{width}.png"
            page.screenshot(path=str(p), full_page=True)
            check(p.is_file(), f"Menu {width}px: screenshot written ({p.name})")

            _click_nav(page, "Find")
            _search_and_pick(page, GDANSK_QUERY)

            if width == 1280:
                page.evaluate("window.scrollTo(0, 0)")
                _settle(page, 500)
                top_p = shot_dir / "smoke_find_top_1280.png"
                page.screenshot(path=str(top_p), full_page=False)
                check(top_p.is_file(), f"Find top-of-page 1280px: screenshot written ({top_p.name})")

            _ensure_expander_open(page, "panel_subfields", ".st-key-fig_subfields")
            _settle(page, 1500)
            scroll = page.evaluate("document.documentElement.scrollWidth")
            inner = page.evaluate("window.innerWidth")
            check(scroll <= inner + 2, f"Find {width}px: scrollWidth {scroll} <= innerWidth+2 {inner + 2}")
            p2 = shot_dir / f"smoke_find_{width}.png"
            page.screenshot(path=str(p2), full_page=True)
            check(p2.is_file(), f"Find {width}px: screenshot written ({p2.name})")

            if width in (390, 1280):
                check_subfields_panel_no_overlap(page, width)
            if width == 390:
                subfields_p = shot_dir / "smoke_find_subfields_390.png"
                page.screenshot(path=str(subfields_p), full_page=True)
                check(subfields_p.is_file(),
                      f"Find Top-subfields panel 390px: screenshot written ({subfields_p.name})")
        except Exception as exc:  # noqa: BLE001 -- one width's failure must not skip the rest
            fail_section(f"Screenshots at {width}px", exc)
        finally:
            page.close()


# ------------------------------------------------ Phase 2B/2B-R: the journey --

def _n_plotly(page) -> int:
    return page.locator(".js-plotly-plot").count()


def _settle_figures(page, target: int, timeout_ms: int = 60_000) -> int:
    """Streamlit streams elements in, so a figure count climbs for a while
    after the first plot appears. Poll until the count reaches the floor and
    holds for 3 checks running, rather than a blind sleep."""
    deadline = time.time() + timeout_ms / 1000
    last, stable = -1, 0
    while time.time() < deadline:
        now = _n_plotly(page)
        stable = stable + 1 if now == last and now >= target else 0
        last = now
        if stable >= 3:
            break
        page.wait_for_timeout(800)
    page.wait_for_timeout(1000)
    return last


def _sidebar_basket_n(page) -> int | None:
    """The `{n} of {cap} added` sidebar caption, read wherever it renders:
    Find's own editable basket AND Compare/Collaborate's read-only mirror
    share the exact same template."""
    text = _all_text(page, '[data-testid="stSidebar"] [data-testid="stCaptionContainer"]')
    m = re.search(r"(\d+) of \d+ added", text)
    return int(m.group(1)) if m else None


def _add_l1_candidates(page) -> list[dict]:
    """Downloads the seed's own L1 (subfield-overlap) ranking CSV and returns
    the top 3 rows' `{institution_id, display_name}`."""
    tab = page.locator('[role="tab"]').filter(has_text="L1").first
    tab.click(timeout=ACTION_TIMEOUT_MS)
    _settle(page, 1500)
    with page.expect_download(timeout=ACTION_TIMEOUT_MS) as dl_info:
        page.locator(".st-key-dl_L1 button").first.click(timeout=ACTION_TIMEOUT_MS)
    path = dl_info.value.path()
    with open(path, "r", encoding="utf-8", newline="") as fh:
        rows = list(csv.DictReader(fh))
    return [{"institution_id": r["institution_id"], "display_name": r["display_name"]}
            for r in rows[:3] if r.get("institution_id")]


def check_journey_basket(page) -> list[dict]:
    """Search Gdansk, add 3 candidates off the L1 table plus the seed itself
    (basket = 4) -- COMPARE_CAP=3 means Compare will need to truncate."""
    clear_btn = page.locator(".st-key-basket_clear button")
    if clear_btn.count():
        clear_btn.first.click(timeout=ACTION_TIMEOUT_MS)
        _settle(page, 1500)
    _search_and_pick(page, GDANSK_QUERY)
    heading = _seed_heading(page)
    check("Gda" in heading,
          f"Journey: seed re-loaded to Gdansk before building the basket (got {heading!r})")
    candidates = _add_l1_candidates(page)
    check(len(candidates) == 3, f"Journey: read 3 candidates off the L1 CSV (got {len(candidates)})")
    for row in candidates:
        _add_comparator(page, row["display_name"])
    _add_comparator(page, GDANSK_QUERY)  # the seed itself, same sidebar add box
    n = _basket_count(page)
    check(n == 4, f"Journey: basket holds 4 (3 L1 candidates + the seed itself), got {n}")
    _no_exception(page, "Journey basket (L1 candidates + seed)")
    return candidates


def _compare_deeplink_ids(page) -> list[str]:
    loc = page.locator('[data-testid="stCode"]').filter(has_text="?compare=").first
    if loc.count() == 0:
        return []
    text = loc.text_content() or ""
    if "?compare=" not in text:
        return []
    return text.split("?compare=", 1)[1].strip().split(",")


def check_compare_journey(page, candidates: list[dict]) -> dict:
    """2B-R-4/5/6/7/8/9: the cap-3 truncation notice, the overview cards
    (incl. international/company), the metric selectors, the two frontier
    charts and the legend-above-every-chart rule."""
    _click_nav(page, NAV_COMPARE)
    _settle_figures(page, COMPARE_MIN_FIGURES)
    _no_exception(page, "Compare (initial render)")

    names = [c["display_name"] for c in candidates]
    full_text = _full_page_text(page)
    strip = _all_text(page, ".st-key-compare_strip")
    check(any(n in strip for n in names) or "Gda" in strip,
          f"Compare: the overview strip names at least one of the basket's institutions "
          f"(strip[:200]={strip[:200]!r})")

    # 2B-R-4: cap-3 truncation. The 4-item basket must be disclosed AND cut.
    check(CAP_TRUNCATED_SUBSTR in full_text,
          "Compare (2B-R-4): the cap-3 truncation notice renders for a 4-item basket")
    ids3 = _compare_deeplink_ids(page)
    check(len(ids3) == COMPARE_CAP,
          f"Compare (2B-R-4): the deep link is truncated to {COMPARE_CAP} ids (got {len(ids3)}: {ids3})")

    # 2B-R2-9: overview cards (markup now, not st.metric) carry all six KPI
    # facts incl. intl/industrial, a best-value dot, and the name-as-link.
    strip_text = _all_text(page, ".st-key-compare_strip")
    for label in CARD_LABELS:
        check(label in strip_text, f"Compare overview: the {label!r} card renders (2B-R2-9)")
    strip_html = page.evaluate(
        "(() => { const e = document.querySelector('.st-key-compare_strip');"
        " return e ? e.innerHTML : ''; })()")
    check("benchup-kpi" in strip_html,
          "Compare overview: cards render as markup, reusing the Find card class (2B-R2-9)")
    dot_colors = page.evaluate(
        "Array.from(document.querySelectorAll('.st-key-compare_strip span'))"
        ".map(e => getComputedStyle(e).backgroundColor)"
        ".filter(c => c && c !== 'rgba(0, 0, 0, 0)' && c !== 'transparent')")
    check(len(dot_colors) >= 1,
          f"Compare overview (2B-R2-9): at least one best-value dot is painted ({len(dot_colors)})")
    name_hrefs = page.evaluate(
        "Array.from(document.querySelectorAll('.st-key-compare_strip a')).map(a => a.href)")
    check(any("openalex.org/works" in h for h in name_hrefs),
          "Compare overview: an institution name links to its own OpenAlex works")
    strip_buttons = _all_text(page, ".st-key-compare_strip button")
    check("Publications" not in strip_buttons,
          "Compare overview (2B-R2-9): the separate Publications button is gone")

    # 2B-R-12: a legend strip above every chart section.
    legend_hits = page.evaluate(
        "(names) => Array.from(document.querySelectorAll('[data-testid=\"stMarkdownContainer\"]'))"
        ".filter(e => names.every(n => e.textContent.includes(n))"
        " && e.querySelectorAll('span').length >= names.length * 2).length",
        names[:1] + [candidates[0]["display_name"]] if candidates else [])
    # A minimal generic probe (any strip carrying >=2 swatches) is more robust
    # than requiring every institution's own name in each one, since some
    # sections' compared SET differs after the cap-3 truncation.
    generic_legend_hits = page.evaluate(
        "() => Array.from(document.querySelectorAll('[data-testid=\"stMarkdownContainer\"]'))"
        ".filter(e => e.querySelectorAll('span').length >= 2 "
        "&& e.innerHTML.includes('color:')).length")
    check(generic_legend_hits >= MIN_LEGEND_STRIPS,
          f"Compare (2B-R-12): at least {MIN_LEGEND_STRIPS} legend strips render above their charts "
          f"(found {generic_legend_hits})")

    n_figs = _n_plotly(page)
    check(n_figs >= COMPARE_MIN_FIGURES,
          f"Compare: >= {COMPARE_MIN_FIGURES} plotly figures render ({n_figs})")

    # --- 2B-R2-3: the metric selector's own vocabulary, per level -----------
    def _opt_texts(key: str) -> list:
        return [t.strip() for t in
                page.locator(f'.st-key-{key} [data-testid="stRadioOption"]').all_text_contents()
                if t.strip()]

    def _click_opt(key: str, text: str) -> None:
        page.locator(f".st-key-{key}").get_by_text(text, exact=True).first.click(
            timeout=ACTION_TIMEOUT_MS)
        _settle(page, 2500)

    subj_opts, erc_opts, sdg_opts = (_opt_texts("cmp_metric_subject"), _opt_texts("cmp_metric_erc"),
                                     _opt_texts("cmp_metric_sdg"))
    for level, opts in (("subject", subj_opts), ("ERC", erc_opts), ("SDG", sdg_opts)):
        check(VOL_TOP10_LABEL not in opts,
              f"Compare {level} (2B-R2-3): the retired top-decile TAB is not offered ({opts})")
    check(set(SUBJECT_METRIC_LABELS) <= set(subj_opts),
          f"Compare subject (2B-R2-3): offers Share/Specialisation/PP/SDG/Dynamics ({subj_opts})")
    check(VOLUME_LABEL in erc_opts, f"Compare ERC (2B-R2-3): 'Volume' is offered ({erc_opts})")
    check(VOLUME_LABEL in sdg_opts, f"Compare SDG (2B-R2-3): 'Volume' is offered ({sdg_opts})")

    # THE 2B-R LESSON, reapplied: every option the SUBJECT selector offers is
    # actually clicked and its chart redrawn -- "option visible, render path
    # unreached" is the bug class that survived last round.
    for label in subj_opts:
        _click_opt("cmp_metric_subject", label)
        check(page.locator(".st-key-fig_cmp_subject .js-plotly-plot").count() >= 1,
              f"Compare subject = {label!r}: the chart renders")
    _no_exception(page, "Compare (after the full subject metric sweep)")

    # --- 2B-R2-5: row order IDENTICAL across two metric switches (load-     -
    #     bearing check) + the "sort by value" toggle really re-ranks --------
    def _tick_labels(key: str) -> list:
        return page.evaluate(
            "(sel) => { const p = document.querySelector(sel); if (!p) return [];"
            " return Array.from(p.querySelectorAll('.yaxislayer-above .ytick text'))"
            ".map(t => t.textContent); }",
            f".st-key-{key} .js-plotly-plot")

    def _row_names(labels) -> list:
        return [re.split(r"\d", str(t))[0].strip() for t in labels]

    _click_opt("cmp_metric_subject", "Share")
    labels_share = _tick_labels("fig_cmp_subject")
    rows_share = _row_names(labels_share)
    check(len(rows_share) > 5, f"Compare subject (Share): a full set of rows draws ({len(rows_share)})")
    with_numbers = [t for t in labels_share if re.search(r"\d", str(t))]
    check(len(with_numbers) >= len(labels_share) - 1,
          f"Compare subject (2B-R2-3): the volume gutter carries a number on every row "
          f"({len(with_numbers)} of {len(labels_share)})")
    n_refs_share = page.evaluate(
        "(() => { const p = document.querySelector('.st-key-fig_cmp_subject .js-plotly-plot');"
        " return p ? p.querySelectorAll('.shapelayer path').length : 0; })()")

    _click_opt("cmp_metric_subject", "Change in mean annual volume")
    rows_dyn = _row_names(_tick_labels("fig_cmp_subject"))
    common = [t for t in rows_share if t in set(rows_dyn)]
    check(len(common) > 5 and common == [t for t in rows_dyn if t in set(rows_share)],
          f"Compare (2B-R2-5, LOAD-BEARING): row order is IDENTICAL between Share and Dynamics "
          f"({len(common)} common rows)")
    n_refs_dyn = page.evaluate(
        "(() => { const p = document.querySelector('.st-key-fig_cmp_subject .js-plotly-plot');"
        " return p ? p.querySelectorAll('.shapelayer path').length : 0; })()")
    check(n_refs_dyn > n_refs_share,
          f"Compare (2B-R2-4): the Dynamics view draws its reference line, Share does not "
          f"({n_refs_dyn} vs {n_refs_share} shapes)")

    # --- 2B-R2-5: the "sort by value" toggle, tested on the SAME metric/view
    #     it was just captured on (Dynamics has the widest value spread, so
    #     it is the least likely metric to tie its way back to taxonomy order)
    if page.locator(".st-key-cmp_sort_subject").count() >= 1:
        _click_opt("cmp_sort_subject", SORT_VALUE_LABEL)
        ranked = _row_names(_tick_labels("fig_cmp_subject"))
        check(set(ranked) == set(rows_dyn) and ranked != rows_dyn,
              "Compare (2B-R2-5): the row-order toggle re-ranks the same rows by value")
        _click_opt("cmp_sort_subject", SORT_TAXONOMY_LABEL)
        back = _row_names(_tick_labels("fig_cmp_subject"))
        check(back == rows_dyn, "Compare (2B-R2-5): switching back restores the taxonomy order")

    # The dagger lives INSIDE the bar's own value-label text (e.g. "-16.4%†"),
    # a Plotly trace `text` entry -- not the y-axis tick label, which carries
    # only the row name and its volume-gutter numbers (measured: DOM ytick
    # text never contains the glyph even where a real low-volume row exists).
    def _bar_texts(key: str) -> list:
        data = _fig_xy_text(page, f".st-key-{key} .js-plotly-plot")
        return [t for tr in (data or []) for t in tr["text"]]

    dagger_hits = [LOW_VOLUME_GLYPH in str(t) for t in _bar_texts("fig_cmp_subject")]
    if not any(dagger_hits) and "Change in mean annual volume" in erc_opts:
        _click_opt("cmp_metric_erc", "Change in mean annual volume")
        dagger_hits += [LOW_VOLUME_GLYPH in str(t) for t in _bar_texts("fig_cmp_erc")]
    if not any(dagger_hits) and "Change in mean annual volume" in sdg_opts:
        _click_opt("cmp_metric_sdg", "Change in mean annual volume")
        dagger_hits += [LOW_VOLUME_GLYPH in str(t) for t in _bar_texts("fig_cmp_sdg")]
    check(any(dagger_hits),
          "Compare (2B-R2-4): a low-volume marker (dagger glyph) renders on at least one bar's "
          "own value label, searched across subject/ERC/SDG on the Dynamics view")

    _click_opt("cmp_metric_subject", "PP(top10%)")
    n_refs_pp = page.evaluate(
        "(() => { const p = document.querySelector('.st-key-fig_cmp_subject .js-plotly-plot');"
        " return p ? p.querySelectorAll('.shapelayer path').length : 0; })()")
    check(n_refs_pp > n_refs_share,
          f"Compare (2B-R2-4): the PP view draws its index reference too ({n_refs_pp} shapes)")

    _click_opt("cmp_metric_subject", "Share")
    _no_exception(page, "Compare (after the row-order / gutter / reference checks)")

    # --- 2B-R2-13: plain-language "not shown here" disclosures --------------
    caps = _all_text(page, '[data-testid="stCaptionContainer"]')
    check(NOT_OFFERED_HEADER in caps,
          f"Compare (2B-R2-13): the shared header {NOT_OFFERED_HEADER!r} renders")
    check(": " in caps.replace(NOT_OFFERED_HEADER, ""),
          "Compare: at least one 'Measure: reason' disclosure line renders in plain language")
    _no_forbidden_vocab(page, "Compare page")

    # --- 2B-R-9: the frontier map's own top-N slider ------------------------
    map_before = _plotly_point_count(page, ".st-key-fig_cmp_frontier_map .js-plotly-plot")
    try:
        map_slider = page.locator('.st-key-cmp_frontier_topn input[type="range"]').first
        map_slider.press("ArrowLeft", timeout=ACTION_TIMEOUT_MS)
        _settle(page, 2500)
        map_after = _plotly_point_count(page, ".st-key-fig_cmp_frontier_map .js-plotly-plot")
        check(map_before > 0 and map_after > 0 and map_after != map_before,
              f"Compare frontier map (2B-R-9): the top-N slider changes the plotted point count "
              f"({map_before} -> {map_after})")
        map_slider.press("ArrowRight", timeout=ACTION_TIMEOUT_MS)
        _settle(page, 2500)
    except Exception as exc:
        fail_section("Compare frontier map slider", exc)

    check(page.locator(".st-key-fig_cmp_shared_frontier .js-plotly-plot").count() >= 1,
          "Compare frontier (2B-R-9): the diverging 'who holds the shared frontier' chart renders")

    # --- 2B-R2-10: pool selector + domain-colour toggle, by TOPIC-SET
    #     SIGNATURE (never just point count -- a mode swap can tie on count) --
    def _frontier_map_sig() -> str:
        data = _fig_xy_text(page, ".st-key-fig_cmp_frontier_map .js-plotly-plot")
        return "" if not data else "|".join(
            ",".join(f"{v:.4f}" for v in tr["x"]) for tr in data)

    if page.locator(".st-key-cmp_frontier_pool").count() >= 1:
        sig_volume = _frontier_map_sig()
        page.locator(".st-key-cmp_frontier_pool").get_by_text(
            POOL_ELITE_LABEL, exact=True).first.click(timeout=ACTION_TIMEOUT_MS)
        _settle(page, 3000)
        sig_elite = _frontier_map_sig()
        check(bool(sig_volume) and bool(sig_elite) and sig_volume != sig_elite,
              "Compare frontier (2B-R2-10): the pool selector changes the plotted topic set")
        page.locator(".st-key-cmp_frontier_pool").get_by_text(
            POOL_VOLUME_LABEL, exact=True).first.click(timeout=ACTION_TIMEOUT_MS)
        _settle(page, 3000)

    def _frontier_map_legend() -> str:
        # The map's OWN legend strip (`map_legend_strip`) is the markdown block
        # immediately BEFORE the map's own keyed container -- scoped there
        # rather than to the whole page, since the diverging "who holds the
        # shared frontier" chart carries the SAME "held by more than one" chip
        # in its own, unrelated legend and would otherwise false-negative this
        # check regardless of the colour toggle.
        return page.evaluate(
            "(() => { const fig = document.querySelector('.st-key-fig_cmp_frontier_map');"
            " const prev = fig ? fig.previousElementSibling : null;"
            " return prev ? prev.textContent : ''; })()")

    if page.locator(".st-key-cmp_frontier_color").count() >= 1:
        check(LEGEND_SHARED_TEXT in _frontier_map_legend(),
              "Compare frontier: the map's own ownership legend carries the shared chip by default")
        page.locator(".st-key-cmp_frontier_color").get_by_text(
            COLOR_DOMAIN_LABEL, exact=True).first.click(timeout=ACTION_TIMEOUT_MS)
        _settle(page, 3000)
        domain_legend = _frontier_map_legend()
        check(LEGEND_SHARED_TEXT not in domain_legend and bool(domain_legend.strip()),
              f"Compare frontier (2B-R2-10): the domain-colour toggle replaces the map's own "
              f"ownership legend with the broad subject areas (got {domain_legend[:200]!r})")
        check(page.locator(".st-key-fig_cmp_frontier_map .js-plotly-plot").count() >= 1,
              "Compare frontier: the map still renders under the domain-colour toggle")
        page.locator(".st-key-cmp_frontier_color").get_by_text(
            COLOR_OWNER_LABEL, exact=True).first.click(timeout=ACTION_TIMEOUT_MS)
        _settle(page, 3000)
        check(LEGEND_SHARED_TEXT in _frontier_map_legend(),
              "Compare frontier: switching back restores the map's own ownership legend")

    # --- the impact floor toggle ----------------------------------------------
    # 2B-R2-8 moved this page's reading lines out of `st.caption` and into
    # markdown `chart_note` blocks -- the floor's own descriptive text (and
    # the cell counts it reports) live there now, not in a caption.
    before_md = _all_text(page, '[data-testid="stMarkdownContainer"]')
    page.locator('.st-key-cmp_impact_floor [data-testid="stRadioOption"]').last.click(
        timeout=ACTION_TIMEOUT_MS)
    _settle(page, 2500)
    after_md = _all_text(page, '[data-testid="stMarkdownContainer"]')
    check(before_md != after_md,
          "Compare (2B-R2-8): the impact floor toggle changes the page's reading lines")
    page.locator('.st-key-cmp_impact_floor [data-testid="stRadioOption"]').first.click(
        timeout=ACTION_TIMEOUT_MS)
    _settle(page, 2500)

    # --- the workbook (2B-R re-cut sheets) --------------------------------
    with page.expect_download(timeout=120_000) as dl_info:
        page.locator(".st-key-dl_workbook button").first.click(timeout=ACTION_TIMEOUT_MS)
    raw = Path(dl_info.value.path()).read_bytes()
    check(raw[:2] == b"PK", "Compare: the workbook downloads as a real xlsx container")
    book = openpyxl.load_workbook(io.BytesIO(raw))
    check(len(book.sheetnames) == XLSX_SHEET_COUNT,
          f"Compare: the workbook carries exactly {XLSX_SHEET_COUNT} sheets "
          f"({len(book.sheetnames)}: {book.sheetnames})")
    check(XLSX_METHODS_SHEET in book.sheetnames,
          f"Compare: the workbook carries a {XLSX_METHODS_SHEET!r} sheet ({book.sheetnames})")

    # --- remove one shown institution -> the basket (now == cap) refills ---
    before_ids = _compare_deeplink_ids(page)
    page.locator('[class*="st-key-cmp_rm_"] button').first.click(timeout=ACTION_TIMEOUT_MS)
    _settle_figures(page, COMPARE_MIN_FIGURES)
    after_ids = _compare_deeplink_ids(page)
    check(len(after_ids) == COMPARE_CAP,
          f"Compare: after removing one shown institution, the remaining basket (3) still fills the "
          f"comparison to {COMPARE_CAP} (got {len(after_ids)})")
    check(set(after_ids) != set(before_ids),
          f"Compare: removing a shown institution changes the compared set ({before_ids} -> {after_ids})")
    check(CAP_TRUNCATED_SUBSTR not in _full_page_text(page),
          "Compare: the truncation notice is gone once the basket no longer exceeds the cap")
    _no_exception(page, "Compare (after metric/slider/remove interactions)")
    return {"remaining_ids": after_ids}


def _pair_deeplink_ids(page) -> list[str]:
    loc = page.locator('[data-testid="stCode"]').filter(has_text="?pair=").first
    if loc.count() == 0:
        return []
    text = loc.text_content() or ""
    if "?pair=" not in text:
        return []
    return text.split("?pair=", 1)[1].strip().split(",")


def check_handoff(page) -> None:
    """The in-session `st.switch_page` hand-off (Fix X-2B): force the B
    selectbox to the LAST option, read the pair off Compare's own printed
    `?pair=` deep link, click the hand-off button and confirm Collaborate
    opens on that SAME pair in the SAME tab -- basket and tree/basis survive
    the hop."""
    _open_select(page, "cmp_pair_b")
    page.locator('[role="option"]').last.click(timeout=ACTION_TIMEOUT_MS)
    _settle(page, 2000)

    picked_ids = _pair_deeplink_ids(page)
    check(len(picked_ids) == 2 and picked_ids[0] != picked_ids[1],
          f"Compare: the hand-off's own deep link names two distinct ids ({picked_ids})")

    n_before = _sidebar_basket_n(page)
    tree_before = _selectbox_value(page, "tree")
    check(n_before is not None,
          f"Compare: sidebar basket count is readable before the hop (caption gave {n_before!r})")

    page.locator(".st-key-cmp_handoff_open button").first.click(timeout=ACTION_TIMEOUT_MS)
    # 2B-R2-11: Collaborate's tables are hand-built HTML now, not a canvas
    # `st.dataframe` -- the untapped table's own `data-table` hook is what
    # actually signals the page has finished landing.
    page.wait_for_selector('[data-table="collab_untapped"]', timeout=ACTION_TIMEOUT_MS)
    _settle(page, 2500)

    landed_ids = _pair_deeplink_ids(page)
    check(landed_ids == picked_ids,
          f"Collaborate: opened on the SAME pair the hand-off named ({picked_ids} -> {landed_ids})")
    _no_exception(page, "Collaborate (opened from the Compare hand-off)")

    n_after = _sidebar_basket_n(page)
    tree_after = _selectbox_value(page, "tree")
    check(n_after is not None and n_after == n_before,
          f"Basket: the sidebar count survives the hand-off hop ({n_before} -> {n_after})")
    check(tree_after == tree_before,
          f"Scenario: the tree selection survives the hand-off hop ({tree_before!r} -> {tree_after!r})")

    # --- 2B-R-10: the four sections all render ------------------------------
    body_text = _full_page_text(page)
    for header in COLLAB_SECTION_HEADERS:
        check(header in body_text, f"Collaborate (2B-R-10): section header {header!r} renders")

    # --- 2025* on the pulse chart --------------------------------------------
    pulse_points = page.evaluate(
        "(() => { const el = document.querySelector('.st-key-fig_pulse .js-plotly-plot');"
        " if (!el || !el.data) return [];"
        " return el.data.flatMap(t => t.x || []); })()")
    check(any(str(v) == BONUS_YEAR_AXIS_LABEL for v in pulse_points) or BONUS_YEAR_AXIS_LABEL in body_text,
          f"Collaborate pulse: the partial bonus year is starred ({BONUS_YEAR_AXIS_LABEL!r})")

    # --- rank direction: the two ranks read in OPPOSITE directions ----------
    rank_matches = re.findall(r"ranks number\s*\*?\*?(\d+)\*?\*?", body_text)
    check(len(rank_matches) >= 2,
          f"Collaborate pulse: both rank lines render (found {len(rank_matches)} 'ranks number' phrases)")
    if len(rank_matches) >= 2:
        check(rank_matches[0] != rank_matches[1],
              f"Collaborate pulse: the two ranks read asymmetrically, proving no swap "
              f"(got {rank_matches[0]!r} and {rank_matches[1]!r})")

    # --- swap flips the order --------------------------------------------------
    page.locator(".st-key-pair_swap button").first.click(timeout=ACTION_TIMEOUT_MS)
    _settle(page, 2000)
    swapped_ids = _pair_deeplink_ids(page)
    check(swapped_ids == list(reversed(landed_ids)),
          f"Collaborate: swap flips A and B ({landed_ids} -> {swapped_ids})")
    _no_exception(page, "Collaborate (hand-off + swap)")


def check_collab_anchor_pair(app_dir: Path, port: int) -> None:
    """2B-R2-11: the field-breakdown chart, its table's chips/arrows/links,
    the shared-topics table and its slider, the untapped tables, and the two
    deleted gap tables -- all driven on the manager-verified anchor pair
    (Universite de Strasbourg x CNRS, CNRS's own first partner), a fresh
    standalone session reached via `?pair=` exactly like the below-floor
    check. Deliberately NOT run on the journey's hand-off pair (built from an
    arbitrary L1 ranking): a small basket candidate could legitimately land
    below the floor-5 pair-table cutoff, which would fail every one of these
    checks for a reason that is not an app defect."""
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page(viewport={"width": 1280, "height": 1000})
            page.set_default_timeout(ACTION_TIMEOUT_MS)
            base = f"http://127.0.0.1:{port}"
            page.goto(f"{base}/Collaborate?pair={COLLAB_PAIR_A},{COLLAB_PAIR_B}",
                      wait_until="domcontentloaded")
            page.wait_for_selector('[data-table="collab_untapped"]', timeout=ACTION_TIMEOUT_MS)
            _settle(page, 3000)

            found_tables = [t for t in COLLAB_TABLES if page.locator(f'[data-table="{t}"]').count() > 0]
            check(found_tables == list(COLLAB_TABLES),
                  f"Collaborate anchor pair (2B-R2-11): all four tables render ({found_tables})")
            check(page.locator('[data-testid="stDataFrame"]').count() == 0,
                  "Collaborate anchor pair: no canvas grid -- tables are hand-built HTML")

            check(page.locator(".st-key-fig_fields").count() >= 1,
                  "Collaborate anchor pair (2B-R2-11a): the field-breakdown chart renders")
            fields_data = _fig_xy_text(page, ".st-key-fig_fields .js-plotly-plot")
            n_field_vals = sum(len([v for v in tr["x"] if v not in (None, "")])
                               for tr in (fields_data or []))
            check(bool(fields_data) and n_field_vals > 0,
                  f"Collaborate anchor pair: the field chart carries real values ({n_field_vals})")

            field_domains = _table_cells(page, "collab_fields", ".bu-chip", "data-domain")
            n_field_rows = _table_rows(page, "collab_fields")
            check(len(field_domains) == n_field_rows and n_field_rows > 0,
                  f"Collaborate anchor pair: every field row carries one domain chip "
                  f"({len(field_domains)} chips, {n_field_rows} rows)")
            field_arrows = set(_table_cells(page, "collab_fields", ".bu-arrow", "data-arrow"))
            check(bool(field_arrows),
                  f"Collaborate anchor pair: field rows carry a dynamics arrow ({field_arrows})")

            hrefs = _hrefs(page)
            field_hrefs = _table_cells(page, "collab_fields", ".bu-link", "href")
            check(bool(field_hrefs) and all(h in hrefs for h in field_hrefs if h),
                  "Collaborate anchor pair: every field row links its own pair+field co-publications")
            if field_hrefs:
                u = field_hrefs[0]
                check(u.count("authorships.institutions.id:") == 2,
                      f"Collaborate anchor pair: a field link ANDs both institutions ({u!r})")
                check(any(k in u for k in TAXON_FILTER_KEYS),
                      f"Collaborate anchor pair: a field link carries a taxon filter ({u!r})")

            n_topics_default = _table_rows(page, "collab_topics")
            check(n_topics_default == TOPICS_ROWS_DEFAULT,
                  f"Collaborate anchor pair: the topic table opens at its default depth "
                  f"({n_topics_default} rows)")
            topic_domains = _table_cells(page, "collab_topics", ".bu-chip", "data-domain")
            check(len(topic_domains) == 2 * n_topics_default,
                  f"Collaborate anchor pair: each topic row carries a topic AND a subfield chip "
                  f"({len(topic_domains)} of {2 * n_topics_default})")
            topic_arrows = _table_cells(page, "collab_topics", ".bu-arrow", "data-arrow")
            check(len(set(topic_arrows)) > 1,
                  f"Collaborate anchor pair: topic arrows vary row to row ({set(topic_arrows)})")
            topic_hrefs = [h for h in _table_cells(page, "collab_topics", ".bu-link", "href") if h]
            check(bool(topic_hrefs) and all(h in hrefs for h in topic_hrefs),
                  "Collaborate anchor pair: every shown topic row links its own pair+topic co-pubs")
            if topic_hrefs:
                u = topic_hrefs[0]
                check(u.count("authorships.institutions.id:") == 2,
                      f"Collaborate anchor pair: a topic link ANDs both institutions ({u!r})")
                check("primary_topic.id:" in u,
                      f"Collaborate anchor pair: a topic link carries the topic filter ({u!r})")

            topics_slider = page.locator('.st-key-topics_n input[type="range"]').first
            check(topics_slider.count() >= 1, "Collaborate anchor pair: the topic-depth slider renders")
            if topics_slider.count() >= 1:
                check(topics_slider.get_attribute("max") == str(TOPICS_TOP_N_CAP),
                      f"Collaborate anchor pair: the slider's top stop is the shipped cap "
                      f"({TOPICS_TOP_N_CAP})")
                topics_slider.focus()
                topics_slider.press("ArrowRight")
                _settle(page, 3500)
                n_after_step = _table_rows(page, "collab_topics")
                check(n_after_step == n_topics_default + TOPICS_ROWS_STEP,
                      f"Collaborate anchor pair (slider): one step right adds one step of rows "
                      f"({n_topics_default} -> {n_after_step})")
                topics_slider.press("Home")
                _settle(page, 3500)
                n_min = _table_rows(page, "collab_topics")
                check(n_min < n_after_step,
                      f"Collaborate anchor pair (slider): the minimum shrinks the table "
                      f"({n_after_step} -> {n_min})")
                topics_slider.press("End")
                _settle(page, 3500)
                n_max = _table_rows(page, "collab_topics")
                check(n_max > n_min,
                      f"Collaborate anchor pair (slider): the maximum grows it back ({n_max})")

            n_untapped_default = _table_rows(page, "collab_untapped")
            untapped_hrefs = [h for h in _table_cells(page, "collab_untapped", ".bu-link", "href") if h]
            check(bool(untapped_hrefs) and all(h in hrefs for h in untapped_hrefs),
                  "Collaborate anchor pair: every untapped row links its own topic")
            untapped_slider = page.locator('.st-key-untapped_n input[type="range"]').first
            if untapped_slider.count() >= 1 and n_untapped_default > 0:
                untapped_slider.focus()
                untapped_slider.press("ArrowRight")
                _settle(page, 3500)
                check(_table_rows(page, "collab_untapped") != n_untapped_default,
                      "Collaborate anchor pair: the untapped-depth slider changes its own row count")

            body_text = _full_page_text(page)
            check(GAP_TABLE_HEADER_SUBSTR not in body_text.lower(),
                  "Collaborate anchor pair (2B-R2-11f): no 'X does not publish in' gap table renders")
            check(DOWNLOAD_GAPS_TEXT not in body_text,
                  "Collaborate anchor pair (2B-R2-11f): no gap-list CSV download renders")
            check(NOT_OFFERED_HEADER in body_text,
                  f"Collaborate anchor pair (2B-R2-13): the shared {NOT_OFFERED_HEADER!r} header "
                  f"discloses the deletion")
            _no_forbidden_vocab(page, "Collaborate anchor pair page")

            for key, header in (("dl_fields", FIELD_BREAKDOWN_CSV_HEADER),
                                ("dl_topics", JOINT_TOPICS_CSV_HEADER),
                                ("dl_untapped", UNTAPPED_CSV_HEADER)):
                btn = page.locator(f".st-key-{key} button")
                if btn.count() >= 1:
                    # Re-anchored fresh on each pass, and settled first: the
                    # slider interactions above triggered several reruns, and a
                    # download button click that lands mid-rerun can silently
                    # miss the event (measured flake, not an app defect).
                    btn.scroll_into_view_if_needed(timeout=ACTION_TIMEOUT_MS)
                    _settle(page, 1000)
                    with page.expect_download(timeout=60_000) as dl_info:
                        btn.first.click(timeout=ACTION_TIMEOUT_MS)
                    with open(dl_info.value.path(), "r", encoding="utf-8") as fh:
                        got = fh.readline()
                    check(got.strip() == header,
                          f"Collaborate anchor pair: {key} CSV header matches the contract ({got!r})")
            _no_exception(page, "Collaborate (anchor pair, standalone session)")
            browser.close()
    except Exception as exc:
        fail_section("Collaborate anchor pair", exc)


def check_ifremer_crash_seed(app_dir: Path, port: int) -> None:
    """2B-R2-1a: the profile that took the app down at gate 2B-R (Ifremer is
    both an umbrella AND type-corrected). A dedicated, standalone check (fresh
    session, `?seed=`) at all three widths -- the crash class this exists to
    catch is "renders for the common seed, still crashes for the rare one",
    so this must never share a page/session with the Gdansk walk above.

    Also carries the 2B-R2-7 SI-padding bounding-box proof: Ifremer's top
    subfields panel is FA3's own measured worst case (SI 0.17-21.35 at 1280
    px), so this is where a clipped outer-end value label would actually
    show up."""
    base = f"http://127.0.0.1:{port}"
    for width in WIDTHS:
        try:
            with sync_playwright() as p:
                browser = p.chromium.launch()
                page = browser.new_page(viewport={"width": width, "height": 900})
                page.set_default_timeout(ACTION_TIMEOUT_MS)
                page.goto(f"{base}/Find?seed={CRASH_SEED}", wait_until="domcontentloaded")
                page.wait_for_selector('[role="tab"]', timeout=ACTION_TIMEOUT_MS)
                _settle(page, 3000)
                _no_exception(page, f"Ifremer crash seed {width}px")
                check(page.locator(".st-key-profile .benchup-kpi").count() == N_CARDS,
                      f"Ifremer {width}px: {N_CARDS} cards render")
                body = _full_page_text(page)
                m = IDENTITY_TYPE_CORRECTED_RE.search(body)
                check(m is not None,
                      f"Ifremer {width}px: the inline type correction "
                      f"'<type>* (was: <type>)' renders (body has: {m.group(0) if m else None})")
                star_color = page.evaluate(
                    """() => { const spans = Array.from(document.querySelectorAll('.st-key-profile span'));
                        for (const s of spans) { if ((s.textContent || '').trim() === '*') {
                            return getComputedStyle(s).color; } } return null; }""")
                check(star_color is not None, f"Ifremer {width}px: the '*' renders as its own span")
                if star_color:
                    rgbv = [int(x) for x in re.findall(r"\d+", star_color)[:3]]
                    reddish = len(rgbv) == 3 and rgbv[0] > 150 and rgbv[0] > rgbv[1] + 40 and rgbv[0] > rgbv[2] + 40
                    check(reddish, f"Ifremer {width}px: the '*' is coloured red ({star_color})")
                name_href = page.locator(".st-key-profile h3 a").first.get_attribute("href") or ""
                check("openalex.org/works" in name_href,
                      f"Ifremer {width}px: the institution name links to OpenAlex works")
                check(NO_LONGER_ON_FIND not in body,
                      f"Ifremer {width}px: no separate {NO_LONGER_ON_FIND!r} link")
                _ensure_expander_open(page, "panel_subfields", ".st-key-fig_subfields")
                _settle(page, 1500)
                if width >= 1280:
                    # 2B-R2-7's padding fix targets the SI panel's own gutter,
                    # measured by FA3 at 1280px (the worst case: SI 0.17-21.35).
                    # At 390px the whole SI column is ~30px wide and labels
                    # crowd regardless -- FA3's OWN documented pre-existing
                    # issue (Find never passes `stacked=True`), unrelated to
                    # this stream's fix, so this bbox proof is scoped to the
                    # widths the fix actually targets.
                    _outer_label_bbox_check(page, ".st-key-fig_subfields .js-plotly-plot",
                                            f"Ifremer {width}px SI padding (worst-case panel)")
                scroll = page.evaluate("document.documentElement.scrollWidth")
                inner = page.evaluate("window.innerWidth")
                check(scroll <= inner + 2,
                      f"Ifremer {width}px: scrollWidth {scroll} <= innerWidth+2 {inner + 2}")
                browser.close()
        except Exception as exc:  # noqa: BLE001 -- one width's crash must not skip the rest
            fail_section(f"Ifremer crash seed {width}px", exc)


def check_below_floor_pair(app_dir: Path, port: int) -> None:
    """2B-R2-11(g)/2B-R2-12: a REAL below-floor pair (Strasbourg x Bavarian
    Academy of Sciences and Humanities, 2 joint works < the floor of 5)
    renders the honest notice -- pulse, total and links stay, the field AND
    topic breakdowns do not (untapped/adjacent are NOT floor-gated, so they
    still render). Deliberately a FRESH, standalone page/session (this
    asserts no persistence claim), reached via `?pair=` -- the one case in
    this file where `page.goto()` for something other than the very first
    load is correct, per the module docstring."""
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page(viewport={"width": 1280, "height": 900})
            page.set_default_timeout(ACTION_TIMEOUT_MS)
            base = f"http://127.0.0.1:{port}"
            page.goto(f"{base}/Collaborate?pair={BELOW_FLOOR_A_ID},{BELOW_FLOOR_B_ID}",
                      wait_until="domcontentloaded")
            page.wait_for_selector('[data-testid="stSidebarNav"]', state="attached",
                                    timeout=ACTION_TIMEOUT_MS)
            _settle(page, 4000)
            body_text = _full_page_text(page)
            check(BELOW_FLOOR_NOTICE_RE.search(body_text) is not None,
                  f"Collaborate below-floor pair (2B-R2-12, floor {PAIR_FLOOR}): the honest "
                  f"notice renders (looking for pattern {BELOW_FLOOR_NOTICE_RE.pattern!r})")
            check("The relationship, year by year" in body_text,
                  "Collaborate below-floor pair: the pulse section still renders")
            check("Read the publications on OpenAlex" in body_text,
                  "Collaborate below-floor pair: the link-outs section still renders")
            check(page.locator(".st-key-collab_links a").count() >= 3,
                  "Collaborate below-floor pair: all 3 OpenAlex link buttons still render")
            found = [t for t in COLLAB_TABLES if page.locator(f'[data-table="{t}"]').count() > 0]
            check(found == list(COLLAB_TABLES_BELOW_FLOOR),
                  f"Collaborate below-floor pair (2B-R2-12): only the untapped/adjacent tables "
                  f"render, field+topic breakdowns are absent (found {found})")
            _no_exception(page, "Collaborate (below-floor pair, standalone session)")
            browser.close()
    except Exception as exc:
        fail_section("Collaborate below-floor pair", exc)


def check_methods_journey(page) -> None:
    _click_nav(page, NAV_METHODS)
    page.wait_for_selector('[data-testid="stExpander"]', timeout=ACTION_TIMEOUT_MS)
    _settle(page, 1500)

    n_sections = page.locator('[data-testid="stExpander"]').count()
    check(n_sections >= METHODS_MIN_SECTIONS,
          f"Methods: >= {METHODS_MIN_SECTIONS} section expanders render ({n_sections})")

    body = _full_page_text(page)
    leftover = PLACEHOLDER_RE.findall(body)
    check(not leftover, f"Methods: no unresolved {{placeholder}} text on the page (found {leftover[:5]})")

    # 2B-R-11: the lens concordance table ("Reading the lens codes").
    _ensure_expander_open_by_text(page, LENS_CODES_TITLE, "text=(C1)")
    _settle(page, 1000)
    concordance_text = _full_page_text(page)
    check(LENS_CODES_TITLE in concordance_text,
          f"Methods (2B-R-11): the {LENS_CODES_TITLE!r} section renders")
    check("(C1)" in concordance_text and "(L7)" in concordance_text,
          "Methods: the concordance table names both optional lenses' internal ids")
    check(re.search(r"L[0-9]\**\s*\(L0\)", concordance_text) is not None
          or "(L0)" in concordance_text,
          "Methods: the concordance table names the L0 internal id")

    with page.expect_download(timeout=ACTION_TIMEOUT_MS) as dl_info:
        page.locator(".st-key-dl_methods_note button").first.click(timeout=ACTION_TIMEOUT_MS)
    raw = Path(dl_info.value.path()).read_bytes()
    check(len(raw) > 500, f"Methods: the source-note Markdown download is a real document ({len(raw)} bytes)")
    _no_exception(page, "Methods")


def check_narrative_persistence(page) -> dict:
    """The tree/basis scenario reads the same on Compare, Collaborate AND
    Methods's own sidebar selects; the basket's `{n} of {cap} added` sidebar
    count agrees across all three; returning to Find shows the Gdansk seed
    still loaded and the SAME count on Find's own editable list."""
    _click_nav(page, NAV_COMPARE)
    _settle(page, 1500)
    tree_c, basis_c = _selectbox_value(page, "tree"), _selectbox_value(page, "basis")
    check(tree_c == TREE_LABEL_ORIGINAL,
          f"Compare: sidebar taxonomy still {TREE_LABEL_ORIGINAL!r} (got {tree_c!r})")
    check(basis_c == BASIS_LABEL_FRAC,
          f"Compare: sidebar counting basis still {BASIS_LABEL_FRAC!r} (got {basis_c!r})")
    n_compare = _sidebar_basket_n(page)
    check(n_compare is not None, f"Compare: sidebar basket count is readable (caption gave {n_compare!r})")
    scroll_check_widths(page, "Compare", extra_widths=(1920, 390))

    _click_nav(page, NAV_COLLAB)
    _settle(page, 1500)
    tree_l, basis_l = _selectbox_value(page, "tree"), _selectbox_value(page, "basis")
    check(tree_l == TREE_LABEL_ORIGINAL,
          f"Collaborate: sidebar taxonomy still {TREE_LABEL_ORIGINAL!r} (got {tree_l!r})")
    check(basis_l == BASIS_LABEL_FRAC,
          f"Collaborate: sidebar counting basis still {BASIS_LABEL_FRAC!r} (got {basis_l!r})")
    n_collab = _sidebar_basket_n(page)
    check(n_compare is not None and n_compare == n_collab,
          f"Basket: the sidebar count agrees on Compare and Collaborate ({n_compare} vs {n_collab})")
    scroll_check_widths(page, "Collaborate", extra_widths=(1920, 390))

    _click_nav(page, NAV_METHODS)
    _settle(page, 1000)
    tree_m, basis_m = _selectbox_value(page, "tree"), _selectbox_value(page, "basis")
    check(tree_m == TREE_LABEL_ORIGINAL,
          f"Methods: sidebar taxonomy still {TREE_LABEL_ORIGINAL!r} (got {tree_m!r})")
    check(basis_m == BASIS_LABEL_FRAC,
          f"Methods: sidebar counting basis still {BASIS_LABEL_FRAC!r} (got {basis_m!r})")
    n_methods = _sidebar_basket_n(page)
    check(n_compare is not None and n_compare == n_methods,
          f"Basket: the sidebar count agrees on Compare and Methods ({n_compare} vs {n_methods})")
    _no_exception(page, "Methods (persistence hop)")
    scroll_check_widths(page, "Methods", extra_widths=(1920, 390))
    page.set_viewport_size({"width": 1280, "height": 900})
    _settle(page, 800)

    _click_nav(page, "Find")
    _settle(page, 1500)
    heading = _seed_heading(page)
    check("Gda" in heading, f"Find: returning from Methods still shows the Gdansk seed (got {heading!r})")
    n_find = _basket_count(page)
    check(n_compare is not None and n_find == n_compare,
          f"Basket: {n_find} items on Find matches the {n_compare} the sidebar reported on "
          f"Compare/Collaborate")
    return {"n_basket": n_find}


def scroll_check_widths(page, label: str, extra_widths=(1920, 390)) -> None:
    """No horizontal body scroll at every width, on whatever page is
    currently loaded (cross-page requirement: 'no horizontal body scroll at
    1920/1280/390 on every page'). 1280 is asserted by the caller's own
    render check already; this adds the other two."""
    try:
        current = page.viewport_size or {"width": 1280, "height": 900}
        for width in extra_widths:
            page.set_viewport_size({"width": width, "height": 900})
            _settle(page, 900)
            if width == 390:
                _ensure_sidebar_open(page)
                _settle(page, 300)
            scroll = page.evaluate("document.documentElement.scrollWidth")
            inner = page.evaluate("window.innerWidth")
            check(scroll <= inner + 2, f"{label} {width}px: scrollWidth {scroll} <= innerWidth+2 {inner + 2}")
        page.set_viewport_size(current)
        _settle(page, 500)
    except Exception as exc:
        fail_section(f"{label}: extra-width scroll check", exc)


def check_journey_widths(page, shot_dir: Path) -> None:
    """Compare at three widths; one screenshot each for Collaborate and
    Methods at 1280 px (the other two widths are already covered by
    `scroll_check_widths` inside `check_narrative_persistence`)."""
    shot_dir.mkdir(parents=True, exist_ok=True)
    _click_nav(page, NAV_COMPARE)
    _settle_figures(page, COMPARE_MIN_FIGURES)
    for width in WIDTHS:
        page.set_viewport_size({"width": width, "height": 900})
        _settle(page, 1000)
        if width == 390:
            _ensure_sidebar_open(page)
        scroll = page.evaluate("document.documentElement.scrollWidth")
        inner = page.evaluate("window.innerWidth")
        check(scroll <= inner + 2, f"Compare {width}px: scrollWidth {scroll} <= innerWidth+2 {inner + 2}")
        p = shot_dir / f"smoke_compare_{width}.png"
        page.screenshot(path=str(p), full_page=True)
        check(p.is_file(), f"Compare {width}px: screenshot written ({p.name})")
    page.set_viewport_size({"width": 1280, "height": 900})
    _settle(page, 800)

    _click_nav(page, NAV_COLLAB)
    _settle(page, 2000)
    scroll = page.evaluate("document.documentElement.scrollWidth")
    inner = page.evaluate("window.innerWidth")
    check(scroll <= inner + 2, f"Collaborate 1280px: scrollWidth {scroll} <= innerWidth+2 {inner + 2}")
    p = shot_dir / "smoke_collab_1280.png"
    page.screenshot(path=str(p), full_page=True)
    check(p.is_file(), f"Collaborate 1280px: screenshot written ({p.name})")

    _click_nav(page, NAV_METHODS)
    _settle(page, 1500)
    scroll = page.evaluate("document.documentElement.scrollWidth")
    inner = page.evaluate("window.innerWidth")
    check(scroll <= inner + 2, f"Methods 1280px: scrollWidth {scroll} <= innerWidth+2 {inner + 2}")
    p = shot_dir / "smoke_methods_1280.png"
    page.screenshot(path=str(p), full_page=True)
    check(p.is_file(), f"Methods 1280px: screenshot written ({p.name})")


# ------------------------------------------------------------------ main ----

def main() -> int:
    global PORT, BASE_URL
    parser = argparse.ArgumentParser()
    parser.add_argument("--port", type=int, default=8611)
    parser.add_argument("--app-dir", type=str, default=None,
                         help="app/ root to target (default: this file's own app/); "
                              "pass a throwaway copy for the non-vacuity proofs.")
    args = parser.parse_args()
    PORT = args.port
    BASE_URL = f"http://127.0.0.1:{PORT}"
    app_dir = Path(args.app_dir).resolve() if args.app_dir else DEFAULT_APP_DIR
    shot_dir = app_dir / "tests" / "ui" / "screenshots"

    server = _start_server(app_dir, PORT)
    profile_expect: dict = {}
    journey: dict = {}
    try:
        if not _wait_for_port(PORT, timeout=90.0):
            check(False, f"server did not open port {PORT} within timeout")
            return 1

        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page(viewport={"width": 1280, "height": 900})
            page.set_default_timeout(ACTION_TIMEOUT_MS)

            def _run_profile_panels() -> None:
                profile_expect.update(check_profile_and_panels(page) or {})

            sections = [
                ("Menu", lambda: check_menu(page)),
                ("Find search", lambda: check_find_search(page)),
                ("Basket", lambda: check_basket(page)),
                ("Controls placement", lambda: check_controls_placement(page)),
                ("Profile / panels", _run_profile_panels),
                ("Bonus year axis", lambda: check_bonus_year_axis(page)),
                ("SI value labels", lambda: check_si_value_labels(page)),
                ("Frontier slider (both modes)",
                 lambda: check_frontier_slider_modes(page, profile_expect)),
                ("A11 tab overflow (both optional lenses)", lambda: check_tab_overflow_a11(page)),
                ("Benchmark lens guide", lambda: check_benchmark_lens_guide(page)),
                ("Tables / export", lambda: check_tables_and_export(page)),
                ("Institution link popup (A10)", lambda: check_institution_link_popup(page)),
                ("Settings", lambda: check_settings(page)),
                ("Persistence", lambda: check_persistence(page, profile_expect)),
                ("Type filter clear", lambda: check_type_filter_clear(page)),
            ]
            for name, fn in sections:
                try:
                    fn()
                except Exception as exc:  # noqa: BLE001 -- one section's crash must not hang the run
                    fail_section(name, exc)

            try:
                undefined = _find_undefined_l2f_seed(app_dir)
                if undefined is None:
                    check(False, "Undefined lens: no institution with an undefined L2f was found")
                else:
                    check_undefined_lens(page, *undefined)
            except Exception as exc:  # noqa: BLE001
                fail_section("Undefined lens", exc)

            # The full four-page journey, Menu -> Find -> Compare ->
            # Collaborate -> Methods, on the SAME page/session the checks
            # above already built up.
            def _run_compare_journey() -> None:
                journey.update(check_compare_journey(page, journey.get("candidates", [])) or {})

            journey_sections = [
                ("Journey: basket (L1 candidates + seed)",
                 lambda: journey.__setitem__("candidates", check_journey_basket(page))),
                ("Journey: Compare page", _run_compare_journey),
                ("Journey: hand-off to Collaborate", lambda: check_handoff(page)),
                ("Journey: Methods page", lambda: check_methods_journey(page)),
                ("Journey: narrative persistence", lambda: check_narrative_persistence(page)),
                ("Journey: widths + screenshots", lambda: check_journey_widths(page, shot_dir)),
            ]
            for name, fn in journey_sections:
                try:
                    fn()
                except Exception as exc:  # noqa: BLE001 -- one section's crash must not hang the run
                    fail_section(name, exc)

            page.close()
            check_screenshots(browser, shot_dir)
            browser.close()

        # Isolated, standalone sessions -- deliberately outside the shared
        # browser/page above (these checks assert no persistence claim).
        check_ifremer_crash_seed(app_dir, PORT)
        check_collab_anchor_pair(app_dir, PORT)
        check_below_floor_pair(app_dir, PORT)
    finally:
        _stop_server(server)

    failed = [m for ok, m in RESULTS if not ok]
    print(f"\n{len(RESULTS) - len(failed)} of {len(RESULTS)} checks passed")
    if failed:
        for m in failed:
            print("FAILED:", m)
        return 1
    print("ALL CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
