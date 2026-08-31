"""
Acceptance probe for the Compare page as rebuilt in Phase 2B-R2 (stream CP3,
decisions 2B-R2-3/4/5/8/9/10; 2B-R's own 4/5/6/7/9/12 still stand). Same shape
as ops/_probe_collab.py: start `streamlit run pages/2_<scales>_Compare.py` as a
subprocess, drive it headless with Playwright, ALWAYS terminate the server.

Selectors are locale-independent -- the `st-key-<key>` classes the page's own
keyed widgets and containers emit, plus `[data-testid=...]`. Nothing is asserted
against a canvas: the workbook is checked by DOWNLOADING it and opening it with
openpyxl, the figures are counted as plotly roots, the row order is read off the
rendered SVG tick labels, and every VALUE the page renders is checked by
recomputing it here from `lib/compare_data.py` and looking for the page's own
formatting of it in the DOM. A probe that only counted elements would pass a
page whose cards had drifted from the frame.

WHAT THIS ROUND ADDS, and why each one is here rather than in the AppTest file:
  * EVERY SELECTOR OPTION IS DRIVEN TO A RENDER, on every level, in the browser
    (2B-R's lesson: the crash class that survived was "option visible, render
    path unreached"). AppTest proves the option list; only a real rerun proves
    the figure comes back.
  * ROW ORDER IS COMPARED BETWEEN TWO METRIC TABS by reading the y-axis tick
    labels out of the rendered plot (2B-R2-5). The frame-level property is
    tested in pytest; this is the drawn one.
  * THE GUTTERS CARRY NUMBERS and the dynamics gutter carries its raw-delta
    string; the reference dashes are counted as SVG shapes (2B-R2-3/4).
  * THE CARDS' DOT, LINKS AND DELETIONS (2B-R2-9) are checked in the DOM: the
    dot exists, each institution NAME is an OpenAlex link, and the separate
    publications button is gone.
  * THE POOL AND COLOUR TOGGLES (2B-R2-10) are clicked and their effect read
    back -- the pool rule in the note, the domain names in the legend.

Usage:  python ops/_probe_compare.py [--port 8605]
Exit 0 when every check passes; 1 otherwise. Stdout is ASCII-only (cp1252
console).
"""
from __future__ import annotations

import argparse
import io
import re
import socket
import subprocess
import sys
import time
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright

APP_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(APP_DIR))

from lib import charts_compare, compare_data, copy, state, tiles, views_compare  # noqa: E402
from lib.ranked import works_link_named  # noqa: E402

PAGE = "pages/2_⚖️_Compare.py"
DEFAULT_PORT = 8605

STRASBOURG = "I68947357"
SORBONNE = "I39804081"
FREIBURG = "I161046081"
GDANSK = "I40413290"          # the fourth id: over the cap ON PURPOSE
LINK_IDS = [STRASBOURG, SORBONNE, FREIBURG, GDANSK]
SHOWN_IDS = LINK_IDS[:state.COMPARE_CAP]

TREE, BASIS = "bestfit", "frac"

SHOT_DIR = APP_DIR / "tests" / "ui" / "screenshots"
WIDTHS = [1920, 1280, 390]
SHOT_HEIGHT_PX = 2400   # full_page=True is a no-op on Streamlit's scroll container, so the
TALL_SHOT_PX = 5600     # VIEWPORT is the screenshot; the wide widths get a tall one.
HEAD_SHOT_PX = 900      # the true above-the-fold view at 1280 px
MIN_FIGURES = 8         # subject, ERC, SDG, frontier map, shared frontier, impact,
                        # impact by subfield, trends, coverage -- the shared-frontier
                        # chart is the one that can legitimately be absent (an empty
                        # intersection), so the floor is one below the nine drawn.

PORT = DEFAULT_PORT
RESULTS: list[tuple[bool, str]] = []


def check(ok: bool, message: str) -> bool:
    RESULTS.append((bool(ok), message))
    print(("PASS: " if ok else "FAIL: ") + message)
    return bool(ok)


def _wait_for_port(port: int, timeout: float = 60.0) -> bool:
    deadline = time.time() + timeout
    while time.time() < deadline:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            if sock.connect_ex(("127.0.0.1", port)) == 0:
                return True
        time.sleep(0.5)
    return False


def _load(page) -> None:
    page.goto(f"http://127.0.0.1:{PORT}/?compare=" + ",".join(LINK_IDS),
              wait_until="domcontentloaded")
    page.wait_for_selector(".js-plotly-plot", timeout=240_000)
    _settle(page)


def _settle(page, target: int = MIN_FIGURES, timeout_ms: int = 180_000) -> int:
    """Streamlit streams its elements, so the figure count climbs for a while
    after the first plot appears. Wait for it to reach the acceptance floor and
    then stop changing, rather than sleeping a guessed number of seconds."""
    deadline = time.time() + timeout_ms / 1000.0
    last, stable = -1, 0
    while time.time() < deadline:
        now = _n_figures(page)
        stable = stable + 1 if now == last and now >= target else 0
        last = now
        if stable >= 3:
            break
        page.wait_for_timeout(1000)
    page.wait_for_timeout(1500)
    return last


def _n_figures(page) -> int:
    return page.locator(".js-plotly-plot").count()


def _text_of(page, selector: str) -> str:
    return page.evaluate(
        "(sel) => { const e = document.querySelector(sel); return e ? e.textContent : ''; }",
        selector)


def _html_of(page, selector: str) -> str:
    return page.evaluate(
        "(sel) => { const e = document.querySelector(sel); return e ? e.innerHTML : ''; }",
        selector)


def _body_text(page) -> str:
    return page.evaluate("document.body.innerText")


def _captions(page) -> str:
    return page.evaluate(
        "Array.from(document.querySelectorAll('[data-testid=\"stCaptionContainer\"]'))"
        ".map(e => e.textContent).join('|')")


def _markdown(page) -> str:
    """Every markdown block's TEXT. 2B-R2-8 moved this page's prose out of
    captions and into `charts_compare.chart_note`, which renders as markdown."""
    return page.evaluate(
        "Array.from(document.querySelectorAll('[data-testid=\"stMarkdownContainer\"]'))"
        ".map(e => e.textContent).join('|')")


def _tooltips(page) -> str:
    """Every `title=` payload on the page: 2B-R2-8's methodology lives there,
    behind the `?` of a chart note, and a probe that read only visible text
    could not tell "moved into the tooltip" from "deleted"."""
    return page.evaluate(
        "Array.from(document.querySelectorAll('[title]')).map(e => e.getAttribute('title'))"
        ".join('|')")


def _warnings(page) -> str:
    return page.evaluate(
        "Array.from(document.querySelectorAll('[data-testid=\"stAlertContainer\"]'))"
        ".map(e => e.textContent).join('|')")


def _n_legends(page, names) -> int:
    """Legend strips: a markdown container whose text names every compared
    institution and which carries the chip spans. `charts.chip_legend_html`
    writes inline styles the browser normalises, so the count is taken on
    STRUCTURE and TEXT, never on an attribute substring."""
    return page.evaluate(
        "(names) => Array.from(document.querySelectorAll('[data-testid=\"stMarkdownContainer\"]'))"
        ".filter(e => names.every(n => e.textContent.includes(n))"
        " && e.querySelectorAll('span').length >= names.length * 2).length",
        list(names))


def _click_option(page, container_key: str, text: str) -> None:
    """Click one option of a keyed radio, then wait for the rerun to finish
    redrawing (a rerun rebuilds every figure on the page, so the count dips
    before it comes back)."""
    page.locator(f".st-key-{container_key}").get_by_text(text, exact=True).first.click()
    page.wait_for_timeout(2000)
    _settle(page)


def _radio_options(page, container_key: str) -> list:
    """The LABELS a keyed radio actually offers, read off the DOM -- so "every
    option is driven" is driven against what the page shows, never against what
    this file believes it shows."""
    return page.evaluate(
        "(key) => Array.from(document.querySelectorAll("
        "  '.st-key-' + key + ' [role=\"radiogroup\"] label'))"
        ".map(e => e.textContent.trim()).filter(t => t.length)",
        container_key)


def _pick_selectbox(page, container_key: str, text: str) -> None:
    """Open a keyed selectbox and choose an option by typing it: Streamlit's
    combobox filters as you type and Enter takes the first match, which needs
    no assumption about where the portal renders the option list."""
    box = page.locator(f".st-key-{container_key} input").first
    box.click()
    box.fill(text)
    page.wait_for_timeout(800)
    box.press("Enter")
    page.wait_for_timeout(2000)
    _settle(page)


def _first_literal(template: str) -> str:
    segments = [s for s in re.split(r"\{[^{}]*\}", template) if s.strip()]
    return segments[0].strip()


DASH_CHARS = "-\N{EN DASH}\N{EM DASH}\N{MINUS SIGN}"


def _dashless(text: str) -> str:
    """Streamlit renders every caption through markdown, which rewrites a
    double hyphen into a typographic dash. A probe that compared the source
    string byte-for-byte against the DOM would fail on the one note that carries
    a parenthetical -- `compare_data.DYNAMICS_DENOM_NOTE`, the dynamics note
    2B-R-6 requires verbatim -- for a typographic reason and nothing else.
    Dropping every dash from BOTH sides keeps the check about the words and the
    numbers, which is what the decision is about."""
    return "".join(c for c in text if c not in DASH_CHARS)


def _set_slider(page, container_key: str, target: int) -> int:
    """Streamlit's slider is an `input[type=range]` driven by react-aria: it
    takes keyboard steps, not a click at a coordinate (a click on the track is
    what a human does, but its landing value depends on pixel geometry, which
    is exactly the kind of thing a probe must not assert on). Arrow keys walk
    it by its own step until the value is the one asked for."""
    knob = page.locator(f".st-key-{container_key} input[type=range]").first
    knob.focus()
    for _ in range(20):
        now = int(knob.input_value())
        if now == target:
            break
        knob.press("ArrowLeft" if now > target else "ArrowRight")
        page.wait_for_timeout(300)
    page.wait_for_timeout(2000)
    _settle(page)
    return int(knob.input_value())


def _ctx():
    return views_compare._bundle()["ctx"]


def _names(ids) -> dict:
    idx = _ctx()["index_by_id"]
    return {i: str(idx.loc[i, "display_name"]) for i in ids}


# ------------------------------------------------------- the subject chart --

SUBJECT_FIG = "fig_cmp_subject"


def _fig_root(page, key: str) -> str:
    return f".st-key-{key} .js-plotly-plot"


def _tick_labels(page, key: str) -> list:
    """The y-axis tick labels of ONE figure, as drawn. Plotly writes them as
    SVG text under the y axis layer; the gutter numbers 2B-R2-3 appends live
    inside the SAME label (they are part of the tick string), which is why the
    caller strips them rather than this reading them out of a second place."""
    return page.evaluate(
        "(sel) => { const p = document.querySelector(sel); if (!p) return [];"
        " return Array.from(p.querySelectorAll('.yaxislayer-above .ytick text'))"
        ".map(t => t.textContent); }",
        _fig_root(page, key))


def _row_names(labels) -> list:
    """A tick label reduced to its taxon NAME: everything up to the first digit.
    The 2B-R2-3 gutter appends the volumes INTO the tick string, so a raw
    comparison of two tabs' labels would compare their numbers as well as their
    order -- which is not the property under test. Used on the SUBJECT chart
    only: field and subfield names carry no digit of their own (the SDG labels
    do, in front, which is why this is not pointed at that chart)."""
    return [re.split(r"\d", str(text))[0].strip() for text in labels]


def _n_shapes(page, key: str, kind: str) -> int:
    """How many elements matching `kind` one figure holds -- reference dashes
    and row rules are layout shapes, bubbles are scatter paths, so this counts
    the ink 2B-R2-4/10 ask for without asserting on a pixel."""
    return page.evaluate(
        "([sel, q]) => { const p = document.querySelector(sel); if (!p) return 0;"
        " return p.querySelectorAll(q).length; }",
        [_fig_root(page, key), kind])


def _strip_dot_colors(page) -> list:
    """The COMPUTED background colour of every span inside the overview strip
    that has one. The 2B-R2-9 dot is the only painted span there -- the swatch
    is a coloured GLYPH, i.e. `color`, not a background -- so this is the dot
    census, read as computed style rather than out of the markup because the
    browser re-serialises an inline `#RRGGBB` into `rgb(...)`."""
    return page.evaluate(
        "Array.from(document.querySelectorAll('.st-key-compare_strip span'))"
        ".map(e => getComputedStyle(e).backgroundColor)"
        ".filter(c => c && c !== 'rgba(0, 0, 0, 0)' && c !== 'transparent')")


def _rgb(hexcol: str) -> str:
    h = hexcol.lstrip("#")
    return "rgb({}, {}, {})".format(*(int(h[i:i + 2], 16) for i in (0, 2, 4)))


# --------------------------------------------------------------- the page --

def _probe_overview(page) -> None:
    """VIZ_SPEC 4.1 + 2B-R-7/2B-R2-9: the cards ARE `compare_data.overview`, so
    every rendered figure is recomputed here and looked for in the strip."""
    names = _names(SHOWN_IDS)
    strip = _text_of(page, ".st-key-compare_strip")
    for iid, name in names.items():
        check(name in strip, f"the overview names {iid}")
    check(_names([GDANSK])[GDANSK] not in strip,
          "the fourth institution of the link is NOT compared (cap of three)")

    frame = compare_data.overview(_ctx(), SHOWN_IDS).set_index("institution_id")
    idx = _ctx()["index_by_id"]
    for iid in SHOWN_IDS:
        facts = views_compare._card_facts(idx.loc[iid], frame.loc[iid])
        missing = [(label, value) for _c, label, value, _t in facts if value not in strip]
        check(not missing, f"every overview value of {iid} reads back from compare_data "
                           f"({len(facts)} values; missing {missing})")

    html = _html_of(page, ".st-key-compare_strip")
    n_cards = html.count(f'class="{tiles.TILE_CLASS}"')
    expected = len(SHOWN_IDS) * len(views_compare.CARD_COLUMNS)
    check(n_cards == expected, f"the strip draws one card per measure per institution "
                               f"({n_cards} of {expected})")

    # 2B-R2-9 the DOT: in the leading institution's own colour, on the measure
    # it leads, recomputed here from the frame and matched as MARKUP.
    leaders = views_compare._leaders(compare_data.overview(_ctx(), SHOWN_IDS))
    slots = views_compare._slots(_ctx(), SHOWN_IDS)
    check(bool(leaders), "at least one measure has a single leader (else the dot is vacuous)")
    painted = _strip_dot_colors(page)
    check(len(painted) == len(leaders),
          f"one best-value dot per measure that has a single leader "
          f"({len(painted)} dots, {len(leaders)} leaders)")
    from lib import palette as PAL
    wanted = {_rgb(PAL.institution_color(slots[iid])) for iid in set(leaders.values())}
    check(wanted <= set(painted),
          f"every dot is painted in its own institution's colour ({sorted(wanted)} in "
          f"{sorted(set(painted))})")

    # 2B-R2-9 the LINK and the DELETION
    hrefs = page.evaluate(
        "Array.from(document.querySelectorAll('.st-key-compare_strip a')).map(a => a.href)")
    for iid, name in names.items():
        want = works_link_named(iid, name)
        check(any(h.startswith(want.split("#")[0]) for h in hrefs),
              f"the institution name of {iid} links to its own publications in OpenAlex")
    check(copy.COMPARE["STRIP_LINK_PUBS"] not in
          page.evaluate("Array.from(document.querySelectorAll('.st-key-compare_strip button'))"
                        ".map(b => b.textContent).join('|')"),
          "the separate publications button is gone from the cards")

    tips = _tooltips(page)
    check(_first_literal(copy.COMPARE["CARD_WINDOW_TIP"]) in tips,
          "the window sentence is now inside the cards' own tooltips")
    check(_first_literal(copy.COMPARE["OVERVIEW_NOTE_TIP"]) in tips,
          "the cards' reading line carries its method in a tooltip")
    check(views_compare._ci_sentence() not in _text_of(page, ".st-key-compare_strip"),
          "the bootstrap-interval line has left the cards")


def _probe_cap(page) -> None:
    warned = _warnings(page)
    check(_first_literal(copy.COMPARE["CAP_TRUNCATED"]) in warned,
          "the over-cap deep link renders the truncation reason")
    n_over = len(LINK_IDS) - state.COMPARE_CAP
    check(copy.COMPARE["CAP_TRUNCATED"].format(cap=state.COMPARE_CAP, n=n_over) in warned,
          f"the reason names the cap and how many it left out ({n_over})")
    code = page.evaluate(
        "Array.from(document.querySelectorAll('[data-testid=\"stCode\"]'))"
        ".map(e => e.textContent).join('|')")
    check("?compare=" + ",".join(SHOWN_IDS) in code,
          "the page prints the deep link for the comparison it actually drew")
    check("?pair=" in code, "the page prints a Collaborate deep link for the chosen pair")


def _probe_legends_and_figures(page) -> None:
    n_figs = _n_figures(page)
    check(n_figs >= MIN_FIGURES, f"the page draws its views ({n_figs} figures, floor "
                                 f"{MIN_FIGURES})")
    legends = _n_legends(page, list(_names(SHOWN_IDS).values()))
    check(legends >= n_figs,
          f"a legend strip sits above every chart ({legends} strips, {n_figs} figures)")
    check(copy.COMPARE["LEGEND_SHARED"] in _body_text(page),
          "the frontier map's legend carries the shared chip")


def _probe_interval_coverage(page) -> None:
    sentence = views_compare._ci_sentence()
    tips = _tooltips(page)
    check(sentence in tips,
          "the exact interval coverage is stated on the page (Methods' own sentence)")
    check(tips.count(sentence) >= 2,
          "the coverage is stated on BOTH impact panels")


def _probe_presentation(page) -> None:
    """2B-R2-8: no grey wall above or below a chart. Every section's reading
    line is one short sentence and its method sits behind the `?`."""
    n_notes = page.evaluate(
        "document.querySelectorAll('[role=\"note\"][title]').length")
    check(n_notes >= MIN_FIGURES,
          f"every chart carries a reading line with a `?` ({n_notes} notes)")
    longest = page.evaluate(
        "Math.max(...Array.from(document.querySelectorAll('[role=\"note\"][title]'))"
        ".map(e => e.parentElement.textContent.trim().length))")
    check(longest <= charts_compare.NOTE_MAX_CHARS + 2,
          f"the longest reading line is one line ({longest} chars, cap "
          f"{charts_compare.NOTE_MAX_CHARS})")
    tips = _tooltips(page)
    check(_first_literal(copy.COMPARE["TIP_GUTTER"]) in tips,
          "the gutter's meaning is stated, in a tooltip")


def _probe_not_offered(page) -> None:
    """2B-R2-3/8/13: the hidden options are disclosed in the SHARED wording,
    with the frame's own plain-language reason, on every level that hides one."""
    caps = _captions(page)
    check(copy.SHARED["NOT_OFFERED_HEADER"] in caps,
          "the page uses the shared 'not shown here, and why' header")
    for level, key in (("field", "cmp_metric_subject"), ("erc", "cmp_metric_erc"),
                       ("sdg", "cmp_metric_sdg")):
        hidden = [m for m in views_compare.SUBJECT_METRICS
                  if not compare_data.metric_frame_available(m, level)]
        lines = [views_compare._not_offered_line(
            views_compare.METRIC_LABELS[m],
            compare_data.UNAVAILABLE_REASON[(m, level)]) for m in hidden]
        check(bool(hidden) and all(_dashless(line) in _dashless(caps) for line in lines),
              f"the {level} section discloses each measure it does not offer "
              f"({len(hidden)} of them)")
    check(views_compare.METRIC_LABELS["vol_top10"] not in
          "|".join(_radio_options(page, "cmp_metric_subject")
                   + _radio_options(page, "cmp_metric_erc")
                   + _radio_options(page, "cmp_metric_sdg")),
          "the top-decile volume is not offered as a tab anywhere (2B-R2-3)")
    check("Volume: Volume" not in caps,
          "a reason that names its own measure is not printed twice")


def _probe_metric_sweep(page) -> None:
    """THE 2B-R LESSON: every option the selector offers is CLICKED and its
    figure redrawn. An option whose render path is never reached is exactly the
    bug that survived the last round."""
    for key in ("cmp_metric_subject", "cmp_metric_erc", "cmp_metric_sdg"):
        options = _radio_options(page, key)
        expected = [views_compare.METRIC_LABELS[m] for m in views_compare.SUBJECT_METRICS
                    if compare_data.metric_frame_available(
                        m, {"cmp_metric_subject": "field", "cmp_metric_erc": "erc",
                            "cmp_metric_sdg": "sdg"}[key])]
        check(options == expected, f"{key} offers exactly what the data can serve "
                                   f"({options})")
        for label in options:
            _click_option(page, key, label)
            check(_n_figures(page) >= MIN_FIGURES,
                  f"{key} = {label!r} renders every view ({_n_figures(page)} figures)")
    # ... and back to the measure the rest of the probe reads
    _click_option(page, "cmp_metric_subject", views_compare.METRIC_LABELS["share"])


def _probe_row_order(page) -> None:
    """2B-R2-5: the DRAWN row order is identical between two metric tabs. Read
    off the rendered tick labels, not off the frame."""
    _click_option(page, "cmp_metric_subject", views_compare.METRIC_LABELS["share"])
    first = _row_names(_tick_labels(page, SUBJECT_FIG))
    _click_option(page, "cmp_metric_subject", views_compare.METRIC_LABELS["dynamics"])
    second = _row_names(_tick_labels(page, SUBJECT_FIG))
    check(len(first) > 5 and len(second) > 5,
          f"both metric tabs draw a full set of rows ({len(first)}, {len(second)})")
    common_a = [t for t in first if t in set(second)]
    common_b = [t for t in second if t in set(first)]
    check(len(common_a) > 5 and common_a == common_b,
          f"the row order is IDENTICAL between two metric tabs ({len(common_a)} common rows)")
    # the toggle really does re-rank
    _click_option(page, "cmp_sort_subject", views_compare.SORT_LABELS["value"])
    ranked = _row_names(_tick_labels(page, SUBJECT_FIG))
    check(set(ranked) == set(second) and ranked != second,
          "the sort-by-value toggle re-ranks the same rows")
    _click_option(page, "cmp_sort_subject", views_compare.SORT_LABELS["taxonomy"])
    back = _row_names(_tick_labels(page, SUBJECT_FIG))
    check(back == second, "switching back restores the taxonomy order")


def _probe_gutters_and_refs(page) -> None:
    """2B-R2-3/4: the gutter carries a raw volume on EVERY metric, the dynamics
    gutter carries its raw-delta string, and pp/sdg_share/dynamics draw the
    reference the frame supplies."""
    _click_option(page, "cmp_metric_subject", views_compare.METRIC_LABELS["share"])
    labels = _tick_labels(page, SUBJECT_FIG)
    with_numbers = [t for t in labels if re.search(r"\d", str(t))]
    check(len(with_numbers) >= len(labels) - 1,
          f"the volume gutter carries numbers on every row ({len(with_numbers)} of "
          f"{len(labels)})")
    n_refs_share = _n_shapes(page, SUBJECT_FIG, ".shapelayer path")

    _click_option(page, "cmp_metric_subject", views_compare.METRIC_LABELS["dynamics"])
    dyn_labels = _tick_labels(page, SUBJECT_FIG)
    arrow = compare_data.DYNAMICS_ARROW
    check(any(arrow in str(t) for t in dyn_labels),
          "the dynamics gutter prints the raw change, from one mean to the other")
    check(any("/" in str(t) for t in dyn_labels),
          "the dynamics gutter names its unit (a year)")
    n_refs_dyn = _n_shapes(page, SUBJECT_FIG, ".shapelayer path")
    check(n_refs_dyn > n_refs_share,
          f"the dynamics view draws reference marks the share view does not "
          f"({n_refs_dyn} vs {n_refs_share} shapes)")
    check(_dashless(compare_data.DYNAMICS_DENOM_NOTE) in _dashless(_tooltips(page)),
          "the dynamics view names both windows, verbatim from the frame")
    check(_first_literal(copy.COMPARE["TIP_REFERENCE"]) in _tooltips(page),
          "the two baselines are named where the reference is drawn")

    _click_option(page, "cmp_metric_subject", views_compare.METRIC_LABELS["pp"])
    n_refs_pp = _n_shapes(page, SUBJECT_FIG, ".shapelayer path")
    check(n_refs_pp > n_refs_share,
          f"the top-decile view draws its index reference ({n_refs_pp} shapes)")
    check(_first_literal(copy.COMPARE["TIP_LOW_VOLUME"]) in _tooltips(page),
          "the low-volume marker is explained where it can fire")
    _click_option(page, "cmp_metric_subject", views_compare.METRIC_LABELS["share"])


def _probe_drill(page) -> None:
    fields = views_compare._fields(tuple(SHOWN_IDS), TREE, BASIS)
    field_name = str(fields["field_name"].iloc[0])
    _pick_selectbox(page, "cmp_field_drill", field_name)
    check(copy.COMPARE["CAPTION_DRILL"].format(field=field_name) in _markdown(page),
          f"the drill renders the subfields of one field ({field_name})")
    for label in _radio_options(page, "cmp_metric_subject"):
        _click_option(page, "cmp_metric_subject", label)
        check(_n_figures(page) >= MIN_FIGURES,
              f"subfield grain, {label!r} renders every view")
    hidden_sub = [m for m in views_compare.SUBJECT_METRICS
                  if not compare_data.metric_frame_available(m, "subfield")]
    line = views_compare._not_offered_line(
        views_compare.METRIC_LABELS[hidden_sub[0]],
        compare_data.UNAVAILABLE_REASON[(hidden_sub[0], "subfield")])
    check(_dashless(line) in _dashless(_captions(page)),
          "the drill discloses the measures the subfield grain cannot serve")
    _pick_selectbox(page, "cmp_field_drill", copy.COMPARE["DRILL_ALL"])
    check(copy.COMPARE["CAPTION_DRILL"].format(field=field_name) not in _markdown(page),
          "leaving the drill returns the subject section to all fields")


def _shared_count_note(top_n: int, pool: str) -> str:
    pooled = views_compare._frontier_pooled(tuple(SHOWN_IDS), TREE, BASIS, top_n, pool)
    n_shared = int((pooled["owner"] == charts_compare.SHARED_OWNER).sum())
    return copy.COMPARE["NOTE_FRONTIER_MAP"].format(
        n_shared=f"{n_shared:,}", n_shown=f"{len(pooled):,}")


def _probe_frontier(page) -> None:
    """2B-R-9 + 2B-R2-10: the slider does real work, the reading line counts the
    shared topics FROM THE DATA, and both new selectors change what is drawn."""
    check(_shared_count_note(views_compare.FRONTIER_TOPN_DEFAULT, "volume") in _markdown(page),
          "the frontier note states the shared-topic count of the plotted cut")
    landed = _set_slider(page, "cmp_frontier_topn", views_compare.FRONTIER_TOPN_MIN)
    check(landed == views_compare.FRONTIER_TOPN_MIN,
          f"the top-N slider moves to its minimum ({landed})")
    check(_shared_count_note(views_compare.FRONTIER_TOPN_MIN, "volume") in _markdown(page),
          "moving the slider re-cuts the pooled map and the note follows it")
    shared = views_compare._shared_long(
        views_compare._shared_frontier(tuple(SHOWN_IDS), TREE, BASIS), SHOWN_IDS)
    check(copy.COMPARE["NOTE_SHARED_FRONTIER"].format(
        n=f"{shared['topic_id'].nunique():,}") in _markdown(page),
        "the shared-frontier list states how many topics more than one of them holds")
    check(copy.COMPARE["FRONTIER_POOL_RULE_VOLUME"][:40] in _tooltips(page),
          "the default pool rule is stated in plain words")

    # the POOL selector
    n_before = _n_shapes(page, "fig_cmp_frontier_map", ".scatterlayer path")
    _click_option(page, "cmp_frontier_pool", views_compare.POOL_LABELS["elite"])
    check(copy.COMPARE["FRONTIER_POOL_RULE_ELITE"][:40] in _tooltips(page),
          "picking the narrower pool restates the pool rule")
    check(_shared_count_note(views_compare.FRONTIER_TOPN_MIN, "elite") in _markdown(page),
          "the narrower pool's own shared count is drawn")
    n_after = _n_shapes(page, "fig_cmp_frontier_map", ".scatterlayer path")
    check(n_after > 0 and n_before > 0,
          f"both pools draw bubbles ({n_before} then {n_after})")
    _click_option(page, "cmp_frontier_pool", views_compare.POOL_LABELS["volume"])

    # the COLOUR-BY toggle
    pooled = views_compare._frontier_pooled(tuple(SHOWN_IDS), TREE, BASIS,
                                            views_compare.FRONTIER_TOPN_MIN, "volume")
    items = views_compare._domain_items(pooled)
    check(bool(items), "the plotted topics carry the domains the colour toggle needs")
    _click_option(page, "cmp_frontier_color", views_compare.COLOR_BY_LABELS["domain"])
    text = _markdown(page)
    check(all(label in text for _d, label in items),
          f"the legend is rebuilt with the broad subject areas ({[l for _d, l in items]})")
    check(_n_figures(page) >= MIN_FIGURES, "the map redraws under the colour toggle")
    _click_option(page, "cmp_frontier_color", views_compare.COLOR_BY_LABELS["owner"])
    check(copy.COMPARE["LEGEND_SHARED"] in _body_text(page),
          "switching back restores the ownership legend")
    _set_slider(page, "cmp_frontier_topn", views_compare.FRONTIER_TOPN_DEFAULT)


def _impact_note(page) -> str:
    segments = [t for t in re.split(r"\{[^{}]*\}", copy.COMPARE["NOTE_IMPACT_SUBFIELDS"])
                if t.strip()]
    marker = max(segments, key=len).strip()
    for text in _markdown(page).split("|"):
        if marker in text:
            return text.strip()
    return ""


def _probe_impact_floor(page) -> None:
    before = _impact_note(page)
    low = min(views_compare.IMPACT_FLOORS)
    _click_option(page, "cmp_impact_floor",
                  copy.COMPARE["IMPACT_FLOOR_OPTION"].format(floor=low))
    after = _impact_note(page)
    check(bool(before) and bool(after) and before != after,
          f"the floor toggle changes the impact reading line ({before!r} -> {after!r})")


def _probe_no_snapshot(page) -> None:
    check("napshot" not in _body_text(page),
          "the snapshot string is gone from this page (2B-R-12)")
    check("napshot" not in _tooltips(page),
          "and it did not survive inside a tooltip")


def _probe_page(page) -> None:
    _load(page)
    _probe_overview(page)
    _probe_cap(page)
    _probe_legends_and_figures(page)
    _probe_interval_coverage(page)
    _probe_presentation(page)
    _probe_no_snapshot(page)
    _probe_not_offered(page)
    _probe_metric_sweep(page)
    _probe_row_order(page)
    _probe_gutters_and_refs(page)
    _probe_drill(page)
    _probe_frontier(page)
    _probe_impact_floor(page)


# ------------------------------------------------------------- the workbook --

def _probe_download(page) -> None:
    """The workbook, taken through the page's real download button, opened with
    openpyxl: one sheet per view plus the re-cut Methods sheet."""
    _load(page)
    with page.expect_download(timeout=120_000) as info:
        page.locator(".st-key-dl_workbook button").first.click()
    path = Path(info.value.path())
    raw = path.read_bytes()
    check(raw[:2] == b"PK", "the workbook downloads as a real xlsx container")
    book = openpyxl.load_workbook(io.BytesIO(raw))
    check(book.sheetnames[0] == copy.COMPARE["XLSX_SHEET_METHODS"],
          f"the first sheet is the Methods sheet ({book.sheetnames[:1]})")
    expected = len(views_compare.SLUGS) + 1
    check(len(book.sheetnames) == expected,
          f"the workbook carries a sheet per view plus Methods ({len(book.sheetnames)} of "
          f"{expected})")
    wanted = {copy.COMPARE["XLSX_SHEET_OVERVIEW"], copy.COMPARE["XLSX_SHEET_SUBJECT_FIELD"],
              copy.COMPARE["VIEW_ERC"], copy.COMPARE["VIEW_SDG"],
              copy.COMPARE["VIEW_FRONTIER_MAP"], copy.COMPARE["VIEW_SHARED_FRONTIER"],
              copy.COMPARE["XLSX_SHEET_IMPACT_INDEX"],
              copy.COMPARE["XLSX_SHEET_IMPACT_SUBFIELDS"],
              copy.COMPARE["VIEW_TRENDS"], copy.COMPARE["VIEW_COVERAGE"]}
    check(wanted <= set(book.sheetnames),
          f"the sheet names are the page's own view names ({sorted(wanted - set(book.sheetnames))})")
    values = [str(c.value) for row in book[book.sheetnames[0]].iter_rows()
              for c in row if c.value is not None]
    check(copy.VERDICT_LINE in values, "the Methods sheet carries the standing reading line")
    check(copy.COMPARE["XLSX_ROW_SNAPSHOT"] not in values,
          "the Methods sheet has no snapshot row (2B-R-12)")
    joined = " ".join(values)
    both = all(views_compare._window(w) in joined
               for w in (compare_data.DYNAMICS_W1, compare_data.DYNAMICS_W2))
    check(both, "the Methods sheet names BOTH dynamics windows (2B-R-6)")
    check(copy.COMPARE["XLSX_ROW_CAP"] in values and copy.COMPARE["XLSX_ROW_FLOORS"] in values,
          "the Methods sheet states the comparison cap and the floors in force")
    check(views_compare._ci_sentence() in values,
          "the Methods sheet states the interval coverage")
    for key in ("XLSX_ROW_POOL", "XLSX_ROW_COLOUR", "XLSX_ROW_SORT"):
        check(copy.COMPARE[key] in values,
              f"the Methods sheet records the control the reader was on ({key})")
    check(views_compare.POOL_LABELS[views_compare.POOL_DEFAULT] in values
          and views_compare.SORT_LABELS[views_compare.SORT_DEFAULT] in values,
          "and it records their VALUES, in the page's own words")


# ----------------------------------------------------------------- widths --

def _probe_widths(browser) -> None:
    SHOT_DIR.mkdir(parents=True, exist_ok=True)
    for width in WIDTHS:
        height = TALL_SHOT_PX if width >= 1280 else SHOT_HEIGHT_PX
        page = browser.new_page(viewport={"width": width, "height": height})
        _load(page)
        scroll = page.evaluate("document.documentElement.scrollWidth")
        inner = page.evaluate("window.innerWidth")
        check(scroll <= inner + 2,
              f"{width} px: scrollWidth {scroll} <= innerWidth+2 {inner + 2}")
        path = SHOT_DIR / f"cp3_compare_{width}.png"
        page.screenshot(path=str(path), full_page=True)
        print("Saved screenshot:", path)
        check(path.is_file(), f"{width} px: screenshot written")
        page.close()
    page = browser.new_page(viewport={"width": 1280, "height": HEAD_SHOT_PX})
    _load(page)
    top = SHOT_DIR / "cp3_compare_top_1280.png"
    page.screenshot(path=str(top), full_page=False)
    print("Saved screenshot:", top)
    check(top.is_file(), "1280 px: the page head, above the fold")
    page.close()


def main() -> int:
    global PORT
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--port", type=int, default=DEFAULT_PORT,
                        help="port to run the Streamlit server on")
    PORT = parser.parse_args().port

    server = subprocess.Popen(
        [sys.executable, "-m", "streamlit", "run", PAGE,
         "--server.headless", "true", "--server.port", str(PORT)],
        cwd=str(APP_DIR), stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)
    try:
        if not _wait_for_port(PORT):
            print("FAIL: server did not open port", PORT)
            return 1
        with sync_playwright() as p:
            browser = p.chromium.launch()
            context = browser.new_context(viewport={"width": 1280, "height": 1000},
                                          accept_downloads=True)
            page = context.new_page()
            _probe_page(page)
            _probe_download(page)
            page.close()
            context.close()
            _probe_widths(browser)
            browser.close()
    finally:
        server.terminate()
        try:
            server.wait(timeout=10)
        except subprocess.TimeoutExpired:
            server.kill()
            server.wait(timeout=10)
    failed = [m for ok, m in RESULTS if not ok]
    print(f"\n{len(RESULTS) - len(failed)} of {len(RESULTS)} checks passed")
    if failed:
        for m in failed:
            print("FAILED:", m)
        return 1
    print("ALL CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
