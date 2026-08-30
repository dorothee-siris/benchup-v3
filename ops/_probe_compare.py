"""
Acceptance probe for the REBUILT Compare page (BUILD_PLAN_2BR.md stream CP,
decisions 2B-R-4/5/6/7/8/9/12; VIZ_SPEC S2 quater 4.1 ... 4.7). Same shape as
ops/_probe_collab.py: start `streamlit run pages/2_<scales>_Compare.py` as a
subprocess, drive it headless with Playwright, ALWAYS terminate the server.

Selectors are locale-independent -- the `st-key-<key>` classes the page's own
keyed widgets and containers emit, plus `[data-testid=...]`. Nothing is asserted
against a canvas: the workbook is checked by DOWNLOADING it and opening it with
openpyxl, the figures are counted as plotly roots, and every VALUE the page
renders is checked by recomputing it here from `lib/compare_data.py` and looking
for the page's own formatting of it in the DOM. A probe that only counted
elements would pass a page whose cards had drifted from the frame.

The comparison is injected through the real `?compare=A,B,C,D` query parameter
the shipped deep link produces (`lib.selection.deeplink`) with FOUR ids, one
more than 2B-R-4's hard cap of three: the cap, and the reason it renders, are
therefore probed on the live URL path, not only through AppTest.

Usage:  python ops/_probe_compare.py [--port 8605]
Exit 0 when every check passes; 1 otherwise. Stdout is ASCII-only (cp1252
console).
"""
from __future__ import annotations

import argparse
import io
import re
import socket
import subprocess
import sys
import time
from pathlib import Path

import openpyxl
from playwright.sync_api import sync_playwright

APP_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(APP_DIR))

from lib import charts_compare, compare_data, copy, state, views_compare  # noqa: E402

PAGE = "pages/2_⚖️_Compare.py"
DEFAULT_PORT = 8605

STRASBOURG = "I68947357"
SORBONNE = "I39804081"
FREIBURG = "I161046081"
GDANSK = "I40413290"          # the fourth id: over the cap ON PURPOSE
LINK_IDS = [STRASBOURG, SORBONNE, FREIBURG, GDANSK]
SHOWN_IDS = LINK_IDS[:state.COMPARE_CAP]

TREE, BASIS = "bestfit", "frac"

SHOT_DIR = APP_DIR / "tests" / "ui" / "screenshots"
WIDTHS = [1920, 1280, 390]
SHOT_HEIGHT_PX = 2400   # full_page=True is a no-op on Streamlit's scroll container, so the
TALL_SHOT_PX = 5600     # VIEWPORT is the screenshot; the wide widths get a tall one.
HEAD_SHOT_PX = 900      # the true above-the-fold view at 1280 px
MIN_FIGURES = 8         # subject, ERC, SDG, frontier map, shared frontier, impact,
                        # impact by subfield, trends, coverage -- the shared-frontier
                        # chart is the one that can legitimately be absent (an empty
                        # intersection), so the floor is one below the nine drawn.

PORT = DEFAULT_PORT
RESULTS: list[tuple[bool, str]] = []


def check(ok: bool, message: str) -> bool:
    RESULTS.append((bool(ok), message))
    print(("PASS: " if ok else "FAIL: ") + message)
    return bool(ok)


def _wait_for_port(port: int, timeout: float = 60.0) -> bool:
    deadline = time.time() + timeout
    while time.time() < deadline:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            if sock.connect_ex(("127.0.0.1", port)) == 0:
                return True
        time.sleep(0.5)
    return False


def _load(page) -> None:
    page.goto(f"http://127.0.0.1:{PORT}/?compare=" + ",".join(LINK_IDS),
              wait_until="domcontentloaded")
    page.wait_for_selector(".js-plotly-plot", timeout=240_000)
    _settle(page)


def _settle(page, target: int = MIN_FIGURES, timeout_ms: int = 180_000) -> int:
    """Streamlit streams its elements, so the figure count climbs for a while
    after the first plot appears. Wait for it to reach the acceptance floor and
    then stop changing, rather than sleeping a guessed number of seconds."""
    deadline = time.time() + timeout_ms / 1000.0
    last, stable = -1, 0
    while time.time() < deadline:
        now = _n_figures(page)
        stable = stable + 1 if now == last and now >= target else 0
        last = now
        if stable >= 3:
            break
        page.wait_for_timeout(1000)
    page.wait_for_timeout(1500)
    return last


def _n_figures(page) -> int:
    return page.locator(".js-plotly-plot").count()


def _text_of(page, selector: str) -> str:
    return page.evaluate(
        "(sel) => { const e = document.querySelector(sel); return e ? e.textContent : ''; }",
        selector)


def _body_text(page) -> str:
    return page.evaluate("document.body.innerText")


def _captions(page) -> str:
    return page.evaluate(
        "Array.from(document.querySelectorAll('[data-testid=\"stCaptionContainer\"]'))"
        ".map(e => e.textContent).join('|')")


def _warnings(page) -> str:
    return page.evaluate(
        "Array.from(document.querySelectorAll('[data-testid=\"stAlertContainer\"]'))"
        ".map(e => e.textContent).join('|')")


def _n_legends(page, names) -> int:
    """Legend strips: a markdown container whose text names every compared
    institution and which carries the chip spans. `charts.chip_legend_html`
    writes inline styles the browser normalises, so the count is taken on
    STRUCTURE and TEXT, never on an attribute substring."""
    return page.evaluate(
        "(names) => Array.from(document.querySelectorAll('[data-testid=\"stMarkdownContainer\"]'))"
        ".filter(e => names.every(n => e.textContent.includes(n))"
        " && e.querySelectorAll('span').length >= names.length * 2).length",
        list(names))


def _click_option(page, container_key: str, text: str) -> None:
    """Click one option of a keyed radio, then wait for the rerun to finish
    redrawing (a rerun rebuilds every figure on the page, so the count dips
    before it comes back)."""
    page.locator(f".st-key-{container_key}").get_by_text(text, exact=True).first.click()
    page.wait_for_timeout(2000)
    _settle(page)


def _pick_selectbox(page, container_key: str, text: str) -> None:
    """Open a keyed selectbox and choose an option by typing it: Streamlit's
    combobox filters as you type and Enter takes the first match, which needs
    no assumption about where the portal renders the option list."""
    box = page.locator(f".st-key-{container_key} input").first
    box.click()
    box.fill(text)
    page.wait_for_timeout(800)
    box.press("Enter")
    page.wait_for_timeout(2000)
    _settle(page)


def _first_literal(template: str) -> str:
    segments = [s for s in re.split(r"\{[^{}]*\}", template) if s.strip()]
    return segments[0].strip()


DASH_CHARS = "-\N{EN DASH}\N{EM DASH}\N{MINUS SIGN}"


def _dashless(text: str) -> str:
    """Streamlit renders every caption through markdown, which rewrites a
    double hyphen into a typographic dash. A probe that compared the source
    string byte-for-byte against the DOM would fail on the one caption that
    carries a parenthetical -- `compare_data.DYNAMICS_DENOM_NOTE`, the dynamics
    note 2B-R-6 requires verbatim -- for a typographic reason and nothing else.
    Dropping every dash from BOTH sides keeps the check about the words and the
    numbers, which is what the decision is about."""
    return "".join(c for c in text if c not in DASH_CHARS)


def _set_slider(page, container_key: str, target: int) -> int:
    """Streamlit's slider is an `input[type=range]` driven by react-aria: it
    takes keyboard steps, not a click at a coordinate (a click on the track is
    what a human does, but its landing value depends on pixel geometry, which
    is exactly the kind of thing a probe must not assert on). Arrow keys walk
    it by its own step until the value is the one asked for."""
    knob = page.locator(f".st-key-{container_key} input[type=range]").first
    knob.focus()
    for _ in range(20):
        now = int(knob.input_value())
        if now == target:
            break
        knob.press("ArrowLeft" if now > target else "ArrowRight")
        page.wait_for_timeout(300)
    page.wait_for_timeout(2000)
    _settle(page)
    return int(knob.input_value())


def _ctx():
    return views_compare._bundle()["ctx"]


def _names(ids) -> dict:
    idx = _ctx()["index_by_id"]
    return {i: str(idx.loc[i, "display_name"]) for i in ids}


# --------------------------------------------------------------- the page --

def _probe_overview(page) -> None:
    """VIZ_SPEC 4.1 + 2B-R-7: the cards ARE `compare_data.overview`, so every
    rendered figure is recomputed here and looked for in the strip."""
    names = _names(SHOWN_IDS)
    strip = _text_of(page, ".st-key-compare_strip")
    for iid, name in names.items():
        check(name in strip, f"the overview names {iid}")
    check(_names([GDANSK])[GDANSK] not in strip,
          "the fourth institution of the link is NOT compared (cap of three)")

    frame = compare_data.overview(_ctx(), SHOWN_IDS).set_index("institution_id")
    idx = _ctx()["index_by_id"]
    for iid in SHOWN_IDS:
        facts = views_compare._card_facts(idx.loc[iid], frame.loc[iid])
        missing = [(label, value) for label, value, _h, _s in facts if value not in strip]
        check(not missing, f"every overview value of {iid} reads back from compare_data "
                           f"({len(facts)} values; missing {missing})")
    intervals = [views_compare._interval(frame.loc[iid, "ci_low"], frame.loc[iid, "ci_high"])
                 for iid in SHOWN_IDS]
    check(all(i in strip for i in intervals),
          "each card carries its own bootstrap interval beside the point estimate")


def _probe_cap(page) -> None:
    warned = _warnings(page)
    check(_first_literal(copy.COMPARE["CAP_TRUNCATED"]) in warned,
          "the over-cap deep link renders the truncation reason")
    n_over = len(LINK_IDS) - state.COMPARE_CAP
    check(copy.COMPARE["CAP_TRUNCATED"].format(cap=state.COMPARE_CAP, n=n_over) in warned,
          f"the reason names the cap and how many it left out ({n_over})")
    code = page.evaluate(
        "Array.from(document.querySelectorAll('[data-testid=\"stCode\"]'))"
        ".map(e => e.textContent).join('|')")
    check("?compare=" + ",".join(SHOWN_IDS) in code,
          "the page prints the deep link for the comparison it actually drew")
    check("?pair=" in code, "the page prints a Collaborate deep link for the chosen pair")


def _probe_legends_and_figures(page) -> None:
    n_figs = _n_figures(page)
    check(n_figs >= MIN_FIGURES, f"the page draws its views ({n_figs} figures, floor "
                                 f"{MIN_FIGURES})")
    legends = _n_legends(page, list(_names(SHOWN_IDS).values()))
    check(legends >= n_figs,
          f"a legend strip sits above every chart ({legends} strips, {n_figs} figures)")
    check(copy.COMPARE["LEGEND_SHARED"] in _body_text(page),
          "the frontier map's legend carries the shared chip")


def _probe_interval_coverage(page) -> None:
    sentence = views_compare._ci_sentence()
    check(sentence in _captions(page),
          "the exact interval coverage is stated on the page (Methods' own sentence)")
    check(_captions(page).count(sentence) >= 2,
          "the coverage is stated beside BOTH the cards and the impact panel")


def _probe_selectors(page) -> None:
    """2B-R-5: one metric selector per section, unavailable options hidden WITH
    the frame's own reason, and a level change that retires the current metric
    clamps instead of raising."""
    hidden_erc = [m for m in views_compare.ERC_METRICS
                  if not compare_data.metric_frame_available(m, "erc")]
    lines = [copy.COMPARE["METRIC_HIDDEN_LINE"].format(
        metric=views_compare.METRIC_LABELS[m],
        reason=compare_data.UNAVAILABLE_REASON[(m, "erc")]) for m in hidden_erc]
    caps = _captions(page)
    check(bool(hidden_erc) and all(line in caps for line in lines),
          f"the ERC section discloses the frame's reason for each hidden metric "
          f"({len(hidden_erc)} of them)")

    # switch the subject metric: same chart count, and SI brings its own caption
    before = _n_figures(page)
    _click_option(page, "cmp_metric_subject", views_compare.METRIC_LABELS["si"])
    check(_n_figures(page) == before, "switching the subject metric redraws the same views")
    check(_first_literal(copy.FIND["CAPTION_SI"]) in _captions(page),
          "the specialisation metric brings the specialisation caption with it")

    # ... and the dynamics metric states BOTH windows, from the frame itself
    _click_option(page, "cmp_metric_subject", views_compare.METRIC_LABELS["dynamics"])
    caps = _captions(page)
    check(_dashless(compare_data.DYNAMICS_DENOM_NOTE) in _dashless(caps),
          "the dynamics view names both windows, verbatim from the frame")
    check(views_compare._window(compare_data.DYNAMICS_W1) not in
          views_compare._window(compare_data.DYNAMICS_W2),
          "the two dynamics windows are distinct")

    # drill into one field: the level changes and the retired metric is clamped
    fields = views_compare._fields(tuple(SHOWN_IDS), TREE, BASIS)
    field_name = str(fields["field_name"].iloc[0])
    _pick_selectbox(page, "cmp_field_drill", field_name)
    check(copy.COMPARE["CAPTION_DRILL"].format(field=field_name) in _captions(page),
          f"the drill renders the subfields of one field ({field_name})")
    hidden_sub = [m for m in views_compare.SUBJECT_METRICS
                  if not compare_data.metric_frame_available(m, "subfield")]
    line = copy.COMPARE["METRIC_HIDDEN_LINE"].format(
        metric=views_compare.METRIC_LABELS[hidden_sub[0]],
        reason=compare_data.UNAVAILABLE_REASON[(hidden_sub[0], "subfield")])
    check(line in _captions(page),
          "the drill discloses the metrics the subfield grain cannot serve")
    check(_n_figures(page) >= MIN_FIGURES,
          "the page still draws every view after the drill")
    _pick_selectbox(page, "cmp_field_drill", copy.COMPARE["DRILL_ALL"])
    check(copy.COMPARE["CAPTION_DRILL"].format(field=field_name) not in _captions(page),
          "leaving the drill returns the subject section to all fields")


def _shared_count_caption(top_n: int) -> str:
    pooled = views_compare._frontier_pooled(tuple(SHOWN_IDS), TREE, BASIS, top_n)
    n_shared = int((pooled["owner"] == charts_compare.SHARED_OWNER).sum())
    return copy.COMPARE["CAPTION_FRONTIER_SHARED_COUNT"].format(
        n_shared=f"{n_shared:,}", n_shown=f"{len(pooled):,}")


def _probe_frontier(page) -> None:
    """2B-R-9: the pooled map's slider does real work, and the caption counts
    the shared topics FROM THE DATA (the head of the ranking is near-degenerate,
    which the picture cannot say for itself)."""
    check(_shared_count_caption(views_compare.FRONTIER_TOPN_DEFAULT) in _captions(page),
          "the frontier caption states the shared-topic count of the plotted cut")
    landed = _set_slider(page, "cmp_frontier_topn", views_compare.FRONTIER_TOPN_MIN)
    check(landed == views_compare.FRONTIER_TOPN_MIN,
          f"the top-N slider moves to its minimum ({landed})")
    check(_shared_count_caption(views_compare.FRONTIER_TOPN_MIN) in _captions(page),
          "moving the slider re-cuts the pooled map and the caption follows it")
    shared = views_compare._shared_long(
        views_compare._shared_frontier(tuple(SHOWN_IDS), TREE, BASIS), SHOWN_IDS)
    check(copy.COMPARE["CAPTION_SHARED_TOTAL"].format(
        n=f"{shared['topic_id'].nunique():,}") in _captions(page),
        "the shared-frontier list states how many topics more than one of them holds")
    _set_slider(page, "cmp_frontier_topn", views_compare.FRONTIER_TOPN_DEFAULT)


def _impact_caption(page) -> str:
    segments = [t for t in re.split(r"\{[^{}]*\}", copy.COMPARE["CAPTION_IMPACT_SHOWN"])
                if t.strip()]
    marker = max(segments, key=len).strip()
    for text in _captions(page).split("|"):
        if marker in text:
            return text.strip()
    return ""


def _probe_impact_floor(page) -> None:
    before = _impact_caption(page)
    low = min(views_compare.IMPACT_FLOORS)
    _click_option(page, "cmp_impact_floor",
                  copy.COMPARE["IMPACT_FLOOR_OPTION"].format(floor=low))
    after = _impact_caption(page)
    check(bool(before) and bool(after) and before != after,
          f"the floor toggle changes the impact caption ({before!r} -> {after!r})")


def _probe_no_snapshot(page) -> None:
    check("napshot" not in _body_text(page),
          "the snapshot string is gone from this page (2B-R-12)")


def _probe_page(page) -> None:
    _load(page)
    _probe_overview(page)
    _probe_cap(page)
    _probe_legends_and_figures(page)
    _probe_interval_coverage(page)
    _probe_no_snapshot(page)
    _probe_selectors(page)
    _probe_frontier(page)
    _probe_impact_floor(page)


# ------------------------------------------------------------- the workbook --

def _probe_download(page) -> None:
    """The workbook, taken through the page's real download button, opened with
    openpyxl: one sheet per view plus the re-cut Methods sheet."""
    _load(page)
    with page.expect_download(timeout=120_000) as info:
        page.locator(".st-key-dl_workbook button").first.click()
    path = Path(info.value.path())
    raw = path.read_bytes()
    check(raw[:2] == b"PK", "the workbook downloads as a real xlsx container")
    book = openpyxl.load_workbook(io.BytesIO(raw))
    check(book.sheetnames[0] == copy.COMPARE["XLSX_SHEET_METHODS"],
          f"the first sheet is the Methods sheet ({book.sheetnames[:1]})")
    expected = len(views_compare.SLUGS) + 1
    check(len(book.sheetnames) == expected,
          f"the workbook carries a sheet per view plus Methods ({len(book.sheetnames)} of "
          f"{expected})")
    wanted = {copy.COMPARE["XLSX_SHEET_OVERVIEW"], copy.COMPARE["XLSX_SHEET_SUBJECT_FIELD"],
              copy.COMPARE["VIEW_ERC"], copy.COMPARE["VIEW_SDG"],
              copy.COMPARE["VIEW_FRONTIER_MAP"], copy.COMPARE["VIEW_SHARED_FRONTIER"],
              copy.COMPARE["XLSX_SHEET_IMPACT_INDEX"],
              copy.COMPARE["XLSX_SHEET_IMPACT_SUBFIELDS"],
              copy.COMPARE["VIEW_TRENDS"], copy.COMPARE["VIEW_COVERAGE"]}
    check(wanted <= set(book.sheetnames),
          f"the sheet names are the page's own view names ({sorted(wanted - set(book.sheetnames))})")
    values = [str(c.value) for row in book[book.sheetnames[0]].iter_rows()
              for c in row if c.value is not None]
    check(copy.VERDICT_LINE in values, "the Methods sheet carries the standing reading line")
    check(copy.COMPARE["XLSX_ROW_SNAPSHOT"] not in values,
          "the Methods sheet has no snapshot row (2B-R-12)")
    joined = " ".join(values)
    both = all(views_compare._window(w) in joined
               for w in (compare_data.DYNAMICS_W1, compare_data.DYNAMICS_W2))
    check(both, "the Methods sheet names BOTH dynamics windows (2B-R-6)")
    check(copy.COMPARE["XLSX_ROW_CAP"] in values and copy.COMPARE["XLSX_ROW_FLOORS"] in values,
          "the Methods sheet states the comparison cap and the floors in force")
    check(views_compare._ci_sentence() in values,
          "the Methods sheet states the interval coverage")


# ----------------------------------------------------------------- widths --

def _probe_widths(browser) -> None:
    SHOT_DIR.mkdir(parents=True, exist_ok=True)
    for width in WIDTHS:
        height = TALL_SHOT_PX if width >= 1280 else SHOT_HEIGHT_PX
        page = browser.new_page(viewport={"width": width, "height": height})
        _load(page)
        scroll = page.evaluate("document.documentElement.scrollWidth")
        inner = page.evaluate("window.innerWidth")
        check(scroll <= inner + 2,
              f"{width} px: scrollWidth {scroll} <= innerWidth+2 {inner + 2}")
        path = SHOT_DIR / f"cp_compare_{width}.png"
        page.screenshot(path=str(path), full_page=True)
        print("Saved screenshot:", path)
        check(path.is_file(), f"{width} px: screenshot written")
        page.close()
    page = browser.new_page(viewport={"width": 1280, "height": HEAD_SHOT_PX})
    _load(page)
    top = SHOT_DIR / "cp_compare_top_1280.png"
    page.screenshot(path=str(top), full_page=False)
    print("Saved screenshot:", top)
    check(top.is_file(), "1280 px: the page head, above the fold")
    page.close()


def main() -> int:
    global PORT
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--port", type=int, default=DEFAULT_PORT,
                        help="port to run the Streamlit server on")
    PORT = parser.parse_args().port

    server = subprocess.Popen(
        [sys.executable, "-m", "streamlit", "run", PAGE,
         "--server.headless", "true", "--server.port", str(PORT)],
        cwd=str(APP_DIR), stdout=subprocess.DEVNULL, stderr=subprocess.STDOUT)
    try:
        if not _wait_for_port(PORT):
            print("FAIL: server did not open port", PORT)
            return 1
        with sync_playwright() as p:
            browser = p.chromium.launch()
            context = browser.new_context(viewport={"width": 1280, "height": 1000},
                                          accept_downloads=True)
            page = context.new_page()
            _probe_page(page)
            _probe_download(page)
            page.close()
            context.close()
            _probe_widths(browser)
            browser.close()
    finally:
        server.terminate()
        try:
            server.wait(timeout=10)
        except subprocess.TimeoutExpired:
            server.kill()
            server.wait(timeout=10)
    failed = [m for ok, m in RESULTS if not ok]
    print(f"\n{len(RESULTS) - len(failed)} of {len(RESULTS)} checks passed")
    if failed:
        for m in failed:
            print("FAILED:", m)
        return 1
    print("ALL CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
